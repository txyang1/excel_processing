import os
import time
import json
import pandas as pd
import schedule
from watchdog.observers.polling import PollingObserver as Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import win32com.client  # pip install pywin32

# === 加载配置 ===
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
with open(os.path.join(BASE_DIR, "unified_config_auto2.json"), 'r', encoding='utf-8') as f:
    cfg = json.load(f)

folders         = cfg['folders']
ORIG_DIR        = os.path.join(BASE_DIR, folders['orig_dir'])
JIRA_DIR        = os.path.join(BASE_DIR, folders['jira_dir'])
OCTANE_DIR      = os.path.join(BASE_DIR, folders['octane_dir'])
paths           = cfg['paths']
sheet_name      = cfg['sheet']['target_sheet']
sources         = cfg['sources']
clear_old       = cfg.get('settings', {}).get('clear_old_highlight','N').upper() == 'Y'
fund_patterns   = cfg.get('fund_function_patterns', {})
owner_patterns  = cfg.get('owner_root_cause_patterns', {})

def is_row_blank(ws, row):
    for c in range(1, ws.max_column+1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True

def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break

def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, key_col).value not in (None, ""):
            return r
    return 1

def _refresh_pivot_tables(xlsx_path, sheet_name):
    """
    通过 COM 自动化，将指定 Sheet 上的所有 PivotTable 源范围更新
    到当前数据区域（从 A1 到 最后行+列），并刷新它们。
    """
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
    ws = wb.Worksheets(sheet_name)

    # 找到数据区域的末尾行列
    last_row = ws.Cells(ws.Rows.Count, 1).End(win32com.client.constants.xlUp).Row
    last_col = ws.Cells(1, ws.Columns.Count).End(win32com.client.constants.xlToLeft).Column

    # 构造 SourceData 区域：Sheet!$A$1:$Z$123
    top_left     = ws.Cells(1,1).Address(True, True, 1)               # "$A$1"
    bottom_right = ws.Cells(last_row, last_col).Address(True, True, 1)
    source_data  = f"{sheet_name}!{top_left}:{bottom_right}"

    # 更新每一个 PivotTable
    for pt in ws.PivotTables():
        cache = wb.PivotCaches().Create(
            SourceType=1,     # xlDatabase
            SourceData=source_data
        )
        pt.ChangePivotCache(cache)
        pt.RefreshTable()

    wb.Save()
    wb.Close(False)
    excel.Quit()
    print(f"  → PivotTable 源已更新并刷新: {source_data}")

def update_in_place(new_fp):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] 更新触发: {new_fp}")

    # 判断来源
    folder = os.path.basename(os.path.dirname(new_fp))
    if folder == os.path.basename(JIRA_DIR):
        source_key = "Jira"
    elif folder == os.path.basename(OCTANE_DIR):
        source_key = "Octane"
    else:
        return

    orig_fp = os.path.join(ORIG_DIR, paths['original_file'])
    src_cfg = sources[source_key]
    df_new  = pd.read_excel(new_fp) if src_cfg['read_method']=="excel" else pd.read_csv(new_fp)

    # 用 openpyxl 更新原表
    wb = load_workbook(orig_fp)
    ws = wb[sheet_name]

    # 清除旧高亮
    if clear_old:
        green = {"8ED973","008ED973","FF8ED973"}
        blue  = {'FFADD8E6','00ADD8E6','ADD8E6'}
        gray  = {'FFC0C0C0','00C0C0C0','C0C0C0'}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                pf = cell.fill
                if getattr(pf,'patternType',None)=="solid":
                    argb = pf.fgColor.rgb or pf.fgColor.value
                    if argb in green|blue|gray:
                        cell.fill = PatternFill()

    trim_trailing_blank_rows(ws)

    header2col = { ws.cell(1,c).value: c for c in range(1, ws.max_column+1) }
    id_col     = header2col['ID']
    mapping    = src_cfg['mapping']
    date_col   = src_cfg.get('date_col')

    # 构建已有 ID→行 映射
    id2row = {
        ws.cell(r,id_col).value: r
        for r in range(2, ws.max_row+1)
        if ws.cell(r,id_col).value
    }

    # 如果是 Excel，提取超链接
    id2url = {}
    if src_cfg['read_method']=="excel":
        tmpwb = load_workbook(new_fp)
        tmpws = tmpwb.active
        cid   = next((c for c in range(1, tmpws.max_column+1)
                      if tmpws.cell(1,c).value=="ID"), None)
        if cid:
            for r in range(2, tmpws.max_row+1):
                cell = tmpws.cell(r,cid)
                if cell.hyperlink:
                    id2url[cell.value] = cell.hyperlink.target

    last_row    = find_last_data_row(ws, id_col)
    update_fill = PatternFill('solid', fgColor='ADD8E6')
    new_fill    = PatternFill('solid', fgColor='8ED973')

    id_key = next(k for k,v in mapping.items() if v=="ID")

    # 逐行更新或追加
    for _, new_row in df_new.iterrows():
        nid = new_row.get(id_key)
        if pd.isna(nid) or not nid:
            continue

        if nid in id2row:
            r = id2row[nid]
            # 跳过已关闭
            phase = ws.cell(r, header2col.get('Phase')).value or ""
            if any(s in phase for s in ("Concluded","Closed","Resolved")):
                continue

            # 更新 Creation time
            if date_col:
                raw = new_row.get(date_col)
                if pd.notna(raw):
                    try:
                        dt = pd.to_datetime(raw).to_pydatetime()
                        c  = header2col[mapping[date_col]]
                        cell = ws.cell(r,c)
                        if cell.value != dt:
                            cell.value         = dt
                            cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'
                            cell.fill          = update_fill
                    except:
                        pass

            # 更新其他字段
            for nk, ok in mapping.items():
                if nk==date_col or ok=="Function":
                    continue
                val = new_row.get(nk)
                if pd.isna(val) or val == "":
                    continue
                # Involved I-Step 等自定义逻辑可接入此处…
                c = header2col.get(ok)
                if c and ws.cell(r,c).value != val:
                    ws.cell(r,c).value = val
                    ws.cell(r,c).fill  = update_fill

        else:
            last_row += 1
            for hdr, cidx in header2col.items():
                val = ""
                for nk, ok in mapping.items():
                    if ok == hdr:
                        tmp = new_row.get(nk)
                        if pd.notna(tmp):
                            val = tmp
                        break
                cell = ws.cell(last_row, cidx)
                cell.value = val
                if val:
                    cell.fill = new_fill
                    if hdr == mapping.get(date_col):
                        cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'
            if nid in id2url:
                c = ws.cell(last_row, id_col)
                c.hyperlink, c.style = id2url[nid], 'Hyperlink'

    # 填 Days 公式
    if "Days" in header2col and "Creation time" in header2col:
        colL = get_column_letter(header2col["Creation time"])
        for r in range(2, ws.max_row+1):
            ws.cell(r, header2col["Days"]).value = f'=DATEDIF(${colL}{r},TODAY(),"D")'

    # 保存回原文件
    wb.save(orig_fp)
    print(f"[{ts}] 更新完成：已写回 {orig_fp}")

    # ---- 调用 COM 去更新并刷新 PivotTable ----
    _refresh_pivot_tables(orig_fp, sheet_name)

class FolderHandler(FileSystemEventHandler):
    def on_created(self, event):
        self._check(event.src_path)
    on_modified = on_created
    def on_moved(self, event):
        self._check(event.dest_path)
    def _check(self, path):
        if os.path.isdir(path):
            return
        fld = os.path.basename(os.path.dirname(path))
        if fld in (folders['jira_dir'], folders['octane_dir']):
            update_in_place(path)

def main():
    for d in (ORIG_DIR, JIRA_DIR, OCTANE_DIR):
        os.makedirs(d, exist_ok=True)

    observer = Observer()
    handler  = FolderHandler()
    observer.schedule(handler, path=JIRA_DIR, recursive=False)
    observer.schedule(handler, path=OCTANE_DIR, recursive=False)
    observer.start()
    print(f"监控目录：{JIRA_DIR}, {OCTANE_DIR}")

    # 保留每日 08:00 全量扫描（可选）
    schedule.every().day.at("08:00").do(
        lambda: [
            update_in_place(os.path.join(JIRA_DIR, fn))
            for fn in os.listdir(JIRA_DIR) if fn.lower().endswith((".xlsx",".csv"))
        ] + [
            update_in_place(os.path.join(OCTANE_DIR, fn))
            for fn in os.listdir(OCTANE_DIR) if fn.lower().endswith((".xlsx",".csv"))
        ]
    )

    try:
        while True:
            schedule.run_pending()
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

if __name__ == "__main__":
    main()
