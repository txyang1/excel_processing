import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 定义文件路径（请根据实际情况修改）
original_file = 'original.xlsx'
new_file = 'new.xlsx'
output_file = 'updated_excel.xlsx'

# 1. 读取原始数据和新数据
original_df = pd.read_excel(original_file)
new_df = pd.read_excel(new_file)

# 2. 定义异常数据纠错函数
def error_correction(value, column_name):
    """
    对异常数据进行纠错
    - 如果数据为空或者格式不正确，则返回预设的默认值（可以根据需要扩充）
    """
    # 示例：假设 'Days in phase' 应为数字，如果检测到异常则返回 0
    if column_name == 'Days in phase':
        try:
            corrected = float(value)
        except (ValueError, TypeError):
            corrected = 0
        return corrected
    # 针对其他列，可根据需要添加规则
    # 如果没有特殊要求，直接返回原值
    return value

# 3. 定义映射关系（根据实际情况定义映射字典）
# 例如：当 "Defect finder" 为特定值时，"Phase Group" 应更新为对应值
mapping = {
    "FinderA": "PhaseGroup1",
    "FinderB": "PhaseGroup2",
    # 继续添加映射关系...
}

# 4. 更新数据并记录修改位置（用于后续高亮）
# 使用列表记录已更新的单元格位置，记录格式：(行号, 列名)
updated_cells = []

# 假设每条记录都有唯一标识符 "ID"
for i, new_row in new_df.iterrows():
    # 根据 'ID' 字段匹配原始数据中的行（假设每条记录ID唯一）
    matched = original_df[original_df['ID'] == new_row['ID']]
    if not matched.empty:
        orig_index = matched.index[0]
        # 遍历所有列，对比数据并更新
        for col in new_df.columns:
            new_value = error_correction(new_row[col], col)
            orig_value = original_df.loc[orig_index, col]
            if new_value != orig_value:
                original_df.loc[orig_index, col] = new_value
                updated_cells.append((orig_index, col))
        # 处理映射关系：针对 "Defect finder" 列
        new_defect = error_correction(new_row["Defect finder"], "Defect finder")
        if new_defect in mapping:
            mapped_phase = mapping[new_defect]
            orig_phase = original_df.loc[orig_index, "Phase Group"]
            if mapped_phase != orig_phase:
                original_df.loc[orig_index, "Phase Group"] = mapped_phase
                updated_cells.append((orig_index, "Phase Group"))

# 5. 将更新后的 DataFrame 写入新的 Excel 文件
original_df.to_excel(output_file, index=False)

# 6. 使用 openpyxl 加载写好的文件，并对更新的单元格进行高亮显示
wb = load_workbook(output_file)
ws = wb.active

# 设置高亮填充（这里以黄色为例）
highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 获取表头（第一行）以便确定列的位置
header = [cell.value for cell in ws[1]]

# 对每个更新的单元格进行高亮处理
for row_index, col_name in updated_cells:
    # 注意：pandas 的 index 是从0开始，而 openpyxl 的行号从1开始，并且第一行为表头
    excel_row = row_index + 2  # 加2：1表示从0到1的转变，另1表示跳过表头
    # 确定列号（openpyxl的列是从1开始的）
    if col_name in header:
        col_index = header.index(col_name) + 1
        ws.cell(row=excel_row, column=col_index).fill = highlight_fill

# 保存修改后的文件
wb.save(output_file)

print("数据更新完成，已生成文件：", output_file)
