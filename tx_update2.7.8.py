import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# === 0. 加载配置 ===
with open("config.json", "r", encoding="utf-8") as f:
    cfg = json.load(f)

original_file            = cfg["paths"]["original_file"]
new_file                 = cfg["paths"]["new_file"]
updated_file             = cfg["paths"]["updated_file"]
target_sheet             = cfg["sheet"]["target_sheet"]
mapping                  = cfg["column_mapping"]
fund_function_mapping    = cfg["fund_function_mapping"]
owner_root_cause_mapping = cfg["owner_root_cause_mapping"]

# === 1. 读取数据 ===
df_orig = pd.read_excel(original_file, sheet_name=target_sheet)
df_new  = pd.read_excel(new_file)

# === 2. 打开原表，用于直接读写格式与超链接 ===
wb = load_workbook(original_file)
ws = wb[target_sheet]

# 构建“表头→列号”映射
header2col = {
    ws.cell(row=1, column=c).value: c
    for c in range(1, ws.max_column + 1)
}

# 提取原表中 ID 列号，以及 ID→行号 和 原超链接
id_col = header2col["ID"]
id2row = {}
id2url_orig = {}
for r in range(2, ws.max_row + 1):
    cell = ws.cell(row=r, column=id_col)
    val = cell.value
    if val is not None:
        id2row[val] = r
        if cell.hyperlink:
            id2url_orig[val] = cell.hyperlink.target

# === 3. 提取新表中 ID→超链接 （用普通模式打开才能 access .hyperlink） ===
new_wb = load_workbook(new_file)
new_ws = new_wb.active
id_col_new = next(
    (c for c in range(1, new_ws.max_column+1)
     if new_ws.cell(1,c).value == "ID"),
    None
)
id2url_new = {}
if id_col_new:
    for r in range(2, new_ws.max_row+1):
        c = new_ws.cell(row=r, column=id_col_new)
        if c.hyperlink:
            id2url_new[c.value] = c.hyperlink.target

# 高亮样式
green_fill  = PatternFill("solid", fgColor="00FF00")
yellow_fill = PatternFill("solid", fgColor="FFFF00")

# === 4. 遍历新表，更新原有行或追加新行 ===
for _, new_row in df_new.iterrows():
    new_id = new_row.get("ID", "")
    if new_id in id2row:
        # —— 更新已有行，只更新非空映射字段并染绿 —— 
        r = id2row[new_id]
        # 1) 可选：更新 ID 超链接
        if new_id in id2url_new:
            cell = ws.cell(r, id_col)
            cell.hyperlink = id2url_new[new_id]
            cell.style     = "Hyperlink"
        # 2) 更新其它映射字段
        for new_key, orig_key in mapping.items():
            val = new_row.get(new_key, "")
            if pd.notna(val) and val != "":
                c = header2col.get(orig_key)
                if c:
                    cell = ws.cell(r, c)
                    cell.value = val
                    cell.fill  = green_fill
        # 3) 更新 Function
        if "Found in function" in header2col and "Function" in header2col:
            fund_val = ws.cell(r, header2col["Found in function"]).value or ""
            for k, v in fund_function_mapping.items():
                if k in fund_val:
                    c = header2col["Function"]
                    cell = ws.cell(r, c)
                    cell.value = v
                    cell.fill  = green_fill
                    break
        # 4) 更新 Root cause
        if "Owner" in header2col and "Root cause" in header2col:
            ow = ws.cell(r, header2col["Owner"]).value or ""
            for k, v in owner_root_cause_mapping.items():
                if k in ow:
                    c = header2col["Root cause"]
                    cell = ws.cell(r, c)
                    cell.value = v
                    cell.fill  = green_fill
                    break

    else:
        # —— 追加新行，并只为非空值单元格染黄 —— 
        # 构造一行值列表，按原表列顺序
        row_vals = []
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for hdr in headers:
            v = ""
            for new_key, orig_key in mapping.items():
                if orig_key == hdr:
                    vv = new_row.get(new_key, "")
                    if pd.notna(vv): v = vv
                    break
            row_vals.append(v)
        ws.append(row_vals)
        new_r = ws.max_row

        # 更新 ID 超链接
        if new_id in id2url_new:
            cell = ws.cell(new_r, id_col)
            cell.hyperlink = id2url_new[new_id]
            cell.style     = "Hyperlink"

        # 只为那些实际写入了非空值的单元格染黄
        for new_key, orig_key in mapping.items():
            vv = new_row.get(new_key, "")
            if pd.notna(vv) and vv != "":
                c = header2col.get(orig_key)
                if c:
                    ws.cell(new_r, c).fill = yellow_fill

        # 追加时也要做 Function 和 Root cause 映射
        if "Found in function" in header2col and "Function" in header2col:
            fund_val = new_row.get("Found in function", "") or ""
            for k, v in fund_function_mapping.items():
                if k in fund_val:
                    c = header2col["Function"]
                    ws.cell(new_r, c).value = v
                    ws.cell(new_r, c).fill  = yellow_fill
                    break

        if "Owner" in header2col and "Root cause" in header2col:
            ow = new_row.get("Owner", "") or ""
            for k, v in owner_root_cause_mapping.items():
                if k in ow:
                    c = header2col["Root cause"]
                    ws.cell(new_r, c).value = v
                    ws.cell(new_r, c).fill  = yellow_fill
                    break

# === 5. 填写 Days、Octane or Jira、Open>20 days、No TIS ===
# Days 列公式
days_idx = header2col.get("Days")
if days_idx:
    for r in range(2, ws.max_row + 1):
        ws.cell(r, days_idx).value = f'=DATEDIF($J{r},TODAY(),"D")'

# Octane or Jira
oir_idx = header2col.get("Octane or Jira")
if oir_idx:
    fn = os.path.basename(new_file)
    val = "Octane" if "Octane" in fn else ("Jira" if "Jira" in fn else "")
    for r in range(2, ws.max_row + 1):
        # 这里只覆盖新追加行或全部？可根据需求改为 start_row
        ws.cell(r, oir_idx).value = val

# Open > 20 days
open20_idx = header2col.get("Open >20 days") or header2col.get("Open > 20 days")
if open20_idx and days_idx:
    dl = get_column_letter(days_idx)
    for r in range(2, ws.max_row + 1):
        ws.cell(r, open20_idx).value = f'=IF({dl}{r}>20,1,0)'

# No TIS
nt_idx = header2col.get("No TIS")
pl_idx = header2col.get("Planned closing version")
ti_idx = header2col.get("Target I-Step:")
if nt_idx and pl_idx and ti_idx:
    pl_l = get_column_letter(pl_idx)
    ti_l = get_column_letter(ti_idx)
    for r in range(2, ws.max_row + 1):
        ws.cell(r, nt_idx).value = f'=IF(OR({pl_l}{r}<>"",{ti_l}{r}<>""),0,1)'

# === 6. 保存 ===
wb.save(updated_file)
print("更新完成！超链接保留／重建，存在行更新染绿，新增行染黄，结果保存在", updated_file)

