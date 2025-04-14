import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --------------------------
# 1. 读取原表和新表数据
# --------------------------
df_original = pd.read_excel("original.xlsx")
df_new = pd.read_excel("new.xlsx")

# --------------------------
# 2. 根据原表的列（类型）筛选新表的数据
# --------------------------
# 找出原表的所有列，过滤新表中存在的列
common_columns = [col for col in df_original.columns if col in df_new.columns]

# 生成新表的子集，新表只保留原表中已有的列
df_new_subset = df_new[common_columns].copy()

# 如果原表中有而新表中没有的列，则在 df_new_subset 中创建该列，并填入空值（或默认值）
for col in df_original.columns:
    if col not in df_new_subset.columns:
        df_new_subset[col] = ""

# 调整 df_new_subset 的列顺序与原表一致
df_new_subset = df_new_subset[df_original.columns]

# --------------------------
# 3. 对需要纠错的字段进行处理（例如：修正 Defect finder 的名称）
# --------------------------
def error_correction(value, column_name):
    """
    此函数实现对特定列数据的纠错，比如修正 Defect finder 列中的错误拼写或大小写。
    """
    if column_name == "Defect finder":
        # 示例：将可能的错误名称映射为正确名称（注意映射时区分大小写）
        correction_mapping = {
            "finderA": "FinderA",
            "FiderA": "FinderA",
            "finderB": "FinderB",
            # 根据实际需要增加其他规则
        }
        return correction_mapping.get(value, value)
    return value

# 假设原表中如果存在 "Defect finder" 列，则对新表该列数据进行纠错处理
if "Defect finder" in df_new_subset.columns:
    df_new_subset["Defect finder"] = df_new_subset["Defect finder"].apply(lambda x: error_correction(x, "Defect finder"))

# --------------------------
# 4. 将新表数据追加到原表中
# --------------------------
# 记录原表的行数，以便后续高亮显示新添加的行
original_row_count = df_original.shape[0]

# 合并数据（向下追加）
df_updated = pd.concat([df_original, df_new_subset], ignore_index=True)

# 保存为新的总表文件
updated_file = "updated_excel.xlsx"
df_updated.to_excel(updated_file, index=False)

# --------------------------
# 5. 使用 openpyxl 对新添加的行进行高亮显示
# --------------------------
wb = load_workbook(updated_file)
ws = wb.active

# 定义高亮填充（黄色）
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Excel 中第一行为表头，所以新行的起始行号 = 原表行数 + 2
start_row = original_row_count + 2
for row in range(start_row, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).fill = highlight_fill

wb.save(updated_file)
print(f"数据追加完成，保存为 {updated_file} ，新添加的行已用黄色高亮标识。")
