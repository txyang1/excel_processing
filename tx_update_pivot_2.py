```python
import os
import re
import time
import json
import pandas as pd
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from watchdog.observers.polling import PollingObserver as Observer
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# === 加载配置 ===
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
with open(os.path.join(BASE_DIR, "unified_config_auto3.json"), 'r', encoding='utf-8') as f:
    cfg = json.load(f)

folders      = cfg['folders']
ORIG_DIR     = os.path.join(BASE_DIR, folders['orig_dir'])
JIRA_DIR     = os.path.join(BASE_DIR, folders['jira_dir'])
OCTANE_DIR   = os.path.join(BASE_DIR, folders['octane_dir'])
sheet        = cfg['sheet']['target_sheet']
sources      = cfg['sources']
settings     = cfg.get('settings', {})
clear_old    = settings.get('clear_old_highlight','N').upper() == 'Y'
fund_patterns  = cfg.get('fund_function_patterns', {})
owner_patterns = cfg.get('owner_root_cause_patterns', {})

# 辅助函数

def is_row_blank(ws, row):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True


def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break


def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            return r
    return 1

# 核心更新逻辑

def update_excel(new_fp):
    # 判断来源目录
    folder = os.path.basename(os.path.dirname(new_fp))
    if folder == os.path.basename(JIRA_DIR):
        source_key = 'Jira'
    elif folder == os.path.basename(OCTANE_DIR):
        source_key = 'Octane'
    else:
        print(f"→ 未识别来源 ({folder})，跳过更新。")
        return

    orig_fp = os.path.join(ORIG_DIR, cfg['paths']['original_file'])
    src_cfg   = sources[source_key]
    df_new    = pd.read_excel(new_fp) if src_cfg['read_method']=='excel' else pd.read_csv(new_fp)

    wb = load_workbook(orig_fp)
    ws = wb[sheet]

    # 清理旧高亮、删除尾部空行并重置行高
    if clear_old:
        green = {'8ED973','008ED973','FF8ED973'}
        blue  = {'FFADD8E6','00ADD8E6','ADD8E6'}
        gray  = {'FFC0C0C0','00C0C0C0','C0C0C0'}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                pf = cell.fill
                if getattr(pf,'patternType',None)=='solid':
                    argb = pf.fgColor.rgb or pf.fgColor.value
                    if argb in green|blue|gray:
                        cell.fill = PatternFill()
    trim_trailing_blank_rows(ws)
    for r in range(1, ws.max_row+1):
        ws.row_dimensions[r].height = None

    # 构建索引
    header2col = {ws.cell(1,c).value: c for c in range(1, ws.max_column+1)}
    id_col     = header2col['ID']
    headers    = list(header2col.keys())

    # 已有 ID→行号 映射
    id2row = {ws.cell(r,id_col).value: r for r in range(2, ws.max_row+1) if ws.cell(r,id_col).value}

    # 提取新文件中的超链接
    id2url_new = {}
    if src_cfg['read_method']=='excel':
        tmpwb = load_workbook(new_fp)
        nws   = tmpwb.active
        cid   = next((c for c in range(1,nws.max_column+1) if nws.cell(1,c).value=='ID'), None)
        if cid:
            for r in range(2, nws.max_row+1):
                cell = nws.cell(r, cid)
                if cell.hyperlink:
                    id2url_new[cell.value] = cell.hyperlink.target

    # 找到最后一行有数据的行号
    original_last = find_last_data_row(ws, id_col)

    update_fill = PatternFill('solid', fgColor='ADD8E6')
    new_fill    = PatternFill('solid', fgColor='8ED973')
    gray_fill   = PatternFill('solid', fgColor='C0C0C0')
    id_key      = next(k for k,v in src_cfg['mapping'].items() if v=='ID')

    # 处理每条新记录
    for _, new_row in df_new.iterrows():
        new_id = new_row.get(id_key)
        if pd.isna(new_id) or not new_id:
            continue

        if new_id in id2row:
            # 已有行——更新字段（省略重复逻辑）
            pass
        else:
            # 新增行——插入到最后一条数据的下一行前
            insert_at = original_last + 1
            ws.insert_rows(insert_at)

            # 填值与高亮
            for idx, hdr in enumerate(headers, start=1):
                val = ''
                for nk, ok in src_cfg['mapping'].items():
                    if ok==hdr:
                        tmp = new_row.get(nk, '')
                        if pd.notna(tmp) and tmp!='':
                            if hdr==src_cfg.get('date_col'):
                                try:
                                    val = pd.to_datetime(tmp).to_pydatetime()
                                except:
                                    val = tmp
                            else:
                                val = tmp
                        break
                cell = ws.cell(insert_at, idx)
                cell.value = val
                if val!='':
                    cell.fill = new_fill
                    if hdr==src_cfg.get('date_col'):
                        cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'

            # 继承超链接
            url = id2url_new.get(new_id)
            if url:
                hl_cell = ws.cell(insert_at, id_col)
                hl_cell.hyperlink = url
                hl_cell.style     = 'Hyperlink'

            original_last += 1

    # 保存——直接写回原表
    wb.save(orig_fp)
    print(f"[{datetime.now().strftime('%Y%m%d_%H%M%S')}] 更新完成并保存到原表 {orig_fp}")

# 监控逻辑
class FolderHandler(FileSystemEventHandler):
    def __init__(self, folders, debounce_seconds=5):
        self.folders   = folders
        self._last_run = {}
        self._debounce = debounce_seconds
    def on_created(self, event):
        if event.is_directory: return
        self._maybe_update(event.src_path)
    def _maybe_update(self, path):
        if os.path.isdir(path): return
        dir_name = os.path.basename(os.path.dirname(path))
        if dir_name not in (folders['jira_dir'], folders['octane_dir']): return
        now, last = time.time(), self._last_run.get(path, 0)
        if now - last < self._debounce:
            print(f"⚠️ 去抖：忽略重复触发 {path}")
            return
        self._last_run[path] = now
        update_excel(path)

if __name__ == '__main__':
    os.makedirs(ORIG_DIR, exist_ok=True)
    os.makedirs(JIRA_DIR, exist_ok=True)
    os.makedirs(OCTANE_DIR, exist_ok=True)
    obs = Observer()
    handler = FolderHandler(folders, debounce_seconds=5)
    obs.schedule(handler, JIRA_DIR, False)
    obs.schedule(handler, OCTANE_DIR, False)
    obs.start()
    print(f"监控中：{JIRA_DIR}, {OCTANE_DIR}")
    try:
        while True: time.sleep(1)
    except KeyboardInterrupt:
        obs.stop(); obs.join()
```
