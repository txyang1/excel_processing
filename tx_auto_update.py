import time
import os
import pandas as pd
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -------------------------------
# 1. 文件路径及参数设置
# -------------------------------
original_file = "original.xlsx"      # 原始 Excel 文件，包含多个 sheet
new_file = "new.xlsx"                # 新数据 Excel 文件（Teams/OneDrive 同步后的文件）
target_sheet = "Octane and jira"     # 目标更新的工作表名称
updated_file = "updated_excel.xlsx"  # 更新后的文件保存路径

# 定义映射关系：新表列名 -> 原表对应的列名
mapping = {
    "Name": "名字",
    "age": "年龄",
    "sex": "性别"
}

# -------------------------------
# 2. 定义自动更新函数
# -------------------------------
def auto_update():
    try:
        print("开始更新文件...")
        # 读取原始 Excel 中目标工作表数据（保留原有数据）
        df_orig = pd.read_excel(original_file, sheet_name=target_sheet)
        
        # 读取新表数据
        df_new = pd.read_excel(new_file)
        
        # 获取原表所有列（确保结构一致）
        orig_columns = list(df_orig.columns)
        
        # 构造新行数据列表，每一行数据以字典形式存储，key 为原表的列名
        new_rows = []
        for idx, new_row in df_new.iterrows():
            row_data = {}
            for col in orig_columns:
                # 遍历原表的每一列，查找是否在映射关系中有对应的新表列
                matched = False
                for new_key, orig_key in mapping.items():
                    if col == orig_key:  # 当前原表列有映射关系
                        value = new_row.get(new_key, "")
                        if pd.isna(value):
                            value = ""
                        row_data[col] = value
                        matched = True
                        break
                if not matched:
                    # 原表中该列不在映射关系中，则填入空值
                    row_data[col] = ""
            new_rows.append(row_data)
        
        # 生成与原表结构一致的新行 DataFrame
        df_new_rows = pd.DataFrame(new_rows, columns=orig_columns)
        
        # 追加新数据到原表末尾（原有数据不更改）
        df_updated_target = pd.concat([df_orig, df_new_rows], ignore_index=True)
        
        # -------------------------------
        # 保留原 Excel 文件中其他 sheet，不更改原有数据，仅更新目标 sheet
        # -------------------------------
        wb = load_workbook(original_file)
        if target_sheet in wb.sheetnames:
            ws_target = wb[target_sheet]
            wb.remove(ws_target)
        
        # 将更新后的目标 sheet 写入工作簿，保留其他 sheet 不变
        with pd.ExcelWriter(updated_file, engine="openpyxl") as writer:
            writer.book = wb
            writer.sheets = {ws.title: ws for ws in wb.worksheets}
            df_updated_target.to_excel(writer, sheet_name=target_sheet, index=False)
            writer.save()
        
        # -------------------------------
        # 对追加的新行进行高亮显示（黄色）
        # -------------------------------
        wb_updated = load_workbook(updated_file)
        ws = wb_updated[target_sheet]
        
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Excel 第一行为表头，数据行从第二行开始，原有数据行数为 df_orig 的行数
        original_rows = df_orig.shape[0]
        start_row = original_rows + 2  # 新数据起始行号
        
        for row in range(start_row, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = highlight_fill
        
        wb_updated.save(updated_file)
        print(f"更新完成！新数据已按照映射关系追加到工作表 '{target_sheet}' 的末尾，结果保存在 {updated_file}")
    except Exception as e:
        print("更新过程中发生错误:", e)

# -------------------------------
# 3. 定义 watchdog 事件处理器
# -------------------------------
class ExcelUpdateHandler(FileSystemEventHandler):
    # 仅针对 .xlsx 文件
    def on_modified(self, event):
        if not event.is_directory:
            if os.path.basename(event.src_path) == os.path.basename(new_file):
                print(f"检测到 {new_file} 被修改: {event.src_path}")
                auto_update()

    def on_created(self, event):
        if not event.is_directory:
            if os.path.basename(event.src_path) == os.path.basename(new_file):
                print(f"检测到新文件 {new_file} 被创建: {event.src_path}")
                auto_update()

# -------------------------------
# 4. 启动监控程序
# -------------------------------
if __name__ == "__main__":
    # 获取 new.xlsx 文件所在目录（假设 new.xlsx 存放在 OneDrive 同步目录中）
    folder_to_watch = os.path.dirname(os.path.abspath(new_file))
    event_handler = ExcelUpdateHandler()
    observer = Observer()
    observer.schedule(event_handler, path=folder_to_watch, recursive=False)
    observer.start()
    print(f"开始监控文件夹: {folder_to_watch} 中的文件 {new_file}")
    
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()
