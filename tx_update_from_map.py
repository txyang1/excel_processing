import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -------------------------------
# 1. 设置文件路径、目标工作表名称与映射关系
# -------------------------------
# 注意：此处 original.xlsx 是包含多个 sheet 的原始文件
# 我们只操作其中名称为 "Octane and jira" 的工作表
original_file = "original.xlsx"
new_file = "new.xlsx"
target_sheet = "Octane and jira"    # 目标 sheet
updated_file = "updated_excel.xlsx"  # 更新后保存的文件

# 定义映射关系（new 表列名 -> 原表列名）
mapping = {
    "Name": "名字",
    "age": "年龄",
    "sex": "性别"
}

# -------------------------------
# 2. 读取原表中目标工作表的数据
# -------------------------------
df_orig = pd.read_excel(original_file, sheet_name=target_sheet)

# -------------------------------
# 3. 读取新表的数据
# -------------------------------
df_new = pd.read_excel(new_file)

# -------------------------------
# 4. 根据映射关系构造新行数据
# -------------------------------
# 我们要求最终生成的每一行必须拥有原表所有的列；
# 若该原表列在映射关系中找到了对应新表的列，则填入相应数据，否则留空。
orig_columns = list(df_orig.columns)
new_rows = []  # 用于存储新行数据

for idx, new_row in df_new.iterrows():
    # 对每一行生成一个字典，其 key 与原表列对应
    row_data = {}
    for col in orig_columns:
        # 遍历原表中每一列，看看是否有映射关系（mapping中的值）与之匹配
        mapped = False
        for new_col, orig_col in mapping.items():
            if orig_col == col:
                # 如果在新表中找到了对应列，则提取数据，如果该单元格没有数据则赋空值
                value = new_row.get(new_col, "")
                if pd.isna(value):
                    value = ""
                row_data[col] = value
                mapped = True
                break
        if not mapped:
            # 原表该列不在映射关系中，则填入空值
            row_data[col] = ""
    new_rows.append(row_data)

# 生成新行的数据框
df_new_rows = pd.DataFrame(new_rows, columns=orig_columns)

# -------------------------------
# 5. 将新数据追加到原表数据的下面
# -------------------------------
df_updated = pd.concat([df_orig, df_new_rows], ignore_index=True)

# 写入到新的 Excel 文件
df_updated.to_excel(updated_file, index=False)

# -------------------------------
# 6. 用 openpyxl 对新增加的行进行高亮标记（黄色）
# -------------------------------
# 注意：Excel 第一行为表头，所以数据行从第2行开始
wb = load_workbook(updated_file)
ws = wb[target_sheet]

highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 原有数据的行数（不含表头）
original_rows = df_orig.shape[0]
# 新追加行的起始行号：原有数据行数 + 2（1 行为表头，另 1 为偏移）
start_row = original_rows + 2

for row in range(start_row, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).fill = highlight_fill

wb.save(updated_file)
print(f"完成数据追加。新数据已按照映射关系追加到原表的末尾，结果保存在 {updated_file}，新行已用黄色高亮显示。")
