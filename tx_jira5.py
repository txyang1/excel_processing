import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# -------------------------------
# 辅助函数：清理尾部空行 & 定位最后有数据行
# -------------------------------
def is_row_blank(ws, row):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True

def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break

def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            return r
    return 1

# -------------------------------
# 1. 文件路径及参数
# -------------------------------
original_file = r"data/Ticket summary (1).xlsx"
new_file      = r"data/EC-EF-2 tickets (CodeCraft Jira) 2025-04-17T09_51_12+0200.csv"
updated_file  = "tx_jira_5.xlsx"
target_sheet  = "Octane and jira"


# 映射关系：CSV 列 -> 原表列
mapping = {
    "Issue key":         "ID",
    "Created":           "Creation time",
    "Summary":           "Name",
    "Status":            "Phase",
    "Reporter":          "Defect finder",
    "Assignee":          "Owner",
    "Affects Version/s": "Target I-Step:",
}

# Fund→Function 映射
fund_function_mapping = {
   "adapt speed to route geometry [01.02.02.15.02.14]": "ASRG",
    
    "change lane [01.02.02.15.02.07]": "CL",
    "Speed Limit Info 21.0 Mid [SLI21.0_Mi] [01.02.02.02.20]": "CL",
   
    "allow hands-off driving 130 [01.02.02.15.02.01]": "HOO130",
    "allow hands-off driving 130 [01.02.02.15.02.01], allow hands-off driving 60 [01.02.02.15.02.02]": "HOO130",
    
    "keep distance [01.02.02.15.02.11]": "KD",
    
    "keep lane [01.02.02.15.02.10]": "KL/KLE",
    "keep lane extended [01.02.02.15.02.08]": "KL/KLE",
    "BS2": "Kl/KLE",
    "Motion Planning [01.02.02.15.02.08.07], Motion Planning [01.02.02.15.02.10.08]": "Kl/KLE",
    "Processing SRR - Detection  [01.02.01.02.08.03.01.13.05.02]": "Kl/KLE",
   
    "display assisted view [01.02.02.15.02.20]": "Adview",
    
    "stop and go at traffic lights [01.02.02.15.02.06]": "SGTL",
    
    "Speed Limit Info [SLI] (incl. No Passing Info)  [01.02.02.09.09.01]": "SLI",
    "ADAS  Interaction with Navigation [01.04.03.01.03.01.01.04], FKT_display_settings_Speed_Limit_Info_SLI21.0_basis [01.02.02.01.03.01.01.03.02.03], Provide Navigation 2.0 [01.04.03.01.03.06]": "SLI",
    "ADAS  Interaction with Navigation [01.04.03.01.03.01.01.04], FKT_display_settings_Speed_Limit_Info_SLI21.0_basis [01.02.02.01.03.01.03.02.02.04.02.03], Provide Navigation 2.0 [01.04.03.01.03.06]": "SLI",
    
    "indicate traffic sign and lights [01.02.02.15.03.01]": "TSLI",
    
    "ADAS  Interaction with Navigation [01.04.03.01.03.01.01.04],allow hands-off driving 130 [01.02.02.15.02.01]": "SAM-China",
    
    "Environment Detection for PA [01.02.02.15.04.03.01.03]": "Parking",
    "Parking Assistant [01.02.02.15.04.03.01]": "Parking",
    
    "stop and go at right of way situations [01.02.02.15.02.04]": "SGROW",
    
    "Autonomous Emergency Braking [01.02.02.15.03.03]": "Kufu function test",
    
  
    "Implement basic platform of Vehicle-to-Everything (V2X)": "V2X",
   
    "Processing SRR - Detection  [01.02.01.02.08.01.02.13.05.02]": "Safty",
   
    "ADAS  Interaction with Navigation [01.04.03.01.03.01.01.04]": "HOO-SAM",
    "ADAS  Interaction with Navigation [01.04.03.01.03.01.01.04], Steering and Lane Control Assistant 3 [LSA3] [01.02.02.12.02.02]": "HOO-SAM",

}

# Owner→Root cause 映射
owner_root_cause_mapping = {
    "Niklas Haeuser" :"Condition evaluate",
    "Ruomeng Guan": "HPL",
    "Cristina delVal": "Icon issue",
    "Daniel Albetal": "IKS",
    "Christoph Romainczyk": "Kufu function issue",
    "Aifa Zhou": "Map issue",
    "Zed Zhang": "Motion plan",
    "zedzhang": "Motion plan",
    "Matthias Stark": "Object fusion",
    "JianLin Zhang": "Obstacle",
    "Fabiao Wang": "Road model",
    "Han Jia": "Road strategy",
    "hanjia": "Road strategy",
    "Juan Carlos Fuentes Michel": "SRR/MRR",
    "juan-carlosfuentes-michel": "SRR/MRR",
    "Meijie Fu": "GNSS",
    "Bingchao Tang": "Traffic light fusion",
    "Liang Xue": "Traffic sign fusion",

    "Daniel Suerth": "CV issue",
    "Ao Zhang": "CV issue",
    "Fisher Yu": "CV issue",
    "Tilmann Bidinger": "CV issue",
    "Regine Graf-Roch": "CV issue",
    "Katarzyna Rzonca": "CV issue",
    "aozhangpartner": "CV issue",
    "Ingo Yang": "CV issue"
}

# 高亮样式
green_fill  = PatternFill("solid", fgColor="00FF00")
yellow_fill = PatternFill("solid", fgColor="FFFF00")

# -------------------------------
# 2. 读取 CSV
# -------------------------------
df_new = pd.read_csv(new_file)

# -------------------------------
# 3. 打开原表并清理空行
# -------------------------------
wb = load_workbook(original_file)
ws = wb[target_sheet]
trim_trailing_blank_rows(ws)

# 构建表头→列号映射 & 列头列表
header2col = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
headers    = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
id_col = header2col["ID"]
original_last = find_last_data_row(ws, id_col)
# 提取原表 ID→行号
id_col = header2col["ID"]
id2row = {
    ws.cell(r, id_col).value: r
    for r in range(2, ws.max_row + 1)
    if ws.cell(r, id_col).value
}

# 定位追加起始行
last_row = find_last_data_row(ws, id_col)

# -------------------------------
# 4. 遍历 CSV，更新或追加
# -------------------------------
for _, new_row in df_new.iterrows():
    new_id = new_row.get("Issue key", "")
    if pd.isna(new_id) or new_id == "":
        continue

    if new_id in id2row:
        # 更新已有行
        r = id2row[new_id]
        # 1) Created -> datetime
        created_raw = new_row.get("Created", "")
        if pd.notna(created_raw) and created_raw != "":
            try:
                ts = pd.to_datetime(created_raw)
                dt = ts.to_pydatetime()
                c = header2col["Creation time"]
                cell = ws.cell(r, c)
                if cell.value != dt:
                    cell.value = dt
                    cell.number_format = "m/d/yyyy h:mm:ss AM/PM"
                    cell.fill = green_fill
            except:
                pass
        # 2) 其他映射字段
        for nk, ok in mapping.items():
            if nk == "Created":
                continue
            new_val = new_row.get(nk, "")
            if pd.isna(new_val) or new_val == "":
                continue
            # 前缀替换 for Target I-Step:
            if nk == "Affects Version/s":
                s = str(new_val)
                if s.startswith("G070") or s.startswith("U006"):
                    new_val = "NA05" + s[4:]
            c = header2col.get(ok)
            cell = ws.cell(r, c)
            if cell.value != new_val:
                cell.value = new_val
                cell.fill  = green_fill
        # 3) Function 列
        if "Found in function" in header2col and "Function" in header2col:
            fund_val = ws.cell(r, header2col["Found in function"]).value or ""
            for k, v in fund_function_mapping.items():
                if k in fund_val:
                    c = header2col["Function"]
                    cell = ws.cell(r, c)
                    if cell.value != v:
                        cell.value = v
                        cell.fill  = green_fill
                    break
        # 4) Root cause 列
        if "Owner" in header2col and "Root cause" in header2col:
            ow = ws.cell(r, header2col["Owner"]).value or ""
            for k, v in owner_root_cause_mapping.items():
                if k in ow:
                    c = header2col["Root cause"]
                    cell = ws.cell(r, c)
                    if cell.value != v:
                        cell.value = v
                        cell.fill  = green_fill
                    break

    else:
        # 追加新行
        last_row += 1
        for idx, hdr in enumerate(headers, start=1):
            val = ""
            for nk, ok in mapping.items():
                if ok == hdr:
                    tmp = new_row.get(nk, "")
                    if pd.notna(tmp) and tmp != "":
                        # Created 转 datetime
                        if nk == "Created":
                            try:
                                ts = pd.to_datetime(tmp)
                                val = ts.to_pydatetime()
                            except:
                                val = tmp
                        else:
                            val = tmp
                        # 前缀替换
                        if nk == "Affects Version/s":
                            s = str(val)
                            if s.startswith("G070") or s.startswith("U006"):
                                val = "NA05" + s[4:]
                    break
            cell = ws.cell(last_row, idx)
            cell.value = val
            if val != "":
                cell.fill = yellow_fill
                if hdr == "Created":
                    cell.number_format = "m/d/yyyy h:mm:ss AM/PM"
        # Function & Root cause for new row
        if "Found in function" in header2col and "Function" in header2col:
            fund_val = new_row.get("Found in function", "") or ""
            for k, v in fund_function_mapping.items():
                if k in fund_val:
                    c = header2col["Function"]
                    ws.cell(last_row, c).value = v
                    ws.cell(last_row, c).fill  = yellow_fill
                    break
        if "Owner" in header2col and "Root cause" in header2col:
            ow = new_row.get("Owner", "") or ""
            for k, v in owner_root_cause_mapping.items():
                if k in ow:
                    c = header2col["Root cause"]
                    ws.cell(last_row, c).value = v
                    ws.cell(last_row, c).fill  = yellow_fill
                    break

# -------------------------------
# 5. 填充公式和其它列
# -------------------------------
max_row = ws.max_row

# 5.1 Days 列
days_idx     = header2col.get("Days")
creation_idx = header2col.get("Creation time")
if days_idx and creation_idx:
    creation_col = get_column_letter(creation_idx)
    for r in range(2, max_row + 1):
        ws.cell(r, days_idx).value = f'=DATEDIF(${creation_col}{r},TODAY(),"D")'

# 5.2 Octane or Jira 列 (CSV 来自 Jira)
oir_idx = header2col.get("Octane or Jira")
if oir_idx:
    for r in range(original_last+1, max_row + 1):
        ws.cell(r, oir_idx).value = "Jira"

# 5.3 Open > 20 days 列
open20_idx = header2col.get("Open >20 days") or header2col.get("Open > 20 days")
if open20_idx and days_idx:
    dl = get_column_letter(days_idx)
    for r in range(2, max_row + 1):
        ws.cell(r, open20_idx).value = f'=IF({dl}{r}>20,1,0)'

# 5.4 No TIS 列
nt_idx = header2col.get("No TIS")
pl_idx = header2col.get("Planned closing version")
ti_idx = header2col.get("Target I-Step:")
if nt_idx and pl_idx and ti_idx:
    pl_l = get_column_letter(pl_idx)
    ti_l = get_column_letter(ti_idx)
    for r in range(2, max_row + 1):
        ws.cell(r, nt_idx).value = f'=IF(OR({pl_l}{r}<>"",{ti_l}{r}<>""),0,1)'

# -------------------------------
# 6. 保存结果
# -------------------------------
wb.save(updated_file)
print("更新完成！Creation time 已转为日期对象，Days 公式可用，差异染绿，新行染黄，结果保存在", updated_file)
