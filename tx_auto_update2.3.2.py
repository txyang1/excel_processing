# auto_excel_update.py

import os
import time
import json
import pandas as pd
import schedule
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# === 加载配置 ===
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
with open(os.path.join(BASE_DIR, "unified_config_auto.json"), 'r', encoding='utf-8') as f:
    cfg = json.load(f)

folders    = cfg['folders']
ORIG_DIR   = os.path.join(BASE_DIR, folders['orig_dir'])
JIRA_DIR   = os.path.join(BASE_DIR, folders['jira_dir'])
OCTANE_DIR = os.path.join(BASE_DIR, folders['octane_dir'])

paths      = cfg['paths']
sheet      = cfg['sheet']['target_sheet']
sources    = cfg['sources']
settings   = cfg.get('settings', {})
clear_old  = settings.get('clear_old_highlight', 'N').upper() == 'Y'
fund_pats  = cfg.get('fund_function_patterns', {})
owner_pats = cfg.get('owner_root_cause_patterns', {})

def is_row_blank(ws, row):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True

def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break

def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, key_col).value not in (None, ""):
            return r
    return 1

def update_excel(new_fp):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    print(f"[{ts}] 开始更新: {new_fp}")

    folder = os.path.basename(os.path.dirname(new_fp))
    if folder == os.path.basename(JIRA_DIR):
        source_key  = "Jira"
        jira_flag   = 'Y'
        octane_flag = 'N'
    elif folder == os.path.basename(OCTANE_DIR):
        source_key  = "Octane"
        jira_flag   = 'N'
        octane_flag = 'Y'
    else:
        print(f" → 未识别来源 ({folder})，跳过。")
        return

    # 原表与新版输出路径
    orig_fp  = os.path.join(ORIG_DIR, paths['original_file'])
    template = paths['updated_file']
    filename = template.format(time=ts, jira=jira_flag, octane=octane_flag)
    upd_fp   = os.path.join(os.path.dirname(new_fp), filename)

    src_cfg   = sources[source_key]
    read_meth = src_cfg['read_method']
    date_col  = src_cfg.get('date_col')
    mapping   = src_cfg['mapping']

    # 读取新表
    df_new = pd.read_excel(new_fp) if read_meth == 'excel' else pd.read_csv(new_fp)

    # 打开原表
    wb = load_workbook(orig_fp)
    ws = wb[sheet]

    # 清除旧高亮
    if clear_old:
        green = {"8ED973","008ED973","FF8ED973"}
        blue  = {'FFADD8E6','00ADD8E6','ADD8E6'}
        gray  = {'FFC0C0C0','00C0C0C0','C0C0C0'}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                pf = cell.fill
                if getattr(pf, 'patternType', None) == 'solid':
                    argb = pf.fgColor.rgb or pf.fgColor.value
                    if argb in green|blue|gray:
                        cell.fill = PatternFill()

    trim_trailing_blank_rows(ws)

    header2col = { ws.cell(1,c).value: c for c in range(1, ws.max_column+1) }
    id_col     = header2col['ID']
    id2row     = {
        ws.cell(r, id_col).value: r
        for r in range(2, ws.max_row+1)
        if ws.cell(r, id_col).value
    }

    # Excel 超链接收集
    id2url_new = {}
    if read_meth == 'excel':
        tmpwb = load_workbook(new_fp)
        nws   = tmpwb.active
        cid   = next((c for c in range(1, nws.max_column+1) if nws.cell(1,c).value=='ID'), None)
        if cid:
            for r in range(2, nws.max_row+1):
                cell = nws.cell(r,cid)
                if cell.hyperlink:
                    id2url_new[cell.value] = cell.hyperlink.target

    last_row      = find_last_data_row(ws, id_col)
    update_fill   = PatternFill('solid', fgColor='ADD8E6')
    new_fill      = PatternFill('solid', fgColor='8ED973')

    id_key = next(k for k,v in mapping.items() if v == 'ID')

    # 更新或追加
    for _, new_row in df_new.iterrows():
        nid = new_row.get(id_key)
        if pd.isna(nid) or not nid:
            continue

        if nid in id2row:
            r = id2row[nid]
            phase = ws.cell(r, header2col.get('Phase')).value or ''
            if any(x in phase for x in ('Concluded','Closed','Resolved')):
                continue

            # 更新日期
            if date_col:
                raw = new_row.get(date_col)
                if pd.notna(raw):
                    try:
                        dt = pd.to_datetime(raw).to_pydatetime()
                        c  = header2col[mapping[date_col]]
                        cell = ws.cell(r,c)
                        if cell.value != dt:
                            cell.value         = dt
                            cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'
                            cell.fill          = update_fill
                    except:
                        pass

            # 更新其他字段
            for nk, ok in mapping.items():
                if nk == date_col or ok == 'Function':
                    continue
                val = new_row.get(nk)
                if pd.isna(val) or val == "":
                    continue
                if ok == "Involved I-Step" and str(val).startswith(("G070","U006")):
                    val = "NA05" + str(val)[4:]
                c = header2col.get(ok)
                if c and ws.cell(r,c).value != val:
                    ws.cell(r,c).value = val
                    ws.cell(r,c).fill  = update_fill

            # Root cause
            if "Owner" in header2col and "Root cause" in header2col:
                ow = str(ws.cell(r,header2col["Owner"]).value or '').lower()
                for cause, names in owner_pats.items():
                    if any(n.lower() in ow for n in names):
                        cell = ws.cell(r,header2col["Root cause"])
                        if cell.value != cause:
                            cell.value = cause
                            cell.fill  = update_fill
                        break

        else:
            last_row += 1
            for hdr, cidx in header2col.items():
                val = ""
                for nk, ok in mapping.items():
                    if ok == hdr:
                        tmp = new_row.get(nk)
                        if pd.notna(tmp):
                            val = tmp
                        break
                cell = ws.cell(last_row, cidx)
                cell.value = val
                if val:
                    cell.fill = new_fill
                    if hdr == mapping.get(date_col):
                        cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'

            # 超链接
            if nid in id2url_new:
                c = ws.cell(last_row, id_col)
                c.hyperlink, c.style = id2url_new[nid], 'Hyperlink'

    # 填充 Days 公式
    if "Days" in header2col and "Creation time" in header2col:
        colL = get_column_letter(header2col["Creation time"])
        for r in range(2, ws.max_row+1):
            ws.cell(r, header2col["Days"]).value = f'=DATEDIF(${colL}{r},TODAY(),"D")'

    wb.save(upd_fp)
    print(f"[{ts}] 完成，保存至: {upd_fp}")

class FolderHandler(FileSystemEventHandler):
    def on_created(self, event):
        if not event.is_directory:
            fld = os.path.basename(os.path.dirname(event.src_path))
            if fld in (folders['jira_dir'], folders['octane_dir']):
                update_excel(event.src_path)
    on_modified = on_created

def main():
    # 确保目录存在
    for d in (ORIG_DIR, JIRA_DIR, OCTANE_DIR):
        os.makedirs(d, exist_ok=True)

    # 启动 Watchdog
    observer = Observer()
    observer.schedule(FolderHandler(), path=JIRA_DIR, recursive=False)
    observer.schedule(FolderHandler(), path=OCTANE_DIR, recursive=False)
    observer.start()
    print(f"监控目录：{JIRA_DIR}，{OCTANE_DIR}")

    # 定时每日 08:00 全量扫描
    schedule.every().day.at("08:00").do(lambda: [
        update_excel(os.path.join(JIRA_DIR, fn)) for fn in os.listdir(JIRA_DIR) if fn.lower().endswith((".xlsx", ".csv"))
    ] + [
        update_excel(os.path.join(OCTANE_DIR, fn)) for fn in os.listdir(OCTANE_DIR) if fn.lower().endswith((".xlsx", ".csv"))
    ])
    print("已设定每日 08:00 全量扫描更新任务。")

    try:
        while True:
            schedule.run_pending()
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

if __name__ == "__main__":
    main()
