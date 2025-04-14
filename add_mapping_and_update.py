import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -------------------------------
# 1. 设置文件及目标 sheet 名称
# -------------------------------
original_file = "original.xlsx"       # 包含多个 sheet 的原始 Excel
new_file = "new.xlsx"                 # 新数据文件
target_sheet = "Octane and jira"      # 仅更新该工作表
updated_file = "updated_excel.xlsx"   # 保存更新后的 Excel

# -------------------------------
# 2. 读取原始 Excel 中目标工作表数据
# -------------------------------
df_original = pd.read_excel(original_file, sheet_name=target_sheet)

# -------------------------------
# 3. 读取新数据，并提取原表中已有列的数据
# -------------------------------
df_new = pd.read_excel(new_file)

# 保留两表中共有的列
common_columns = [col for col in df_original.columns if col in df_new.columns]
df_new_subset = df_new[common_columns].copy()

# 对于原表中存在但新数据中缺失的列，补充空值（或可默认值）
for col in df_original.columns:
    if col not in df_new_subset.columns:
        df_new_subset[col] = ""
        
# 调整 df_new_subset 列的顺序与原表一致
df_new_subset = df_new_subset[df_original.columns]

# -------------------------------
# 4. 针对特定字段进行纠错：例如修正 "Defect finder" 列中的常见错误
# -------------------------------
def error_correction(value, column_name):
    if column_name == "Defect finder":
        # 纠错映射：例如将 "finderA" 或 "FiderA" 修正为 "FinderA"
        correction_mapping = {
            "finderA": "FinderA",
            "FiderA": "FinderA",
            "finderB": "FinderB"
            # 可根据实际情况增加其他规则
        }
        return correction_mapping.get(value, value)
    return value

if "Defect finder" in df_new_subset.columns:
    df_new_subset["Defect finder"] = df_new_subset["Defect finder"].apply(lambda x: error_correction(x, "Defect finder"))

# -------------------------------
# 5. 利用映射关系更新 "Function" 列
# -------------------------------
# 定义从 Defect finder 到 Function 映射的字典（可自定义扩充）
defect_finder_to_function_mapping = {
    "Li Hu": "CL",
    # 如有其他映射，可添加，如： "Zhang Wei": "XX", ...
}

def update_function_with_mapping(row):
    # 假设原来的 Function 列来源于 "Assigned ECU Group"（或其它已有字段）
    current_func = row.get("Function", "")
    defect_finder = row.get("Defect finder", "")
    if pd.notnull(defect_finder) and defect_finder in defect_finder_to_function_mapping:
        mapped_value = defect_finder_to_function_mapping[defect_finder]
        # 若当前 Function 为空，则直接使用映射值；否则追加，确保不重复
        if pd.isna(current_func) or current_func == "":
            return mapped_value
        else:
            if mapped_value not in str(current_func):
                return f"{current_func}, {mapped_value}"
    return current_func

# 如果原表中没有 "Function" 列，而你希望其来源于其它字段（例如 "Assigned ECU Group"），
# 可在此处进行处理。这里假定原数据中已有 "Function" 列，若无则需要先创建。
if "Function" in df_new_subset.columns:
    df_new_subset["Function"] = df_new_subset.apply(update_function_with_mapping, axis=1)
    
# -------------------------------
# 6. 将新数据追加到原始数据中
# -------------------------------
original_row_count = df_original.shape[0]  # 记录原有数据行数

# 追加新数据（注意两边列顺序已经一致）
df_updated_sheet = pd.concat([df_original, df_new_subset], ignore_index=True)

# -------------------------------
# 7. 将更新后的数据写回 Excel，保持其它 sheet 不变
# -------------------------------
# 首先利用 openpyxl 加载原始 Excel 工作簿（包含多个 sheet）
wb = load_workbook(original_file)

# 删除原来名为 target_sheet 的工作表（后面用新数据替换）
if target_sheet in wb.sheetnames:
    ws_to_remove = wb[target_sheet]
    wb.remove(ws_to_remove)

# 利用 ExcelWriter 将更新后的目标 sheet 写入工作簿中
with pd.ExcelWriter(original_file, engine='openpyxl') as writer:
    writer.book = wb
    df_updated_sheet.to_excel(writer, sheet_name=target_sheet, index=False)
    writer.save()

# 为了方便后续核查，复制为 updated_file
wb.save(updated_file)

# -------------------------------
# 8. 高亮显示新增的行（新数据部分）——使用 openpyxl
# -------------------------------
wb_updated = load_workbook(updated_file)
ws = wb_updated[target_sheet]

# 定义高亮填充色（黄色）
highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Excel 第一行为表头，所以数据行从第二行开始
start_row = original_row_count + 2  # 原始行数 + 表头偏移
for row in range(start_row, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).fill = highlight_fill

wb_updated.save(updated_file)
print(f"更新完成，已生成 {updated_file}。目标工作表 '{target_sheet}' 中的新数据行已用黄色高亮显示。")
