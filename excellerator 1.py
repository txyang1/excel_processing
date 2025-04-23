import os
import sys
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# Directory containing the Excel files
directory = input("Please enter the path to the excel folder: ")
timestamp = input("Please enter the timestamp (e.g. 20250421_0905) ")
istep_and_cw = input("Please enter istep and CW  (e.g. 490CW16) ")
#directory = '/home/qxz1ssb/Desktop/tools/01_python/ecef2testdriveexcellerator/excels/CW16'
#timestamp = "20250421_0905"
#istep_and_cw = "490CW16"

# Dictionary to hold all data with headers as keys
data_dict = {}
# List to hold the file names for each row
file_names = []

# Check if summary.xlsx already exists in the directory
summary_file_path = os.path.join(directory, f"summary_{timestamp}_{istep_and_cw}.xlsx")
existing_summary = os.path.exists(summary_file_path)

if existing_summary:
    # Load the existing summary file
    wb = load_workbook(summary_file_path)
    ws_trigger = wb['TRIGGER']
    
    # Read the existing summary data into a dataframe
    existing_summary_df = pd.read_excel(summary_file_path, sheet_name='TRIGGER')
    
    # Extract existing file names
    existing_file_names = existing_summary_df['File Name'].unique().tolist()
else:
    existing_file_names = []

# Loop through all files in the directory
for filename in os.listdir(directory):
    if filename.endswith(".xlsm") and not filename.startswith("summary") and filename not in existing_file_names:
        # Construct full file path
        file_path = os.path.join(directory, filename)
        
        # Load the Excel file
        xls = pd.ExcelFile(file_path)
        
        # Check if 'TRIGGER' sheet exists in the Excel file
        if 'TRIGGER' in xls.sheet_names:
            # Read the 'TRIGGER' sheet into a dataframe, skipping the first 6 rows
            df = pd.read_excel(file_path, sheet_name='TRIGGER', skiprows=6)
            
            # Drop the 8th row (which is the first row after skipping 6 rows)
            df = df.drop(df.index[0])

            # Delete unnecessary columns 
            cols_to_delete = df.columns[33:200] 
            df = df.drop(columns=cols_to_delete)
            
            # Extract headers from the 7th row
            headers = df.columns.tolist()

            # Append the timestamp and istep_and_cw to the filename
            appended_filename = f"{filename}_{timestamp}_{istep_and_cw}"
            
            # Add modified filename to the file names list for each row in the dataframe
            file_names.extend([appended_filename] * len(df))
            
            # Iterate through the headers and add data to the dictionary
            for header in headers:
                if header not in data_dict:
                    data_dict[header] = df[header].tolist()
                else:
                    data_dict[header].extend(df[header].tolist())

        else:
            print("\nIgnored excel(s) due to missing TRIGGER sheet:")
            print(filename)

try:
    # Create a summary dataframe with headers as columns
    new_data_df = pd.DataFrame(data_dict)
except Exception as e:
    print(f"\nError: {e}")
    print("One or more Excels have a wrong format, please delete one by one to find the wrong Excel and repair it. You can start by checking row 7.")
    sys.exit()

# Add the file names column
new_data_df.insert(0, 'File Name', file_names)

# Format the "Date" column if it exists
if 'Date' in new_data_df.columns:
    new_data_df['Date'] = pd.to_datetime(new_data_df['Date'], errors='coerce').dt.strftime('%b %d')

# Add "Comment" column in the end
new_data_df['Comment'] = ''

# Move the "File Name" column after the "Comment" column
cols = new_data_df.columns.tolist()
cols.append(cols.pop(cols.index('File Name')))
new_data_df = new_data_df[cols]

if existing_summary:
    # Append new data to the existing summary dataframe
    summary_df = pd.concat([existing_summary_df, new_data_df], ignore_index=True)
else:
    # Use new data as the summary dataframe
    summary_df = new_data_df

# Save the concatenated dataframe to the summary Excel file
summary_df.to_excel(summary_file_path, index=False, engine='openpyxl', sheet_name='TRIGGER')

# Load the summary file to modify it
wb = load_workbook(summary_file_path)
ws_trigger = wb['TRIGGER']

# Add filters to the columns
ws_trigger.auto_filter.ref = ws_trigger.dimensions

# Create or update the OVERVIEW sheet
if 'OVERVIEW' in wb.sheetnames:
    ws_overview = wb['OVERVIEW']
    ws_overview.delete_rows(2, ws_overview.max_row - 1)
else:
    ws_overview = wb.create_sheet(title='OVERVIEW')

# Get the "Status" column values from the TRIGGER sheet
status_values = summary_df['Status']

# Count the occurrences of each unique value in the "Status" column
status_counts = status_values.value_counts()

# Write the "Status" values and their counts to the OVERVIEW sheet
ws_overview.append(['Status', 'Count'])
for status, count in status_counts.items():
    ws_overview.append([status, count])

# Insert a bar chart based on the "Status" and "Count" data
chart = BarChart()
chart.title = "Status Counts"
chart.x_axis.title = "Status"
chart.y_axis.title = "Count"

data = Reference(ws_overview, min_col=2, min_row=1, max_row=len(status_counts) + 1, max_col=2)
categories = Reference(ws_overview, min_col=1, min_row=2, max_row=len(status_counts) + 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

ws_overview.add_chart(chart, "E5")

# Define the statuses to exclude
excluded_statuses = ["done", "OCT-Ticket", "known", "Event", "ToDo done"]

# Check if all "Status" values for each "File Name" are in the excluded statuses
in_analysis_files = []
done_files = []

for file_name in summary_df['File Name'].unique():
    file_status = summary_df[summary_df['File Name'] == file_name]['Status']
    if all(status in excluded_statuses or pd.isna(status) for status in file_status):
        done_files.append(file_name)
    else:
        in_analysis_files.append(file_name)

# Write the "In Analysis" file names to the OVERVIEW sheet starting from cell A20
start_row = 20
ws_overview.cell(row=start_row, column=1, value='In Analysis')

for i, file_name in enumerate(in_analysis_files, start=start_row + 1):
    ws_overview.cell(row=i, column=1, value=file_name)

# Add the "Done" header starting from cell A20 after the In Analysis list
done_start_row = start_row + len(in_analysis_files) + 2
ws_overview.cell(row=done_start_row, column=1, value='Done')

for i, file_name in enumerate(done_files, start=done_start_row + 1):
    ws_overview.cell(row=i, column=1, value=file_name)

# Hide columns: #, Excel-Session, GPS Position, Road_Ext_QU, Software, Unnamed: 11, Unnamed: 12, Unnamed: 14, Unnamed: 16, Canape-Folder, BI, Error occurence, TIS, Category, Cluster, CANape device 1, CANape device 2
columns_to_hide = ['#', 'Excel-Session', 'GPS Position', 'Road_Ext_QU', 'Software', 'Unnamed: 11', 'Unnamed: 12', 'Unnamed: 14', 'Unnamed: 16', 'Canape-Folder', 'BI', 'Error occurence', 'TIS', 'Category', 'Cluster', 'CANape device 1', 'CANape device 2']
for col in columns_to_hide:
    if col in summary_df.columns:
        col_idx = summary_df.columns.get_loc(col) + 1
        ws_trigger.column_dimensions[get_column_letter(col_idx)].hidden = True

# Set the zoom level to 80%
ws_trigger.sheet_view.zoomScale = 80

# Auto adjust column width for Timestamp, File Name, and headers
def auto_adjust_column_width(ws, df, columns):
    for col in columns:
        if col in df.columns:
            col_idx = df.columns.get_loc(col) + 1
            max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2
            max_length = min(max_length, 30)  # Set maximum width to 30
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length

auto_adjust_column_width(ws_trigger, summary_df, ['File Name', 'Date'] + summary_df.columns.tolist())

# Make columns bigger: Event, Analysis, Solution
columns_to_enlarge = ['Event', 'Analysis', 'Solution']
for col in columns_to_enlarge:
    if col in summary_df.columns:
        col_idx = summary_df.columns.get_loc(col) + 1
        ws_trigger.column_dimensions[get_column_letter(col_idx)].width = 50

# Set specific column widths
specific_columns_to_enlarge = ['Canape-Trigger', 'Vigem-Trigger', 'Datacenter', 'Comment']
for col in specific_columns_to_enlarge:
    if col in summary_df.columns:
        col_idx = summary_df.columns.get_loc(col) + 1
        ws_trigger.column_dimensions[get_column_letter(col_idx)].width = 20

# Set the header background color to light grey
header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
for cell in ws_trigger[1]:
    cell.fill = header_fill

# Define status options and their corresponding colors
status_options = {
    "open": "FFCCCC",  # Light Red
    "forward": "FFCCCC",  # Light Red
    "in analysis": "FFFFCC",  # Light Yellow
    "done": "CCFFCC",  # Light Green
    "OCT-Ticket create": "FFFFCC",  # Light Yellow
    "OCT-Ticket": "CCFFCC",  # Light Green
    "Evaluation open": "FFFFCC",  # Light Yellow
    "known": "CCFFCC",  # Light Green
    "new trigger": "FFCCCC",  # Light Red
    "see": "E0E0E0",  # Light Grey
    "see Cluster": "E0E0E0",  # Light Grey
    "Event": "CCCCFF",  # Light Blue
    "ToDo": "FFCCCC",  # Light Red
    "ToDo done": "CCFFCC"  # Light Green
}

# Add data validation for the "Status" column
status_list = ",".join(status_options.keys())
dv = DataValidation(type="list", formula1=f'"{status_list}"', allow_blank=True)
ws_trigger.add_data_validation(dv)

status_col_idx = summary_df.columns.get_loc('Status') + 1
for row in range(2, ws_trigger.max_row + 1):
    cell = ws_trigger.cell(row=row, column=status_col_idx)
    dv.add(cell)

# Apply conditional formatting based on the status options
for status, color in status_options.items():
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    rule = CellIsRule(operator="equal", formula=[f'"{status}"'], fill=fill)
    ws_trigger.conditional_formatting.add(f"{get_column_letter(status_col_idx)}2:{get_column_letter(status_col_idx)}{ws_trigger.max_row}", rule)

# Save the workbook
wb.save(summary_file_path)

print(f"Summary Excel file created successfully at {summary_file_path}!")