import os
import time
import json
import pandas as pd
import schedule
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# === 目录配置 ===
BASE_DIR   = os.path.abspath(os.path.dirname(__file__))
ORIG_DIR   = os.path.join(BASE_DIR, "Orig_files")
JIRA_DIR   = os.path.join(BASE_DIR, "Jira_files")
OCTANE_DIR = os.path.join(BASE_DIR, "Octane_files")
CONFIG_PATH = os.path.join(BASE_DIR, "unified_config2.7.1.json")

# === 辅助函数 ===
def is_row_blank(ws, row):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True


def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break


def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            return r
    return 1

# === 更新流程 ===
def update_excel(new_fp):
    print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] 开始更新: {new_fp}")
    # 判断源类型
    folder = os.path.basename(os.path.dirname(new_fp))
    if folder == "Jira_files":
        source_key = "Jira"
    elif folder == "Octane_files":
        source_key = "Octane"
    else:
        print(f"未在监控的文件夹中({folder})，跳过。")
        return

    # 加载配置
    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        cfg = json.load(f)

    # 构建路径
    orig_name  = os.path.basename(cfg['paths']['original_file'])
    orig_fp    = os.path.join(ORIG_DIR, orig_name)
    upd_name   = os.path.basename(cfg['paths']['updated_file'])
    dest_dir   = os.path.dirname(new_fp)
    upd_fp     = os.path.join(dest_dir, upd_name)

    sheet      = cfg['sheet']['target_sheet']
    sources    = cfg['sources']
    settings   = cfg.get('settings', {})
    clear_old  = settings.get('clear_old_highlight','N').upper() == 'Y'
    fund_pats  = cfg.get('fund_function_patterns', {})
    owner_pats = cfg.get('owner_root_cause_patterns', {})

    src_cfg    = sources[source_key]
    read_meth  = src_cfg['read_method']
    date_col   = src_cfg.get('date_col')
    mapping    = src_cfg['mapping']

    # 读取新表
    df_new = pd.read_excel(new_fp) if read_meth=='excel' else pd.read_csv(new_fp)

    # 打开原表
    wb = load_workbook(orig_fp)
    ws = wb[sheet]
    if clear_old:
        pink = {'FFC0CB','00FFC0CB','FFFFC0CB'}
        blue = {'FFADD8E6','00ADD8E6','ADD8E6'}
        gray = {'FFC0C0C0','00C0C0C0','C0C0C0'}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                pf = cell.fill
                if getattr(pf, 'patternType', None)=='solid':
                    argb = pf.fgColor.rgb or pf.fgColor.value
                    if argb in pink|blue|gray:
                        cell.fill = PatternFill()
    trim_trailing_blank_rows(ws)

    # 表头列索引
    header2col = {ws.cell(1,c).value: c for c in range(1, ws.max_column+1)}
    id_col      = header2col['ID']

    # 已有ID到行
    id2row = {ws.cell(r,id_col).value: r for r in range(2, ws.max_row+1)
              if ws.cell(r,id_col).value}

    # 超链接收集（仅 Excel 来源）
    id2url = {}
    if read_meth=='excel':
        tmp = load_workbook(new_fp).active
        cid = next((c for c in range(1,tmp.max_column+1) if tmp.cell(1,c).value=='ID'), None)
        if cid:
            for r in range(2, tmp.max_row+1):
                c = tmp.cell(r,cid)
                if c.hyperlink: id2url[c.value] = c.hyperlink.target

    last_row = find_last_data_row(ws, id_col)
    pink_fill = PatternFill('solid', fgColor='FFC0CB')
    blue_fill = PatternFill('solid', fgColor='ADD8E6')

    # 新ID列键
    id_key = next(k for k,v in mapping.items() if v=='ID')
    # 遍历
    for _, nr in df_new.iterrows():
        nid = nr.get(id_key)
        if pd.isna(nid) or not nid: continue
        if nid in id2row:
            r = id2row[nid]
            phase = (ws.cell(r, header2col.get('Phase')).value or '')
            if any(x in phase for x in ('Concluded','Closed','Resolved')): continue
            # 更新字段略，填充 blue
        else:
            last_row += 1
            for hdr, col_idx in header2col.items():
                val = ''
                for nk, ok in mapping.items():
                    if hdr==ok:
                        tmp = nr.get(nk, '')
                        if pd.notna(tmp): val = tmp
                        break
                c = ws.cell(last_row,col_idx)
                c.value = val
                if val: c.fill = pink_fill
            # 超链接
            if nid in id2url:
                c = ws.cell(last_row, id_col)
                c.hyperlink, c.style = id2url[nid], 'Hyperlink'
            # 匹配 Function/Root cause，可按需添加

    # 填充公式（同之前逻辑）
    max_r = ws.max_row
    if 'Days' in header2col and 'Creation time' in header2col:
        colL = get_column_letter(header2col['Creation time'])
        for r in range(2, max_r+1):
            ws.cell(r, header2col['Days']).value = f'=DATEDIF(${colL}{r},TODAY(),"D")'
    if 'Octane or Jira' in header2col:
        for r in range(last_row+1, max_r+1):
            ws.cell(r, header2col['Octane or Jira']].value = source_key

    # 保存
    wb.save(upd_fp)
    print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] 更新完成: {upd_fp}")

# === 事件处理 ===
class FolderHandler(FileSystemEventHandler):
    def on_created(self, event):
        if event.is_directory: return
        folder = os.path.basename(os.path.dirname(event.src_path))
        if folder in ('Jira_files','Octane_files'):
            update_excel(event.src_path)

    def on_modified(self, event):
        self.on_created(event)

# === 启动监控 & 定时 ===
def main():
    # Watchdog
    observer = Observer()
    for d in (JIRA_DIR, OCTANE_DIR):
        os.makedirs(d, exist_ok=True)
        observer.schedule(FolderHandler(), path=d, recursive=False)
    observer.start()
    print(f"监控文件夹: {JIRA_DIR}, {OCTANE_DIR}")

    # 定时每日 08:00 扫描并更新所有文件
    def daily_scan():
        for d in (JIRA_DIR, OCTANE_DIR):
            for fn in os.listdir(d):
                fp = os.path.join(d, fn)
                if os.path.isfile(fp) and fn.lower().endswith(('.xlsx','.csv')):
                    update_excel(fp)
    schedule.every().day.at("08:00").do(daily_scan)
    print("已添加每日 08:00 自动扫描更新任务。")

    try:
        while True:
            schedule.run_pending()
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

if __name__ == '__main__':
    os.makedirs(ORIG_DIR, exist_ok=True)
    os.makedirs(JIRA_DIR, exist_ok=True)
    os.makedirs(OCTANE_DIR, exist_ok=True)
    main()
