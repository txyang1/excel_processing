import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =============================
# 1. 参数设置
# =============================
original_file = "original.xlsx"          # 包含多个 sheet 的原始 Excel
jira_file = "jira.csv"                    # Jira 的 CSV 文件（注意文件格式为 CSV）
target_sheet = "Octane and jira"          # 仅更新该工作表
updated_file = "updated_excel.xlsx"       # 更新后保存的文件

# =============================
# 2. 读取原始 Excel 中目标工作表数据
# =============================
df_original = pd.read_excel(original_file, sheet_name=target_sheet)

# =============================
# 3. 读取 Jira CSV 数据，并重命名字段以匹配原表
# =============================
df_jira = pd.read_csv(jira_file)
# 重命名映射：Jira中的 Summary、issue key、Assignee、Reporter 分别映射到原表的 Name、Ticket no. supplier、Owner、Defect finder
rename_dict = {
    "Summary": "Name",
    "issue key": "Ticket no. supplier",
    "Assignee": "Owner",
    "Reporter": "Defect finder"
}
df_jira.rename(columns=rename_dict, inplace=True)
# 只保留需要参与更新的字段（关键字段 + 映射字段）
df_jira = df_jira[["Ticket no. supplier", "Name", "Owner", "Defect finder"]]

# =============================
# 4. 合并原始数据与 Jira 数据
# =============================
# 以 “Ticket no. supplier” 为唯一键左合并，将 Jira 数据附加，Jira中的字段在合并后后缀为 "_jira"
df_merged = pd.merge(df_original, df_jira, on="Ticket no. supplier", how="left", suffixes=("", "_jira"))

# 如果原表中不存在 “Octane or Jira” 列，则添加该列，默认空字符串
if "Octane or Jira" not in df_merged.columns:
    df_merged["Octane or Jira"] = ""

# =============================
# 5. 对比更新指定字段，并记录更新
# =============================
# 需要更新的字段，注意：更新字段来自原表而更新数据来自 Jira 合并后的 *_jira 列
update_fields = ["Name", "Owner", "Defect finder"]

# 用于记录更新的单元格（格式：(pandas行索引, 列名称)）
updated_cells = []

for idx, row in df_merged.iterrows():
    for field in update_fields:
        jira_field = f"{field}_jira"
        jira_val = row.get(jira_field)
        orig_val = row.get(field)
        # 若 Jira 中该字段有数据，且值不为空，且与原值不同，则更新原表对应字段
        if pd.notnull(jira_val) and str(jira_val).strip() != "" and (pd.isnull(orig_val) or str(orig_val).strip() != str(jira_val).strip()):
            df_merged.at[idx, field] = jira_val
            updated_cells.append((idx, field))
    # 同时，如果该行在 Jira 中存在至少一项更新信息，则将 “Octane or Jira” 列设置为 "jira"
    # 这里判断条件可以是：对应任一 *_jira 字段存在有效数据
    if (pd.notnull(row.get("Name_jira")) and str(row.get("Name_jira")).strip() != "") or \
       (pd.notnull(row.get("Owner_jira")) and str(row.get("Owner_jira")).strip() != "") or \
       (pd.notnull(row.get("Defect finder_jira")) and str(row.get("Defect finder_jira")).strip() != ""):
        df_merged.at[idx, "Octane or Jira"] = "jira"
        # 同时记录此列更新（如果之前的值为空或不同）
        if row.get("Octane or Jira", "").strip().lower() != "jira":
            updated_cells.append((idx, "Octane or Jira"))

# =============================
# 6. 去除合并后不需要的 *_jira 扩展字段，保留原表列结构
# =============================
final_columns = df_original.columns.tolist()
# 如果原表中已含“Octane or Jira”，确保放在最终列中
if "Octane or Jira" not in final_columns and "Octane or Jira" in df_merged.columns:
    final_columns.append("Octane or Jira")
df_final = df_merged[final_columns]

# =============================
# 7. 将更新后的数据写入目标工作表，同时保留其他 sheet 不变
# =============================
# 使用 openpyxl 加载整个工作簿
wb = load_workbook(original_file)
# 删除目标 sheet，后续用更新后的数据替换
if target_sheet in wb.sheetnames:
    ws_to_remove = wb[target_sheet]
    wb.remove(ws_to_remove)
# 利用 ExcelWriter（基于 openpyxl）写入更新后的目标 sheet
with pd.ExcelWriter(original_file, engine='openpyxl') as writer:
    writer.book = wb
    df_final.to_excel(writer, sheet_name=target_sheet, index=False)
    writer.save()
# 保存为更新后的文件（这里另存为 updated_file）
wb.save(updated_file)

# =============================
# 8. 高亮显示因 Jira 更新而修改的单元格
# =============================
wb_updated = load_workbook(updated_file)
ws = wb_updated[target_sheet]
# 设定高亮填充色为黄色
highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
# Excel 第一行为表头，数据从第二行开始，所以 pandas 索引0对应 Excel行2
header = [cell.value for cell in ws[1]]
for (pandas_row, col_name) in updated_cells:
    # Excel 行号 = pandas行索引 + 2（考虑表头）
    excel_row = pandas_row + 2
    if col_name in header:
        excel_col = header.index(col_name) + 1  # openpyxl列号从1开始
        ws.cell(row=excel_row, column=excel_col).fill = highlight_fill
wb_updated.save(updated_file)

print(f"更新完成，已生成 {updated_file}。在工作表 '{target_sheet}' 中已根据 Jira 数据更新了 Name、Owner、Defect finder 字段，并在 \"Octane or Jira\" 列中加入了 'jira' 标识，新更新的单元格均已高亮显示。")




import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============================
# 1. 参数设置
# ============================
original_file = "original.xlsx"          # 包含多个 sheet 的原始 Excel
jira_file = "jira.csv"                   # Jira 的 CSV 文件
target_sheet = "Octane and jira"         # 目标工作表
updated_file = "updated_excel.xlsx"      # 更新后保存的文件

# ============================
# 2. 读取原始总表数据（只处理目标 sheet）
# ============================
df_original = pd.read_excel(original_file, sheet_name=target_sheet)

# ============================
# 3. 读取 Jira CSV 文件，并将字段映射为总表对应的字段名称
# ============================
df_jira = pd.read_csv(jira_file)

# 定义映射：Jira 字段 -> 原总表字段
rename_dict = {
    "Summary": "Name",
    "issue key": "Ticket no. supplier",
    "Assignee": "Owner",
    "Reporter": "Defect finder"
}
df_jira.rename(columns=rename_dict, inplace=True)

# 只保留用于更新的字段；这里我们依赖 "Ticket no. supplier" 来关联其他字段
update_fields = ["Name", "Owner", "Defect finder"]
df_jira = df_jira[["Ticket no. supplier"] + update_fields]

# ============================
# 4. 合并原总表和 Jira 数据
# ============================
# 使用 "Ticket no. supplier" 作为关联键进行左合并，生成新列后缀 "_jira"
df_merged = pd.merge(df_original, df_jira, on="Ticket no. supplier", how="left", suffixes=("", "_jira"))

# ============================
# 5. 根据条件更新：仅当原总表中对应字段为空时，从 Jira 数据更新
# ============================
# 用于记录更新的单元格（用于后续高亮显示），格式为：(pandas 行索引, 字段名称)
updated_cells = []

for idx, row in df_merged.iterrows():
    for field in update_fields:
        jira_val = row.get(f"{field}_jira")  # 来自 Jira 的值
        orig_val = row.get(field)            # 原总表中的值
        # 更新条件：如果原表字段为空（包括 NaN 或仅空白字符串）且 Jira 有非空数据，则更新
        if ((pd.isnull(orig_val)) or (str(orig_val).strip() == "")) \
           and pd.notnull(jira_val) and str(jira_val).strip() != "":
            df_merged.at[idx, field] = jira_val
            updated_cells.append((idx, field))

# ============================
# 6. 删除合并后多出来的 *_jira 辅助列，保留原总表列结构
# ============================
df_final = df_merged[df_original.columns]

# ============================
# 7. 将更新后的数据写入 Excel，保持其它 sheet 不变
# ============================
# 利用 openpyxl 载入整个工作簿
wb = load_workbook(original_file)

# 删除目标工作表以便重写数据
if target_sheet in wb.sheetnames:
    ws_to_remove = wb[target_sheet]
    wb.remove(ws_to_remove)

# 使用 pandas 的 ExcelWriter 写入更新后的目标 sheet（保留其他 sheet 不变）
with pd.ExcelWriter(original_file, engine='openpyxl') as writer:
    writer.book = wb
    df_final.to_excel(writer, sheet_name=target_sheet, index=False)
    writer.save()

# 另存一份更新后的文件
wb.save(updated_file)

# ============================
# 8. 高亮显示因 Jira 更新而填充的单元格（绿色高亮）
# ============================
wb_updated = load_workbook(updated_file)
ws = wb_updated[target_sheet]

# 定义绿色高亮填充
highlight_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

# 读取表头，Excel 第一行是表头；pandas 的行号与 Excel 行号存在偏移（pandas 行索引 0 对应 Excel 第2行）
header = [cell.value for cell in ws[1]]
for (pandas_row, col_name) in updated_cells:
    excel_row = pandas_row + 2  # 加 2：1 为表头，pandas 行索引从 0 开始
    if col_name in header:
        excel_col = header.index(col_name) + 1  # openpyxl 列号从 1 开始
        ws.cell(row=excel_row, column=excel_col).fill = highlight_fill

wb_updated.save(updated_file)
print(f"更新完成，已生成 {updated_file}。只有原表中为空的字段被 Jira 数据补充，并以绿色高亮显示。")
