import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# === 辅助函数 ===
def is_row_blank(ws, row):
    for c in range(1, ws.max_column+1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True

def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break

def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            return r
    return 1

# === 0. 读取配置 ===
with open("unified_config2.json", "r", encoding="utf-8") as f:
    cfg = json.load(f)

paths    = cfg["paths"]
orig_fp  = paths["original_file"]
new_fp   = paths["new_file"]
upd_fp   = paths["updated_file"]
sheet    = cfg["sheet"]["target_sheet"]
sources  = cfg["sources"]
fund_map = cfg["fund_function_mapping"]
own_map  = cfg["owner_root_cause_mapping"]
clear_flag = cfg.get("settings",{}).get("clear_old_highlight","N").upper()


# === 1. 判断数据来源 (Octane/Jira) 并加载映射 ===
base       = os.path.basename(new_fp)
source_key = next((k for k,v in sources.items() if v["pattern"] in base), None)
if not source_key:
    raise RuntimeError("无法识别 new_file 来源 (Octane 或 Jira)")

src_cfg    = sources[source_key]
read_meth  = src_cfg["read_method"]
date_col   = src_cfg.get("date_col")
mapping    = src_cfg["mapping"]

# === 2. 读取新表 ===
df_new = pd.read_excel(new_fp) if read_meth=="excel" else pd.read_csv(new_fp)

# === 3. 打开原表，清除旧的紫/蓝高亮 & 清空尾行 ===
wb = load_workbook(orig_fp)
ws = wb[sheet]

# 要清除的颜色
if clear_flag == "Y":
    purple_codes = {"FF800080", "800080","00800080"}  # 新增用紫色
    blue_codes   = {"FFADD8E6", "ADD8E6","00ADD8E6"}  # 更新用蓝色

    for row in ws.iter_rows(min_row=2, min_col=1,
                            max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            f = cell.fill
            if f.fill_type == "solid":
                col = f.fgColor.rgb or f.fgColor.value
                if col in purple_codes or col in blue_codes:
                    cell.fill = PatternFill(fill_type=None)

trim_trailing_blank_rows(ws)

# 构建表头映射
header2col = {ws.cell(1,c).value:c for c in range(1, ws.max_column+1)}
headers    = list(header2col.keys())
id_col     = header2col["ID"]

# 原表已有 ID→行号
id2row = {
    ws.cell(r,id_col).value:r
    for r in range(2, ws.max_row+1)
    if ws.cell(r,id_col).value
}

# 如果新表是 Excel，还要提取 ID→超链接
id2url_new = {}
if read_meth=="excel":
    nwb = load_workbook(new_fp)
    nws = nwb.active
    col_id = next((c for c in range(1,nws.max_column+1)
                   if nws.cell(1,c).value=="ID"), None)
    if col_id:
        for r in range(2, nws.max_row+1):
            nc = nws.cell(r,col_id)
            if nc.hyperlink:
                id2url_new[nc.value] = nc.hyperlink.target

# 定位追加起始
last_row      = find_last_data_row(ws, id_col)
original_last = last_row

# 高亮样式：更新=蓝色，新增=紫色
blue_fill   = PatternFill("solid", fgColor="ADD8E6")
purple_fill = PatternFill("solid", fgColor="800080")

# 找出映射中“ID”对应的新表列名
id_key = next(k for k,v in mapping.items() if v=="ID")

# === 4. 遍历新表，更新或追加 ===
for _, new_row in df_new.iterrows():
    new_id = new_row.get(id_key, "")
    if pd.isna(new_id) or not new_id:
        continue

    if new_id in id2row:
        # —— 更新已有行，只更新除 Function 之外的列，并标蓝 —— 
        r = id2row[new_id]

        # 更新日期 (Jira)
        if date_col:
            raw = new_row.get(date_col,"")
            if pd.notna(raw) and raw!="":
                try:
                    ts = pd.to_datetime(raw)
                    dt = ts.to_pydatetime()
                    c  = header2col[mapping[date_col]]
                    cell = ws.cell(r,c)
                    if cell.value != dt:
                        cell.value = dt
                        cell.number_format = "m/d/yyyy h:mm:ss AM/PM"
                        cell.fill = blue_fill
                except:
                    pass

        # 更新其它映射字段 (跳过 Function)
        for nk, ok in mapping.items():
            if nk == date_col or ok == "Function":
                continue
            val = new_row.get(nk,"")
            if pd.isna(val) or val=="": 
                continue
            if ok=="Target I-Step:" and (str(val).startswith("G070") or str(val).startswith("U006")):
                val = "NA05"+str(val)[4:]
            c = header2col.get(ok)
            if c:
                cell = ws.cell(r,c)
                if cell.value != val:
                    cell.value = val
                    cell.fill  = blue_fill

        # 更新 Root cause
        if "Owner" in header2col and "Root cause" in header2col:
            ow = ws.cell(r,header2col["Owner"]).value or ""
            for k, v in own_map.items():
                if k in ow:
                    c = header2col["Root cause"]
                    cell = ws.cell(r,c)
                    if cell.value != v:
                        cell.value = v
                        cell.fill  = blue_fill
                    break

    else:
        # —— 追加新行，并标紫 —— 
        last_row += 1
        for idx, hdr in enumerate(headers, start=1):
            val = ""
            for nk, ok in mapping.items():
                if ok == hdr:
                    tmp = new_row.get(nk,"")
                    if pd.notna(tmp) and tmp!="":
                        if hdr == mapping.get(date_col):
                            try:
                                ts = pd.to_datetime(tmp)
                                val = ts.to_pydatetime()
                            except:
                                val = tmp
                        else:
                            val = tmp
                        if ok=="Involved I-Step:" and (str(val).startswith("G070") or str(val).startswith("U006")):
                            val = "NA05"+str(val)[4:]
                    break
            cell = ws.cell(last_row, idx)
            cell.value = val
            if val != "":
                cell.fill = purple_fill
                if hdr == mapping.get(date_col):
                    cell.number_format = "m/d/yyyy h:mm:ss AM/PM"

        # 新行 ID 超链接
        url = id2url_new.get(new_id)
        if url:
            cell = ws.cell(last_row, id_col)
            cell.hyperlink = url
            cell.style     = "Hyperlink"

        # 新行 Function 列（标紫）
        if "Found in function" in header2col and "Function" in header2col:
            fv = new_row.get("Found in function","") or ""
            for k, v in fund_map.items():
                if k in fv:
                    c = header2col["Function"]
                    ws.cell(last_row,c).value = v
                    ws.cell(last_row,c).fill  = purple_fill
                    break

        # 新行 Root cause
        if "Owner" in header2col and "Root cause" in header2col:
            ow = new_row.get("Owner","") or ""
            for k, v in own_map.items():
                if k in ow:
                    c = header2col["Root cause"]
                    ws.cell(last_row,c).value = v
                    ws.cell(last_row,c).fill  = purple_fill
                    break

# === 5. 填公式 & 新增行标记 Octane/Jira ===
max_row = ws.max_row

# Days 列
d_idx  = header2col.get("Days")
ct_idx = header2col.get("Creation time")
if d_idx and ct_idx:
    colL = get_column_letter(ct_idx)
    for r in range(2, max_row+1):
        ws.cell(r, d_idx).value = f'=DATEDIF(${colL}{r},TODAY(),"D")'

# Octane or Jira 列（仅对新增行填充）
oij_idx = header2col.get("Octane or Jira")
if oij_idx:
    for r in range(original_last+1, max_row+1):
        ws.cell(r, oij_idx).value = source_key

# Open > 20 days 列
o20 = header2col.get("Open >20 days") or header2col.get("Open > 20 days")
if o20 and d_idx:
    cd = get_column_letter(d_idx)
    for r in range(2, max_row+1):
        ws.cell(r, o20).value = f'=IF({cd}{r}>20,1,0)'

# No TIS 列
nt = header2col.get("No TIS")
pl = header2col.get("Planned closing version")
ti = header2col.get("Target I-Step:")
if nt and pl and ti:
    pL = get_column_letter(pl)
    tL = get_column_letter(ti)
    for r in range(2, max_row+1):
        ws.cell(r, nt).value = f'=IF(OR({pL}{r}<>"",{tL}{r}<>""),0,1)'

# === 6. 保存 ===
wb.save(upd_fp)
print(f"更新完成 (来源={source_key})，紫色=新增，蓝色=更新。结果保存在 '{upd_fp}'。")
