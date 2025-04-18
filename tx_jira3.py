import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# === 辅助函数（同前） ===
def is_row_blank(ws, row):
    for c in range(1, ws.max_column+1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True

def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break

def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            return r
    return 1

# === 0. 加载配置 ===
with open("config.json", "r", encoding="utf-8") as f:
    cfg = json.load(f)

original_file            = cfg["paths"]["original_file"]
new_file                 = cfg["paths"]["new_file"]
updated_file             = cfg["paths"]["updated_file"]
target_sheet             = cfg["sheet"]["target_sheet"]
excel_mapping            = cfg["excel_mapping"]
jira_mapping             = cfg["jira_mapping"]
fund_function_mapping    = cfg["fund_function_mapping"]
owner_root_cause_mapping = cfg["owner_root_cause_mapping"]

# === 1. 根据文件名选择映射 & 读取新表 ===
basename = os.path.basename(new_file).lower()
if basename.endswith(".csv") and "jira" in basename:
    mapping = jira_mapping
    df_new  = pd.read_csv(new_file)
else:
    mapping = excel_mapping
    df_new  = pd.read_excel(new_file)

# === 2. 打开原表并清理空行 ===
wb = load_workbook(original_file)
ws = wb[target_sheet]
trim_trailing_blank_rows(ws)

# 构建表头→列号映射
header2col = { ws.cell(1, c).value: c for c in range(1, ws.max_column+1) }

# 提取原表 ID→行号 & 超链接
id_col = header2col["ID"]
id2row = {}
id2url = {}
for r in range(2, ws.max_row+1):
    cell = ws.cell(r, id_col)
    if cell.value is not None:
        id2row[cell.value] = r
        if cell.hyperlink:
            id2url[cell.value] = cell.hyperlink.target

# 如果是 Excel，新表超链接也提取
new_id2url = {}
if not (basename.endswith(".csv") and "jira" in basename):
    new_wb  = load_workbook(new_file)
    new_ws  = new_wb.active
    id_col_new = next((c for c in range(1, new_ws.max_column+1)
                       if new_ws.cell(1, c).value=="ID"), None)
    if id_col_new:
        for r in range(2, new_ws.max_row+1):
            c = new_ws.cell(r, id_col_new)
            if c.hyperlink:
                new_id2url[c.value] = c.hyperlink.target

# 样式 & 表头
green_fill  = PatternFill("solid", fgColor="00FF00")
yellow_fill = PatternFill("solid", fgColor="FFFF00")
headers     = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
last_row    = find_last_data_row(ws, id_col)

# === 3. 遍历新表，更新或追加 ===
for _, new_row in df_new.iterrows():
    new_id = new_row.get("ID", "")
    if not new_id:
        continue

    if new_id in id2row:
        # —— 更新现有行，只在差异处染绿 —— 
        r = id2row[new_id]
        # 超链接
        if new_id in new_id2url:
            cell = ws.cell(r, id_col)
            cell.hyperlink = new_id2url[new_id]
            cell.style     = "Hyperlink"
        # 映射字段
        for nk, ok in mapping.items():
            if nk not in new_row:
                continue
            new_val = new_row[nk]
            if pd.isna(new_val) or new_val=="":
                continue
            # 特殊：Involved I-Step 前缀
            if ok=="Involved I-Step":
                s = str(new_val)
                if s.startswith(("G070","U006")):
                    new_val = "NA05"+s[4:]
            c = header2col.get(ok)
            if not c:
                continue
            cell = ws.cell(r, c)
            if cell.value!=new_val:
                cell.value = new_val
                cell.fill  = green_fill
        # Function 列
        if "Found in function" in header2col and "Function" in header2col:
            fv = ws.cell(r, header2col["Found in function"]).value or ""
            for k, v in fund_function_mapping.items():
                if k in fv:
                    c = header2col["Function"]
                    cell = ws.cell(r, c)
                    if cell.value!=v:
                        cell.value = v
                        cell.fill  = green_fill
                    break
        # Root cause
        if "Owner" in header2col and "Root cause" in header2col:
            ow = ws.cell(r, header2col["Owner"]).value or ""
            for k, v in owner_root_cause_mapping.items():
                if k in ow:
                    c = header2col["Root cause"]
                    cell = ws.cell(r, c)
                    if cell.value!=v:
                        cell.value = v
                        cell.fill  = green_fill
                    break

    else:
        # —— 添加新行，只对非空处染黄 —— 
        last_row += 1
        for idx, hdr in enumerate(headers, start=1):
            val = ""
            for nk, ok in mapping.items():
                if ok==hdr:
                    tmp = new_row.get(nk, "")
                    if pd.notna(tmp) and tmp!="":
                        if ok=="Involved I-Step":
                            s = str(tmp)
                            if s.startswith(("G070","U006")):
                                tmp = "NA05"+s[4:]
                        val = tmp
                    break
            cell = ws.cell(last_row, idx)
            cell.value = val
            if val!="":
                cell.fill = yellow_fill
        # ID 超链接
        if new_id in new_id2url:
            cell = ws.cell(last_row, id_col)
            cell.hyperlink = new_id2url[new_id]
            cell.style     = "Hyperlink"
        # Function & Root cause
        if "Found in function" in header2col and "Function" in header2col:
            fv = new_row.get("Found in function","") or ""
            for k, v in fund_function_mapping.items():
                if k in fv:
                    c = header2col["Function"]
                    ws.cell(last_row, c).value = v
                    ws.cell(last_row, c).fill  = yellow_fill
                    break
        if "Owner" in header2col and "Root cause" in header2col:
            ow = new_row.get("Owner","") or ""
            for k, v in owner_root_cause_mapping.items():
                if k in ow:
                    c = header2col["Root cause"]
                    ws.cell(last_row, c).value = v
                    ws.cell(last_row, c).fill  = yellow_fill
                    break

# === 4. 填充公式等 ===
max_row = ws.max_row

# Days
d_i, c_i = header2col.get("Days"), header2col.get("Creation time")
if d_i and c_i:
    col_letter = get_column_letter(c_i)
    for r in range(2, max_row+1):
        ws.cell(r, d_i).value = f'=DATEDIF(${col_letter}{r},TODAY(),"D")'

# Octane or Jira
oir_i = header2col.get("Octane or Jira")
if oir_i:
    v = "Octane" if "octane" in basename else ("Jira" if "jira" in basename else "")
    for r in range(2, max_row+1):
        ws.cell(r, oir_i).value = v

# Open >20 days
o20 = header2col.get("Open >20 days") or header2col.get("Open > 20 days")
if o20 and d_i:
    dl = get_column_letter(d_i)
    for r in range(2, max_row+1):
        ws.cell(r, o20).value = f'=IF({dl}{r}>20,1,0)'

# No TIS
nt_i = header2col.get("No TIS")
pl_i = header2col.get("Planned closing version")
ti_i = header2col.get("Target I-Step:")
if nt_i and pl_i and ti_i:
    pl_l, ti_l = get_column_letter(pl_i), get_column_letter(ti_i)
    for r in range(2, max_row+1):
        ws.cell(r, nt_i).value = f'=IF(OR({pl_l}{r}<>"",{ti_l}{r}<>""),0,1)'

# === 5. 保存 ===
wb.save(updated_file)
print("更新完成，支持 Excel & Jira CSV，映射关系全放 config.json。")
