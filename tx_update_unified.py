import os, json, pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# === 辅助函数 ===
def is_row_blank(ws, row):
    for c in range(1, ws.max_column+1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True

def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break

def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, key_col).value not in (None, ""):
            return r
    return 1

# === 0. 读取配置 ===
with open("config.json", "r", encoding="utf-8") as f:
    cfg = json.load(f)

orig_path   = cfg["paths"]["original_file"]
upd_path    = cfg["paths"]["updated_file"]
sheet_name  = cfg["sheet"]["target_sheet"]
sources     = cfg["sources"]
fund_map    = cfg["fund_function_mapping"]
owner_map   = cfg["owner_root_cause_mapping"]

# === 1. 根据 new_file 判断来源并加载 mapping ===
new_file = cfg["paths"].get("new_file", "")  # 如果从外部传参，可改这里
base = os.path.basename(new_file)
source_key = next((k for k,v in sources.items() if v["pattern"] in base), None)
if not source_key:
    raise RuntimeError("无法识别 new_file 中的来源 (Octane/Jira).")

src_cfg   = sources[source_key]
mapping   = src_cfg["mapping"]
read_meth = src_cfg["read_method"]
date_col  = src_cfg.get("date_col")  # 如 Jira 有 Created 列

# === 2. 读取新表 ===
if read_meth == "excel":
    df_new = pd.read_excel(new_file)
else:
    df_new = pd.read_csv(new_file)

# === 3. 打开原表并预处理 ===
wb = load_workbook(orig_path)
ws = wb[sheet_name]
trim_trailing_blank_rows(ws)

# 构建表头→列号与列序
header2col = { ws.cell(1,c).value: c for c in range(1, ws.max_column+1) }
headers    = list(header2col.keys())
id_col     = header2col["ID"]

# 记录原表已有 ID→行号
id2row = {
    ws.cell(r,id_col).value: r
    for r in range(2, ws.max_row+1)
    if ws.cell(r,id_col).value
}

last_row      = find_last_data_row(ws, id_col)
original_last = last_row

# 高亮样式
green = PatternFill("solid", fgColor="00FF00")
yellow= PatternFill("solid", fgColor="FFFF00")

# === 4. 对每条新数据：更新 or 追加 ===
for _, new_row in df_new.iterrows():
    new_id = new_row.get(next(k for k in mapping if mapping[k]=="ID"), "")
    if pd.isna(new_id) or not new_id:
        continue

    if new_id in id2row:
        # —— 更新已有行 —— 
        r = id2row[new_id]
        # ① 日期列 (只对 Jira 有 Created 转 datetime)
        if date_col:
            raw = new_row.get(date_col,"")
            if pd.notna(raw) and raw!="":
                try:
                    ts = pd.to_datetime(raw)
                    dt = ts.to_pydatetime()
                    c  = header2col[mapping[date_col]]
                    cell = ws.cell(r,c)
                    if cell.value!=dt:
                        cell.value = dt
                        cell.number_format = "m/d/yyyy h:mm:ss AM/PM"
                        cell.fill = green
                except:
                    pass
        # ② 其它映射字段
        for nk, ok in mapping.items():
            if nk==date_col: continue
            val = new_row.get(nk,"")
            if pd.isna(val) or val=="": continue
            # 特殊前缀替换
            if ok=="Target I-Step:" and (str(val).startswith("G070") or str(val).startswith("U006")):
                val = "NA05"+str(val)[4:]
            c = header2col.get(ok)
            cell = ws.cell(r,c)
            if cell.value!=val:
                cell.value=val
                cell.fill = green
        # ③ Function 列
        if "Found in function" in header2col and "Function" in header2col:
            fv = ws.cell(r, header2col["Found in function"]).value or ""
            for k,v in fund_map.items():
                if k in fv:
                    c=header2col["Function"]
                    cell=ws.cell(r,c)
                    if cell.value!=v:
                        cell.value=v; cell.fill=green
                    break
        # ④ Root cause 列
        if "Owner" in header2col and "Root cause" in header2col:
            ow=ws.cell(r,header2col["Owner"]).value or ""
            for k,v in owner_map.items():
                if k in ow:
                    c=header2col["Root cause"]
                    cell=ws.cell(r,c)
                    if cell.value!=v:
                        cell.value=v; cell.fill=green
                    break

    else:
        # —— 追加新行 —— 
        last_row += 1
        for idx,hdr in enumerate(headers, start=1):
            val = ""
            for nk,ok in mapping.items():
                if ok==hdr:
                    tmp = new_row.get(nk,"")
                    if pd.notna(tmp) and tmp!="":
                        # 日期列
                        if hdr==mapping.get(date_col):
                            try:
                                ts= pd.to_datetime(tmp)
                                val=ts.to_pydatetime()
                            except:
                                val=tmp
                        else:
                            val=tmp
                        # 前缀替换
                        if ok=="Target I-Step:" and (str(val).startswith("G070") or str(val).startswith("U006")):
                            val="NA05"+str(val)[4:]
                    break
            cell = ws.cell(last_row, idx)
            cell.value=val
            if val!="":
                cell.fill=yellow
                if hdr==mapping.get(date_col):
                    cell.number_format="m/d/yyyy h:mm:ss AM/PM"
        # Function & Root cause
        if "Found in function" in header2col and "Function" in header2col:
            fv=new_row.get("Found in function","") or ""
            for k,v in fund_map.items():
                if k in fv:
                    c=header2col["Function"]
                    ws.cell(last_row,c).value=v
                    ws.cell(last_row,c).fill=yellow
                    break
        if "Owner" in header2col and "Root cause" in header2col:
            ow=new_row.get("Owner","") or ""
            for k,v in owner_map.items():
                if k in ow:
                    c=header2col["Root cause"]
                    ws.cell(last_row,c).value=v
                    ws.cell(last_row,c).fill=yellow
                    break

# === 5. 填公式 & Octane/Jira 标签等 ===
max_row = ws.max_row

# 5.1 Days
d_idx = header2col.get("Days")
ct_idx= header2col.get("Creation time")
if d_idx and ct_idx:
    colL = get_column_letter(ct_idx)
    for r in range(2, max_row+1):
        ws.cell(r,d_idx).value = f'=DATEDIF(${colL}{r},TODAY(),"D")'

# 5.2 Octane or Jira
oij_idx = header2col.get("Octane or Jira")
if oij_idx:
    tag = source_key  # 要么 "Octane" 要么 "Jira"
    for r in range(original_last+1, max_row+1):
        ws.cell(r,oij_idx).value = tag

# 5.3 Open>20 days
o20 = header2col.get("Open >20 days") or header2col.get("Open > 20 days")
if o20 and d_idx:
    colD=get_column_letter(d_idx)
    for r in range(2, max_row+1):
        ws.cell(r,o20).value = f'=IF({colD}{r}>20,1,0)'

# 5.4 No TIS
nt=header2col.get("No TIS"); pl=header2col.get("Planned closing version"); ti=header2col.get("Target I-Step:")
if nt and pl and ti:
    plL=get_column_letter(pl); tiL=get_column_letter(ti)
    for r in range(2, max_row+1):
        ws.cell(r,nt).value = f'=IF(OR({plL}{r}<>"",{tiL}{r}<>""),0,1)'

# === 6. 保存 ===
wb.save(upd_path)
print(f"完成：来源={source_key}，结果已存于 {upd_path}")
