import os
import time
import pandas as pd
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ===============================
# 1. 文件路径及参数设置
# ===============================
# 原始 Excel 模板，包含多个 sheet，其中目标 sheet 为 "Octane and jira"
original_file = "original.xlsx"
# 最终更新后的结果保存到
updated_file = "updated_excel.xlsx"
# 目标更新的工作表名称
target_sheet = "Octane and jira"
# 定义映射关系（新文件中列名 -> 原表中对应的列名）
mapping = {
    "Name": "名字",
    "age": "年龄",
    "sex": "性别"
}
# 监控文件夹（例如 OneDrive 同步的文件夹），建议使用绝对路径
folder_to_watch = r"C:\Path\To\Your\WatchFolder"  # 请替换成实际监控文件夹路径

# 为避免更新原始 Excel 自身，设定一个不处理的文件名列表
ignored_files = { os.path.basename(original_file), os.path.basename(updated_file) }

# ===============================
# 2. 定义更新逻辑
# ===============================
def update_excel(new_file_path):
    """
    更新函数，读取传入路径的新表数据，然后把新数据按照映射关系追加到原始 Excel 文件中的目标工作表，
    并在更新后对新追加的行以黄色高亮显示，结果保存为 updated_file。
    """
    try:
        print(f"开始更新文件: {new_file_path}")
        # -------------------------------
        # (1) 读取原始 Excel 中目标 sheet 的数据（保留原有数据）
        # -------------------------------
        df_orig = pd.read_excel(original_file, sheet_name=target_sheet)
        
        # -------------------------------
        # (2) 读取新表数据，使用传入的新文件路径
        # -------------------------------
        df_new = pd.read_excel(new_file_path)
        
        # 获取目标 sheet 中所有列（原始表结构）
        orig_columns = list(df_orig.columns)
        
        # 构造新行数据列表，每一行为字典，key 为原表列名
        new_rows = []
        for idx, new_row in df_new.iterrows():
            row_data = {}
            for col in orig_columns:
                matched = False
                # 若原表中的列在映射关系的 value 中，则取新表中对应的列值
                for new_key, orig_key in mapping.items():
                    if col == orig_key:  # 当前列属于映射范围
                        value = new_row.get(new_key, "")
                        # 如果值为 NaN，则置为空字符串
                        if pd.isna(value):
                            value = ""
                        row_data[col] = value
                        matched = True
                        break
                if not matched:
                    # 原表中该列不在映射关系中，默认填入空值
                    row_data[col] = ""
            new_rows.append(row_data)
            
        # 生成与原表结构一致的新行 DataFrame
        df_new_rows = pd.DataFrame(new_rows, columns=orig_columns)
        
        # -------------------------------
        # (3) 将新数据追加到目标 sheet 的末尾（原有数据保持不变）
        # -------------------------------
        df_updated_target = pd.concat([df_orig, df_new_rows], ignore_index=True)
        
        # -------------------------------
        # (4) 保留原始 Excel 中其他 sheet，不更改原有数据，仅更新目标 sheet
        # -------------------------------
        wb = load_workbook(original_file)
        # 删除原始 Excel 中的目标 sheet（后续将用更新后的数据重写）
        if target_sheet in wb.sheetnames:
            ws_target = wb[target_sheet]
            wb.remove(ws_target)
        
        # 写入更新后的目标 sheet到新的工作簿中，同时保留其它 sheet不变
        with pd.ExcelWriter(updated_file, engine="openpyxl") as writer:
            writer.book = wb
            writer.sheets = {ws.title: ws for ws in wb.worksheets}
            df_updated_target.to_excel(writer, sheet_name=target_sheet, index=False)
            writer.save()
        
        # -------------------------------
        # (5) 对新追加的行进行高亮显示（黄色）
        # -------------------------------
        wb_updated = load_workbook(updated_file)
        ws = wb_updated[target_sheet]
        
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # Excel 第一行为表头，因此数据行从第2行开始
        original_rows = df_orig.shape[0]  # 原有数据行数（不包括表头）
        start_row = original_rows + 2      # 新数据起始行号
        
        for row in range(start_row, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = highlight_fill
                
        wb_updated.save(updated_file)
        print(f"更新完成！新数据已追加到工作表 '{target_sheet}' 的末尾，结果保存在 {updated_file}")
    except Exception as e:
        print("更新过程中发生错误:", e)

# ===============================
# 3. 定义 watchdog 事件处理器，监控整个文件夹中所有 Excel 文件
# ===============================
class ExcelFolderHandler(FileSystemEventHandler):
    def on_created(self, event):
        if not event.is_directory and event.src_path.endswith(".xlsx"):
            base = os.path.basename(event.src_path)
            # 忽略原始文件及更新后文件
            if base in ignored_files:
                return
            print(f"检测到新文件: {event.src_path}")
            update_excel(event.src_path)
            
    def on_modified(self, event):
        if not event.is_directory and event.src_path.endswith(".xlsx"):
            base = os.path.basename(event.src_path)
            if base in ignored_files:
                return
            print(f"检测到文件修改: {event.src_path}")
            update_excel(event.src_path)

# ===============================
# 4. 启动文件夹监控
# ===============================
if __name__ == "__main__":
    # 使用绝对路径的 folder_to_watch（确保该目录包含要监控的 Excel 文件）
    folder_to_watch = os.path.abspath(folder_to_watch)
    if not os.path.exists(folder_to_watch):
        print("监控文件夹不存在，请检查路径。")
        exit(1)
    
    event_handler = ExcelFolderHandler()
    observer = Observer()
    observer.schedule(event_handler, path=folder_to_watch, recursive=False)
    observer.start()
    print(f"开始监控文件夹: {folder_to_watch} 中的所有 Excel 文件...")
    
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()
