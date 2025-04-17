import os, json, pandas as pd
from openpyxl import load_workbook, load_workbook as load_wb_readonly
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# === 加载配置 ===
with open("config.json", "r", encoding="utf-8") as f:
    cfg = json.load(f)

# 文件路径
original_file = cfg["paths"]["original_file"]
new_file      = cfg["paths"]["new_file"]
updated_file  = cfg["paths"]["updated_file"]

# Sheet 名称
target_sheet  = cfg["sheet"]["target_sheet"]

# 各种映射
mapping                  = cfg["column_mapping"]
fund_function_mapping    = cfg["fund_function_mapping"]
owner_root_cause_mapping = cfg["owner_root_cause_mapping"]

# -------------------------------
# 1. 读取原表与新表
# -------------------------------
df_orig = pd.read_excel(original_file, sheet_name=target_sheet)
df_new  = pd.read_excel(new_file)
orig_columns = list(df_orig.columns)

# -------------------------------
# 2. 从 new_file 中提取 ID->URL 的映射
# -------------------------------
new_wb = load_wb_readonly(new_file, read_only=True, data_only=True)
new_ws = new_wb.active  # 如果要指定 sheet: new_ws = new_wb["SheetName"]

# 找到新表中“ID”列的列号
id_col_new = None
for c in range(1, new_ws.max_column + 1):
    if new_ws.cell(row=1, column=c).value == "ID":
        id_col_new = c
        break

# 构建映射字典
id2url = {}
if id_col_new:
    for r in range(2, new_ws.max_row + 1):
        cell = new_ws.cell(row=r, column=id_col_new)
        if cell.hyperlink:
            id2url[cell.value] = cell.hyperlink.target

# -------------------------------
# 3. 构造新行数据
# -------------------------------
new_rows = []
for _, new_row in df_new.iterrows():
    rd = {}
    for col in orig_columns:
        for nk, ok in mapping.items():
            if col == ok:
                v = new_row.get(nk, "")
                if pd.isna(v): v = ""
                rd[col] = v
                break
        else:
            rd[col] = ""
    new_rows.append(rd)

df_new_rows = pd.DataFrame(new_rows, columns=orig_columns)

# -------------------------------
# 4. 追加并写入
# -------------------------------
df_updated = pd.concat([df_orig, df_new_rows], ignore_index=True)

wb = load_workbook(original_file)
if target_sheet in wb.sheetnames:
    wb.remove(wb[target_sheet])

with pd.ExcelWriter(updated_file, engine="openpyxl") as writer:
    writer.book = wb
    writer.sheets = {ws.title: ws for ws in wb.worksheets}
    df_updated.to_excel(writer, sheet_name=target_sheet, index=False)
    writer.save()

# -------------------------------
# 5. 高亮 & 公式 & 其它列填充 & 超链接重建
# -------------------------------
wb2 = load_workbook(updated_file)
ws2 = wb2[target_sheet]

# 原有数据行数
orig_n = df_orig.shape[0]
start_row = orig_n + 2

# 高亮新行
hl = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
for r in range(start_row, ws2.max_row + 1):
    for c in range(1, ws2.max_column + 1):
        ws2.cell(row=r, column=c).fill = hl

# 设置 Days 公式
days_idx = next((c for c in range(1, ws2.max_column+1)
                 if ws2.cell(row=1, column=c).value=="Days"), None)
if days_idx:
    for r in range(2, ws2.max_row + 1):
        ws2.cell(row=r, column=days_idx).value = f'=DATEDIF($J{r},TODAY(),"D")'

# 填 Octane or Jira
oir_idx = next((c for c in range(1, ws2.max_column+1)
                 if ws2.cell(row=1, column=c).value=="Octane or Jira"), None)
if oir_idx:
    fn = os.path.basename(new_file)
    val = "Octane" if "Octane" in fn else ("Jira" if "Jira" in fn else "")
    for r in range(start_row, ws2.max_row + 1):
        ws2.cell(row=r, column=oir_idx).value = val

# 填 Open > 20 days
open20_idx = next((c for c in range(1, ws2.max_column+1)
                   if ws2.cell(row=1, column=c).value in ("Open >20 days","Open > 20 days")), None)
if open20_idx and days_idx:
    dl = get_column_letter(days_idx)
    for r in range(start_row, ws2.max_row + 1):
        ws2.cell(row=r, column=open20_idx).value = f'=IF({dl}{r}>20,1,0)'

# 填 No TIS
nt_idx = pl_idx = ti_idx = None
for c in range(1, ws2.max_column+1):
    h = ws2.cell(row=1, column=c).value
    if h=="No TIS":       nt_idx = c
    if h=="Planned closing version": pl_idx = c
    if h=="Target I-Step:":         ti_idx = c
if nt_idx and pl_idx and ti_idx:
    pl_l = get_column_letter(pl_idx)
    ti_l = get_column_letter(ti_idx)
    for r in range(start_row, ws2.max_row + 1):
        ws2.cell(row=r, column=nt_idx).value = f'=IF(OR({pl_l}{r}<>"",{ti_l}{r}<>""),0,1)'

# —— **在这里进行“重建超链接”** ——  
# 找 ID 列索引
id_idx = next((c for c in range(1, ws2.max_column+1)
               if ws2.cell(row=1, column=c).value=="ID"), None)
if id_idx:
    for r in range(start_row, ws2.max_row + 1):
        cell = ws2.cell(row=r, column=id_idx)
        url  = id2url.get(cell.value)
        if url:
            cell.hyperlink = url
            cell.style     = "Hyperlink"

wb2.save(updated_file)
print("全部更新完成，含超链接重建，结果保存在", updated_file)
