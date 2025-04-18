from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

def is_row_blank(ws, row):
    """判断指定行是否所有单元格都为空（或None）"""
    for c in range(1, ws.max_column + 1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True

def trim_trailing_blank_rows(ws):
    """
    从表尾开始，删除所有“完全空白”的行，
    直到遇到第一行非空行为止。
    """
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break

def find_last_data_row(ws, key_col):
    """
    扫描 key_col 列，返回最后一个非空单元格所在的行号，
    如果整列都空，则返回 1（表头行）。
    """
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            return r
    return 1

# —— 在脚本中使用 —— #

# 先加载工作簿和目标 sheet
wb = load_workbook(original_file)
ws = wb[target_sheet]

# 1. 清理尾部空行
trim_trailing_blank_rows(ws)

# 2. 定位最后有数据的行号（以 ID 列为基准）
last_row = find_last_data_row(ws, header2col["ID"])

# 3. 追加时，使用 last_row + 1 作为起始行
new_start = last_row + 1

# 下面在“追加新行”部分，不再用 ws.append()，而是：
#   for idx, hdr in enumerate(headers, start=1):
#       ws.cell(row=new_start, column=idx).value = row_vals[idx-1]
#   然后根据需要给新行单元格染色、添加超链接。
#
# 同理，对于多行追加，每追加一行就更新 new_start += 1。

# 最后保存
wb.save(updated_file)
