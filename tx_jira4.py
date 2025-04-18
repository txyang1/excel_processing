import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# -------------------------------
# 辅助函数：清理尾部空行 & 定位最后有数据的行
# -------------------------------
def is_row_blank(ws, row):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True

def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break

def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            return r
    return 1

# -------------------------------
# 1. 路径及参数
# -------------------------------
original_file = r"data\Ticket summary.xlsx"
new_file      = r"data\EC-EF-2 tickets (CodeCraft Jira) 2025-04-10T05_00_40+0200.csv"
updated_file  = "tx_jira_excel2.0.xlsx"
target_sheet  = "Octane and jira"

# 映射关系
mapping = {
    "Issue key":           "ID",
    "Created":             "Creation time",
    "Summary":             "Name",
    "Status":              "Phase",
    "Reporter":            "Defect finder",
    "Assignee":            "Owner",
    "Affects Version/s":   "Target I-Step:",
}

# Fund→Function 映射
fund_function_mapping = {
    "adapt speed to route geometry [01.02.02.15.02.14]": "ASRG",
    "change lane [01.02.02.15.02.07]":                  "CL",
    "allow hands-off driving 130 [01.02.02.15.02.01]":  "HOO130",
    # …其余映射…
}

# Owner→Root cause 映射（若需要）
owner_root_cause_mapping = {
    "Niklas Haeuser": "Condition evaluate",
    # …其余映射…
}

# 高亮样式
green_fill  = PatternFill("solid", fgColor="00FF00")
yellow_fill = PatternFill("solid", fgColor="FFFF00")

# -------------------------------
# 2. 读取新表 CSV
# -------------------------------
df_new = pd.read_csv(new_file)

# -------------------------------
# 3. 打开原表并清理空行
# -------------------------------
wb = load_workbook(original_file)
ws = wb[target_sheet]
trim_trailing_blank_rows(ws)

# 构建表头→列号映射
header2col = {ws.cell(1, c).value: c for c in range(1, ws.max_column+1)}
headers    = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]

# 提取原表 ID→行号
id_col  = header2col["ID"]
id2row  = {ws.cell(r, id_col).value: r
            for r in range(2, ws.max_row+1)
            if ws.cell(r, id_col).value}

# -------------------------------
# 4. 定位追加起始行
# -------------------------------
last_row = find_last_data_row(ws, id_col)

# -------------------------------
# 5. 遍历新表，更新或追加
# -------------------------------
for _, new_row in df_new.iterrows():
    new_id = new_row.get("Issue key", "")
    if pd.isna(new_id) or new_id == "":
        continue

    if new_id in id2row:
        # 更新已有行
        r = id2row[new_id]
        # 转换 Created 字段格式
        created_val = new_row.get("Created", "")
        if pd.notna(created_val) and created_val != "":
            try:
                dt = pd.to_datetime(created_val, format="%Y-%m-%dT%H:%M:%S%z", utc=True)
                created_str = dt.strftime("%-m/%-d/%Y %I:%M:%S %p")
            except:
                created_str = created_val
            c = header2col.get("Creation time")
            if c:
                cell = ws.cell(r, c)
                if cell.value != created_str:
                    cell.value = created_str
                    cell.fill  = green_fill
        # 其它映射字段
        for nk, ok in mapping.items():
            if nk == "Created":
                continue
            new_val = new_row.get(nk, "")
            if pd.isna(new_val) or new_val == "":
                continue
            # 特殊：Affects Version/s 前缀替换
            if nk == "Affects Version/s":
                s = str(new_val)
                if s.startswith("G070") or s.startswith("U006"):
                    new_val = "NA05" + s[4:]
            c = header2col.get(ok)
            if c:
                cell = ws.cell(r, c)
                if cell.value != new_val:
                    cell.value = new_val
                    cell.fill  = green_fill
        # Function 列
        if "Found in function" in header2col and "Function" in header2col:
            fund_val = ws.cell(r, header2col["Found in function"]).value or ""
            for k, v in fund_function_mapping.items():
                if k in fund_val:
                    c = header2col["Function"]
                    cell = ws.cell(r, c)
                    if cell.value != v:
                        cell.value = v
                        cell.fill  = green_fill
                    break
        # Root cause 列（如需）
        if "Owner" in header2col and "Root cause" in header2col:
            ow = ws.cell(r, header2col["Owner"]).value or ""
            for k, v in owner_root_cause_mapping.items():
                if k in ow:
                    c = header2col["Root cause"]
                    cell = ws.cell(r, c)
                    if cell.value != v:
                        cell.value = v
                        cell.fill  = green_fill
                    break

    else:
        # 追加新行
        last_row += 1
        for idx, hdr in enumerate(headers, start=1):
            val = ""
            for nk, ok in mapping.items():
                if ok == hdr:
                    tmp = new_row.get(nk, "")
                    if pd.notna(tmp) and tmp != "":
                        # Created 字段格式
                        if nk == "Created":
                            try:
                                dt = pd.to_datetime(tmp, format="%Y-%m-%dT%H:%M:%S%z", utc=True)
                                tmp = dt.strftime("%-m/%-d/%Y %I:%M:%S %p")
                            except:
                                pass
                        # 前缀替换
                        if nk == "Affects Version/s":
                            s = str(tmp)
                            if s.startswith("G070") or s.startswith("U006"):
                                tmp = "NA05" + s[4:]
                        val = tmp
                    break
            ws.cell(row=last_row, column=idx).value = val
            if val != "":
                ws.cell(row=last_row, column=idx).fill = yellow_fill
        # Function & Root cause （追加）
        if "Found in function" in header2col and "Function" in header2col:
            fund_val = new_row.get("Found in function", "") or ""
            for k, v in fund_function_mapping.items():
                if k in fund_val:
                    c = header2col["Function"]
                    ws.cell(last_row, c).value = v
                    ws.cell(last_row, c).fill  = yellow_fill
                    break
        if "Owner" in header2col and "Root cause" in header2col:
            ow = new_row.get("Owner", "") or ""
            for k, v in owner_root_cause_mapping.items():
                if k in ow:
                    c = header2col["Root cause"]
                    ws.cell(last_row, c).value = v
                    ws.cell(last_row, c).fill  = yellow_fill
                    break

# -------------------------------
# 6. 填充公式和其它列
# -------------------------------
max_row = ws.max_row

# Days 列
days_idx     = header2col.get("Days")
creation_idx = header2col.get("Creation time")
if days_idx and creation_idx:
    creation_col = get_column_letter(creation_idx)
    for r in range(2, max_row+1):
        ws.cell(r, days_idx).value = f'=DATEDIF(${creation_col}{r},TODAY(),"D")'

# Octane or Jira 列
oir_idx = header2col.get("Octane or Jira")
if oir_idx:
    fn = os.path.basename(new_file)
    val = "Jira"  # CSV 来自 Jira
    for r in range(2, max_row+1):
        ws.cell(r, oir_idx).value = val

# Open > 20 days 列
open20_idx = header2col.get("Open >20 days") or header2col.get("Open > 20 days")
if open20_idx and days_idx:
    dl = get_column_letter(days_idx)
    for r in range(2, max_row+1):
        ws.cell(r, open20_idx).value = f'=IF({dl}{r}>20,1,0)'

# No TIS 列
nt_idx = header2col.get("No TIS")
pl_idx = header2col.get("Planned closing version")
ti_idx = header2col.get("Target I-Step:")
if nt_idx and pl_idx and ti_idx:
    pl_l = get_column_letter(pl_idx)
    ti_l = get_column_letter(ti_idx)
    for r in range(2, max_row+1):
        ws.cell(r, nt_idx).value = f'=IF(OR({pl_l}{r}<>"",{ti_l}{r}<>""),0,1)'

# -------------------------------
# 7. 保存
# -------------------------------
wb.save(updated_file)
print("更新完成！空行已清理，前缀替换、差异染绿、新增染黄、公式已填充，结果保存在", updated_file)
