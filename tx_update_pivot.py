import os
import re
import time
import json
import pandas as pd
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from watchdog.observers.polling import PollingObserver as Observer
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# === 加载配置 ===
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
with open(os.path.join(BASE_DIR, "unified_config_auto3.json"), 'r', encoding='utf-8') as f:
    cfg = json.load(f)

folders    = cfg['folders']
ORIG_DIR   = os.path.join(BASE_DIR, folders['orig_dir'])
JIRA_DIR   = os.path.join(BASE_DIR, folders['jira_dir'])
OCTANE_DIR = os.path.join(BASE_DIR, folders['octane_dir'])

sheet            = cfg['sheet']['target_sheet']
sources          = cfg['sources']
settings         = cfg.get('settings', {})
clear_old        = settings.get('clear_old_highlight','N').upper() == 'Y'
fund_patterns    = cfg.get('fund_function_patterns', {})
owner_patterns   = cfg.get('owner_root_cause_patterns', {})

def is_row_blank(ws, row):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True

def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break

def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            return r
    return 1

def update_excel(new_fp):
    # 判断来源
    folder = os.path.basename(os.path.dirname(new_fp))
    if folder == os.path.basename(JIRA_DIR):
        source_key = "Jira"
    elif folder == os.path.basename(OCTANE_DIR):
        source_key = "Octane"
    else:
        print(f" → 未识别来源 ({folder})，跳过。")
        return

    # 原表路径
    orig_fp = os.path.join(ORIG_DIR, cfg['paths']['original_file'])
    # 读取新数据
    src_cfg   = sources[source_key]
    read_meth = src_cfg['read_method']
    date_col  = src_cfg.get('date_col')
    mapping   = src_cfg['mapping']
    df_new    = pd.read_excel(new_fp) if read_meth=="excel" else pd.read_csv(new_fp)

    # 打开原表
    wb = load_workbook(orig_fp)
    ws = wb[sheet]

    # 清除旧高亮
    if clear_old:
        green = {"8ED973","008ED973","FF8ED973"}
        blue  = {'FFADD8E6','00ADD8E6','ADD8E6'}
        gray  = {'FFC0C0C0','00C0C0C0','C0C0C0'}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                pf = cell.fill
                if getattr(pf,'patternType',None)=="solid":
                    argb = pf.fgColor.rgb or pf.fgColor.value
                    if argb in green|blue|gray:
                        cell.fill = PatternFill()

    # 删除尾部空行并重置行高
    trim_trailing_blank_rows(ws)
    for r in range(1, ws.max_row+1):
        ws.row_dimensions[r].height = None

    # 构建列索引
    header2col = { ws.cell(1,c).value: c for c in range(1, ws.max_column+1) }
    id_col     = header2col['ID']
    headers    = list(header2col.keys())

    # 已有 ID 对应行号
    id2row = {
        ws.cell(r, id_col).value: r
        for r in range(2, ws.max_row+1)
        if ws.cell(r, id_col).value
    }

    # 提取新表中的超链接（若有）
    id2url_new = {}
    if read_meth=="excel":
        tmpwb = load_workbook(new_fp)
        nws   = tmpwb.active
        cid   = next((c for c in range(1, nws.max_column+1)
                      if nws.cell(1,c).value=="ID"), None)
        if cid:
            for r in range(2, nws.max_row+1):
                cell = nws.cell(r, cid)
                if cell.hyperlink:
                    id2url_new[cell.value] = cell.hyperlink.target

    # 找到现有数据区最后一行
    original_last = find_last_data_row(ws, id_col)

    # 准备填充样式
    update_fill = PatternFill("solid", fgColor="ADD8E6")
    new_fill    = PatternFill("solid", fgColor="8ED973")
    gray_fill   = PatternFill("solid", fgColor="C0C0C0")

    id_key = next(k for k,v in mapping.items() if v=="ID")

    # 遍历新记录
    for _, new_row in df_new.iterrows():
        new_id = new_row.get(id_key)
        if pd.isna(new_id) or not new_id:
            continue

        if new_id in id2row:
            # 已有行：更新字段（略，保持原逻辑）
            r = id2row[new_id]
            phase = ws.cell(r, header2col["Phase"]).value or ""
            if any(k in phase for k in ("Concluded","Closed","Resolved")):
                continue

            # 更新日期
            if date_col:
                raw = new_row.get(date_col, "")
                if pd.notna(raw) and raw!="":
                    try:
                        dt = pd.to_datetime(raw).to_pydatetime()
                        c  = header2col[mapping[date_col]]
                        cell = ws.cell(r, c)
                        if cell.value != dt:
                            cell.value         = dt
                            cell.number_format = "m/d/yyyy h:mm:ss AM/PM"
                            cell.fill          = update_fill
                    except:
                        pass
            # 更新其他字段
            for nk, ok in mapping.items():
                if nk==date_col or ok=="Function":
                    continue
                val = new_row.get(nk, "")
                if pd.isna(val) or val=="":
                    continue
                c = header2col.get(ok)
                if c:
                    cell = ws.cell(r, c)
                    if cell.value != val:
                        cell.value = val
                        cell.fill  = update_fill

            # 更新 Root cause
            if "Owner" in header2col and "Root cause" in header2col:
                raw_ow = ws.cell(r, header2col["Owner"]).value or ""
                ow = str(raw_ow).lower()
                for cause, names in owner_patterns.items():
                    if any(n.lower() in ow for n in names):
                        cell = ws.cell(r, header2col["Root cause"])
                        if cell.value != cause:
                            cell.value = cause
                            cell.fill  = update_fill
                        break

        else:
            # 新增行：插入到原有数据区末尾前
            insert_at = original_last + 1
            ws.insert_rows(insert_at)

            # 填值并高亮
            for idx, hdr in enumerate(headers, start=1):
                val = ""
                for nk, ok in mapping.items():
                    if ok == hdr:
                        tmp = new_row.get(nk, "")
                        if pd.notna(tmp) and tmp!="":
                            if hdr == mapping.get(date_col):
                                try:
                                    val = pd.to_datetime(tmp).to_pydatetime()
                                except:
                                    val = tmp
                            else:
                                val = tmp
                        break
                cell = ws.cell(insert_at, idx)
                cell.value = val
                if val != "":
                    cell.fill = new_fill
                    if hdr == mapping.get(date_col):
                        cell.number_format = "m/d/yyyy h:mm:ss AM/PM"

            # 超链接
            url = id2url_new.get(new_id)
            if url:
                cell = ws.cell(insert_at, id_col)
                cell.hyperlink = url
                cell.style     = "Hyperlink"

            # Function via patterns
            if "Found in function" in header2col and "Function" in header2col:
                raw_fv = ws.cell(insert_at, header2col["Found in function"]).value or ""
                fv = str(raw_fv).lower()
                for func_name, kws in fund_patterns.items():
                    if any(kw.lower() in fv for kw in kws):
                        cell = ws.cell(insert_at, header2col["Function"])
                        cell.value = func_name
                        cell.fill  = new_fill
                        break

            # Root cause via patterns
            if "Owner" in header2col and "Root cause" in header2col:
                raw_ow = ws.cell(insert_at, header2col["Owner"]).value or ""
                ow = str(raw_ow).lower()
                for cause, names in owner_patterns.items():
                    if any(n.lower() in ow for n in names):
                        cell = ws.cell(insert_at, header2col["Root cause"])
                        cell.value = cause
                        cell.fill  = new_fill
                        break

            # 更新插入位置
            original_last += 1

    # 填充公式、其它列（同之前逻辑）
    max_row = ws.max_row
    # Days 列
    d_idx  = header2col.get("Days")
    ct_idx = header2col.get("Creation time")
    if d_idx and ct_idx:
        colL = get_column_letter(ct_idx)
        for r in range(2, max_row+1):
            ws.cell(r, d_idx).value = f'=DATEDIF(${colL}{r},TODAY(),"D")'
    # Octane or Jira
    oij_idx = header2col.get("Octane or Jira")
    if oij_idx:
        for r in range(find_last_data_row(ws, id_col)+1, max_row+1):
            ws.cell(r, oij_idx).value = source_key
    # Open >20 days
    o20 = header2col.get("Open >20 days") or header2col.get("Open > 20 days")
    if o20 and d_idx:
        dl = get_column_letter(d_idx)
        for r in range(2, max_row+1):
            ws.cell(r, o20).value = f'=IF({dl}{r}>20,1,0)'
    # No TIS
    nt = header2col.get("No TIS")
    pl = header2col.get("Planned closing version")
    ti = header2col.get("Target I-Step:")
    if nt and pl and ti:
        pL = get_column_letter(pl)
        tL = get_column_letter(ti)
        for r in range(2, max_row+1):
            ws.cell(r, nt).value = f'=IF(OR({pL}{r}<>"",{tL}{r}<>""),0,1)'
    # Top issue
    tags_idx = header2col.get("Tags")
    top_idx  = header2col.get("Top issue Candidiate")
    if tags_idx and top_idx:
        for r in range(2, max_row+1):
            tags_val = ws.cell(r, tags_idx).value or ""
            top_cell = ws.cell(r, top_idx)
            top_val  = top_cell.value or ""
            if "IPN_CN_TopIssue" in tags_val:
                if top_val == "":
                    top_cell.value = "Yes"
                    top_cell.fill  = new_fill
            else:
                if top_val == "Yes":
                    top_cell.fill = gray_fill
    # Rejected ticket
    br_idx  = header2col.get("Blocking reason")
    ph_idx  = header2col.get("Phase")
    rej_idx = header2col.get("Rejected ticket")
    if br_idx and ph_idx and rej_idx:
        for r in range(2, ws.max_row+1):
            br_val = ws.cell(r, br_idx).value
            ph_val = str(ws.cell(r, ph_idx).value or "")
            ws.cell(r, rej_idx).value = 1 if (br_val not in (None,"") and "New" in ph_val) else 0

    # 保存原表
    wb.save(orig_fp)
    print(f"[{datetime.now().strftime('%Y%m%d_%H%M%S')}] 更新完成并保存到原表 {orig_fp}")

class FolderHandler(FileSystemEventHandler):
    def __init__(self, folders, debounce_seconds=5):
        self.folders   = folders
        self._last_run = {}
        self._debounce = debounce_seconds

    def on_created(self, event):
        if event.is_directory:
            return
        self._maybe_update(event.src_path)

    def _maybe_update(self, path):
        if os.path.isdir(path):
            return
        dir_name = os.path.basename(os.path.dirname(path))
        if dir_name not in (folders['jira_dir'], folders['octane_dir']):
            return
        now  = time.time()
        last = self._last_run.get(path, 0)
        if now - last < self._debounce:
            print(f"⚠️ 去抖：忽略短时间内重复触发 {path}")
            return
        self._last_run[path] = now
        update_excel(path)

def main():
    for d in (ORIG_DIR, JIRA_DIR, OCTANE_DIR):
        os.makedirs(d, exist_ok=True)
    observer = Observer()
    handler  = FolderHandler(folders, debounce_seconds=5)
    observer.schedule(handler, path=JIRA_DIR,   recursive=False)
    observer.schedule(handler, path=OCTANE_DIR, recursive=False)
    observer.start()
    print(f"监控：{JIRA_DIR}, {OCTANE_DIR}")
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

if __name__ == "__main__":
    main()
