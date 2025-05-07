import os
import re
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# === 辅助函数 ===
def is_row_blank(ws, row):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True

def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break

def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            return r
    return 1

# === 0. 读取配置 ===
with open("unified_config2.7.1.json", "r", encoding="utf-8") as f:
    cfg = json.load(f)

paths      = cfg["paths"]
orig_fp    = paths["original_file"]
new_fp     = paths["new_file"]
upd_fp     = paths["updated_file"]

sheet      = cfg["sheet"]["target_sheet"]
sources    = cfg["sources"]
settings   = cfg.get("settings", {})
clear_flag = settings.get("clear_old_highlight", "N").upper()

# 模式映射
fund_patterns  = cfg.get("fund_function_patterns", {})
owner_patterns = cfg.get("owner_root_cause_patterns", {})

# === 1. 判断来源并加载映射 ===
base       = os.path.basename(new_fp)
source_key = next((k for k,v in sources.items() if v["pattern"] in base), None)
if not source_key:
    raise RuntimeError("无法识别 new_file 来源 (Octane 或 Jira)")

src_cfg   = sources[source_key]
read_meth = src_cfg["read_method"]
date_col  = src_cfg.get("date_col")
mapping   = src_cfg["mapping"]
#添加判定，如果是Octane先清除之前的颜色
clear_flag = "Y" if src_cfg["pattern"] == "Octane" else "N"
print(f"判断是否清洗更新颜色：",clear_flag)
# === 2. 读取新表 ===
df_new = pd.read_excel(new_fp) if read_meth == "excel" else pd.read_csv(new_fp)

# === 3. 打开原表，清除旧高亮 & 删除尾部空行 ===
wb = load_workbook(orig_fp)
ws = wb[sheet]

if clear_flag == "Y":
    green_codes = {"8ED973","008ED973","FF8ED973"}
    blue_codes   = {"FFADD8E6","00ADD8E6","ADD8E6"}
    gray_codes   = {"FFC0C0C0","00C0C0C0","C0C0C0"}
    for row in ws.iter_rows(min_row=2, min_col=1,
                            max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            pf = cell.fill
            if getattr(pf, 'patternType', None) == "solid":
                argb = pf.fgColor.rgb or pf.fgColor.value
                if argb in green_codes or argb in blue_codes or argb in gray_codes:
                    cell.fill = PatternFill()

trim_trailing_blank_rows(ws)

#重置所有行高
for r in range(1,ws.max_row +1):
    ws.row_dimensions[r].height = None

# 构建表头映射
header2col = { ws.cell(1,c).value: c for c in range(1, ws.max_column + 1) }
headers    = list(header2col.keys())
id_col     = header2col["ID"]

# 原表已有 ID→行
id2row = {
    ws.cell(r,id_col).value: r
    for r in range(2, ws.max_row + 1)
    if ws.cell(r,id_col).value
}

# 对 Excel 来源，提取新表 ID→超链接
id2url_new = {}
if read_meth == "excel":
    nwb    = load_workbook(new_fp)
    nws    = nwb.active
    col_id = next((c for c in range(1, nws.max_column+1)
                   if nws.cell(1,c).value=="ID"), None)
    if col_id:
        for r in range(2, nws.max_row+1):
            cell = nws.cell(r, col_id)
            if cell.hyperlink:
                id2url_new[cell.value] = cell.hyperlink.target

# 定位追加起始行
last_row      = find_last_data_row(ws, id_col)
original_last = last_row

# 高亮样式
update_fill = PatternFill("solid", fgColor="ADD8E6")   # 淡蓝
new_fill    = PatternFill("solid", fgColor="8ED973")   # 淡绿色
gray_fill   = PatternFill("solid", fgColor="C0C0C0")   # 灰色

# 新表中 ID 列对应的键
id_key = next(k for k,v in mapping.items() if v=="ID")

# === 4. 遍历 df_new，更新 or 追加 ===
for _, new_row in df_new.iterrows():
    new_id = new_row.get(id_key, "")
    if pd.isna(new_id) or not new_id:
        continue

    if new_id in id2row:
        # —— 已有行，更新除 Function 之外字段 并染淡蓝 —— 
        r = id2row[new_id]

        # 跳过已结束行
        phase_val = ws.cell(r, header2col["Phase"]).value or ""
        if any(p in phase_val for p in ("Concluded","Closed","Resolved")):
            continue

        # 更新日期 (Jira)
        if date_col:
            raw = new_row.get(date_col, "")
            if pd.notna(raw) and raw != "":
                try:
                    ts = pd.to_datetime(raw)
                    dt = ts.to_pydatetime()
                    c  = header2col[mapping[date_col]]
                    cell = ws.cell(r,c)
                    if cell.value != dt:
                        cell.value         = dt
                        cell.number_format = "m/d/yyyy h:mm:ss AM/PM"
                        cell.fill          = update_fill
                except:
                    pass

        # 更新其他字段 (跳过 Function)
        for nk, ok in mapping.items():
            if nk==date_col or ok=="Function":
                continue
            val = new_row.get(nk, "")
            if pd.isna(val) or val=="":
                continue


            #if ok=="Involved I-Step" and (str(val).startswith("G070") or str(val).startswith("U006")):
                #val = "NA05" + str(val)[4:]
            if ok == "Involved I-Step":
                s = str(val)
                # 1) 如果以 G070 或 U006 开头，沿用原逻辑
                if s.startswith("G070") or s.startswith("U006"):
                    val = "NA05" + s[4:]
                else:
                    # 2) 否则尝试提取全角或半角括号内的编号，如 “（25-07-452 ATS+3...）”
                    m = re.search(r'[（(]([\d-]+)', s)
                    if m:
                        code = m.group(1)               # e.g. "25-07-452"
                        val = f"NA05-{code}"           # 结果 "NA05-25-07-452"



            c = header2col.get(ok)
            if c:
                cell = ws.cell(r,c)
                if cell.value != val:
                    cell.value = val
                    cell.fill  = update_fill

        '''# 更新 Function via patterns
        if "Found in function" in header2col and "Function" in header2col:
            raw_fv = ws.cell(r, header2col["Found in function"]).value or ""
            fv = str(raw_fv).lower()
            for func_name, kws in fund_patterns.items():
                if any(kw.lower() in fv for kw in kws):
                    cell = ws.cell(r, header2col["Function"])
                    if cell.value != func_name:
                        cell.value = func_name
                        cell.fill  = update_fill
                    break'''

        # 更新 Root cause via patterns
        if "Owner" in header2col and "Root cause" in header2col:
            raw_ow = ws.cell(r, header2col["Owner"]).value or ""
            ow = str(raw_ow).lower()
            for cause, names in owner_patterns.items():
                if any(n.lower() in ow for n in names):
                    cell = ws.cell(r, header2col["Root cause"])
                    if cell.value != cause:
                        cell.value = cause
                        cell.fill  = update_fill
                    break

    else:
        # —— 新增行，填值并染粉色 —— 
        last_row += 1
        for idx, hdr in enumerate(headers, start=1):
            val = ""
            for nk, ok in mapping.items():
                if ok == hdr:
                    tmp = new_row.get(nk, "")
                    if pd.notna(tmp) and tmp != "":
                        if hdr == mapping.get(date_col):
                            try:
                                ts = pd.to_datetime(tmp)
                                val = ts.to_pydatetime()
                            except:
                                val = tmp
                        else:
                            val = tmp

                        #if ok=="Involved I-Step:" and (str(val).startswith("G070") or str(val).startswith("U006")):
                            #val = "NA05" + str(val)[4:]
                        if ok == "Involved I-Step":
                            s = str(val)
                            # 1) 如果以 G070 或 U006 开头，沿用原逻辑
                            if s.startswith("G070") or s.startswith("U006"):
                                val = "NA05" + s[4:]
                            else:
                                # 2) 否则尝试提取全角或半角括号内的编号，如 “（25-07-452 ATS+3...）”
                                m = re.search(r'[（(]([\d-]+)', s)
                                if m:
                                    code = m.group(1)               # e.g. "25-07-452"
                                    val = f"NA05-{code}"           # 结果 "NA05-25-07-452"

                    break
            cell = ws.cell(last_row, idx)
            cell.value = val
            if val != "":
                cell.fill = new_fill
                if hdr == mapping.get(date_col):
                    cell.number_format = "m/d/yyyy h:mm:ss AM/PM"

        # 新增 ID 超链接
        url = id2url_new.get(new_id)
        if url:
            cell = ws.cell(last_row, id_col)
            cell.hyperlink = url
            cell.style     = "Hyperlink"

        # 新增 Function via patterns
        if "Found in function" in header2col and "Function" in header2col:
            raw_fv = ws.cell(last_row, header2col["Found in function"]).value or ""
            fv = str(raw_fv).lower()
            for func_name, kws in fund_patterns.items():
                if any(kw.lower() in fv for kw in kws):
                    cell = ws.cell(last_row, header2col["Function"])
                    cell.value = func_name
                    cell.fill  = new_fill
                    break

        # 新增 Root cause via patterns
        if "Owner" in header2col and "Root cause" in header2col:
            raw_ow = ws.cell(last_row, header2col["Owner"]).value or ""
            ow = str(raw_ow).lower()
            for cause, names in owner_patterns.items():
                if any(n.lower() in ow for n in names):
                    cell = ws.cell(last_row, header2col["Root cause"])
                    cell.value = cause
                    cell.fill  = new_fill
                    break

       
# === 5. 填充公式 & 标记 Octane/Jira & 其它列 ===
max_row = ws.max_row

# 5.1 Days 列公式
d_idx  = header2col.get("Days")
ct_idx = header2col.get("Creation time")
if d_idx and ct_idx:
    colL = get_column_letter(ct_idx)
    for r in range(2, max_row + 1):
        ws.cell(r, d_idx).value = f'=DATEDIF(${colL}{r},TODAY(),"D")'

# 5.2 Octane or Jira 列（仅新增）
oij_idx = header2col.get("Octane or Jira")
if oij_idx:
    for r in range(original_last + 1, max_row + 1):
        ws.cell(r, oij_idx).value = source_key

# 5.3 Open > 20 days 列
o20 = header2col.get("Open >20 days") or header2col.get("Open > 20 days")
if o20 and d_idx:
    dl = get_column_letter(d_idx)
    for r in range(2, max_row + 1):
        ws.cell(r, o20).value = f'=IF({dl}{r}>20,1,0)'

# 5.4 No TIS 列
nt = header2col.get("No TIS")
pl = header2col.get("Planned closing version")
ti = header2col.get("Target I-Step:")
if nt and pl and ti:
    pL = get_column_letter(pl)
    tL = get_column_letter(ti)
    for r in range(2, max_row + 1):
        ws.cell(r, nt).value = f'=IF(OR({pL}{r}<>"",{tL}{r}<>""),0,1)'

# 5.5 Top issue 列
tags_idx = header2col.get("Tags")
top_idx =  header2col.get("Top issue Candidiate")
gray_fill = PatternFill("solid", fgColor="C0C0C0")

if tags_idx and top_idx:
    for r in range(2,max_row+1):
        tags_val = ws.cell(r, tags_idx).value or ""
        top_cell = ws.cell(r, top_idx)
        top_val = top_cell.value or ""
        if "IPN_CN_TopIssue" in tags_val:
            if top_val == "":
                top_cell.value = "Yes"
                top_cell.fill = new_fill
        else:
            if top_val == "Yes":
                top_cell.fill = gray_fill

# Rejected ticket
br_idx  = header2col.get("Blocking reason")
ph_idx  = header2col.get("Phase")
rej_idx = header2col.get("Rejected ticket")

if br_idx and ph_idx and rej_idx:
    for r in range(2, ws.max_row + 1):
        br_val = ws.cell(r, br_idx).value
        ph_val = str(ws.cell(r, ph_idx).value or "")
        ws.cell(r, rej_idx).value = 1 if (br_val not in (None, "") and "New" in ph_val) else 0


# === 6. 保存 ===
wb.save(upd_fp)
print(f"更新完成 (来源={source_key})，新增行绿色，高亮；更新行淡蓝，高亮。结果已保存至 '{upd_fp}'.")
