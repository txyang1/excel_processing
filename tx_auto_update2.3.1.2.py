import os
import re
import time
import json
import pandas as pd
import schedule
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from watchdog.observers.polling import PollingObserver as Observer
from datetime import datetime
import pythoncom
from win32com.client import gencache, constants 

# === 加载配置 ===
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
with open(os.path.join(BASE_DIR, "unified_config_auto2.json"), 'r', encoding='utf-8') as f:
    cfg = json.load(f)

folders    = cfg['folders']
ORIG_DIR   = os.path.join(BASE_DIR, folders['orig_dir'])
JIRA_DIR   = os.path.join(BASE_DIR, folders['jira_dir'])
OCTANE_DIR = os.path.join(BASE_DIR, folders['octane_dir'])

paths      = cfg['paths']
sheet      = cfg['sheet']['target_sheet']
sources    = cfg['sources']
settings   = cfg.get('settings', {})
clear_old  = settings.get('clear_old_highlight','N').upper() == 'Y'
fund_patterns  = cfg.get('fund_function_patterns', {})
owner_patterns = cfg.get('owner_root_cause_patterns', {})

# Excel VBA 常量（数字形式）
XL_UP       = -4162   # xlUp
XL_TOLEFT   = -4159   # xlToLeft
XL_DATABASE = 1       # xlDatabase

def is_row_blank(ws, row):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True

def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break

def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            return r
    return 1

def _refresh_pivots_in_workbook(xlsx_path, data_sheet_name):
    """
    用 COM 打开 xlsx_path，基于 data_sheet_name 的 ID 列范围
    更新并刷新整本工作簿中所有 PivotTable。
    """
    pythoncom.CoInitialize()
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))

        try:
            ds = wb.Worksheets(data_sheet_name)
        except Exception:
            print(f"❌ COM: 找不到数据表 “{data_sheet_name}”，跳过 Pivot 刷新")
            wb.Close(False)
            excel.Quit()
            return

        # 1) 找到“ID”列在表头的列号
        header_row = 1
        last_header_col = ds.Cells(header_row, ds.Columns.Count).End(XL_TOLEFT).Column
        id_col_idx = None
        for c in range(1, last_header_col + 1):
            if str(ds.Cells(header_row, c).Value).strip() == "ID":
                id_col_idx = c
                break
        if not id_col_idx:
            print("❌ COM: 未找到 ‘ID’ 列，无法定位数据区域，跳过 Pivot 刷新")
            wb.Close(False)
            excel.Quit()
            return

        # 2) 用 ID 列 End(xlUp) 定位最后一行
        last_row = ds.Cells(ds.Rows.Count, id_col_idx).End(XL_UP).Row
        if last_row <= header_row:
            print("⚠️ COM: ‘ID’ 列无数据，跳过 Pivot 刷新")
            wb.Close(False)
            excel.Quit()
            return

        # 3) 用表头最后一列作为列末
        last_col = last_header_col

        # 4) 构造 SourceData 引用
        top_left     = ds.Cells(header_row, 1).Address    # "$A$1"
        bottom_right = ds.Cells(last_row, last_col).Address
        source_ref   = f"'{data_sheet_name}'!{top_left}:{bottom_right}"
        print(f"  COM: 设置 Pivot 数据源为 {source_ref}")

        # 5) 遍历所有工作表的 PivotTable
        for ws in wb.Worksheets:
            try:
                pts = ws.PivotTables()  # 必须调用
                cnt = pts.Count
            except Exception:
                continue
            if cnt == 0:
                continue
            print(f"  ▶ COM: 工作表 [{ws.Name}] 有 {cnt} 个 PivotTable，开始更新…")
            for i in range(1, cnt+1):
                pt = pts.Item(i)
                try:
                    cache = wb.PivotCaches().Create(
                        SourceType=XL_DATABASE,
                        SourceData=source_ref
                    )
                    pt.ChangePivotCache(cache)
                    pt.RefreshTable()
                    print(f"    ✔️ 已刷新 PivotTable [{pt.Name}]")
                except Exception as e:
                    print(f"    ❌ 刷新 [{pt.Name}] 失败：{e}")

        wb.Save()
        wb.Close(False)
        excel.Quit()
        print("✅ COM: 所有 PivotTable 已更新并刷新\n")
    finally:
        pythoncom.CoUninitialize()

    
    

def update_excel(new_fp):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    print(f"[{ts}] 开始更新: {new_fp}")

    folder = os.path.basename(os.path.dirname(new_fp))
    if folder == os.path.basename(JIRA_DIR):
        source_key  = "Jira"
        jira_flag   = 'Y'
        octane_flag = 'N'
    elif folder == os.path.basename(OCTANE_DIR):
        source_key  = "Octane"
        jira_flag   = 'N'
        octane_flag = 'Y'
    else:
        print(f" → 未识别来源 ({folder})，跳过。")
        return

    # 原表与新版输出路径
    orig_fp  = os.path.join(ORIG_DIR, paths['original_file'])
    template = paths['updated_file']
    filename = template.format(time=ts, jira=jira_flag, octane=octane_flag)
    upd_fp   = os.path.join(os.path.dirname(orig_fp), filename)#

    src_cfg   = sources[source_key]
    read_meth = src_cfg['read_method']
    date_col  = src_cfg.get('date_col')
    mapping   = src_cfg['mapping']

    # 读取 df_new
    df_new = pd.read_excel(new_fp) if read_meth=="excel" else pd.read_csv(new_fp)

    #添加判定，如果是Octane先清除之前的颜色
    clear_old = "Y" if src_cfg["pattern"] == "Octane" else "N"
    print(f"判断是否清洗就更新颜色：",clear_old)

    # 打开原表
    wb = load_workbook(orig_fp)
    ws = wb[sheet]

    # 清除旧高亮
    if clear_old:
        green = {"8ED973","008ED973","FF8ED973"}
        blue  = {'FFADD8E6','00ADD8E6','ADD8E6'}
        gray  = {'FFC0C0C0','00C0C0C0','C0C0C0'}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                pf = cell.fill
                if getattr(pf,'patternType',None)=="solid":
                    argb = pf.fgColor.rgb or pf.fgColor.value
                    if argb in green|blue|gray:
                        cell.fill = PatternFill()

    trim_trailing_blank_rows(ws)
    #重置所有行高
    for r in range(1,ws.max_row +1):
        ws.row_dimensions[r].height = None

    header2col = { ws.cell(1,c).value: c for c in range(1, ws.max_column+1) }
    id_col     = header2col['ID']
    headers    = list(header2col.keys())

    # 已有ID→行
    id2row = {
        ws.cell(r,id_col).value: r
        for r in range(2, ws.max_row+1)
        if ws.cell(r,id_col).value
    }

    # 如果是 Excel，提取超链接
    id2url_new = {}
    if read_meth=="excel":
        tmpwb = load_workbook(new_fp)
        nws   = tmpwb.active
        cid   = next((c for c in range(1,nws.max_column+1)
                      if nws.cell(1,c).value=="ID"), None)
        if cid:
            for r in range(2, nws.max_row+1):
                cell = nws.cell(r,cid)
                if cell.hyperlink:
                    id2url_new[cell.value] = cell.hyperlink.target

    last_row      = find_last_data_row(ws, id_col)
    original_last = last_row

    update_fill = PatternFill("solid", fgColor="ADD8E6")
    new_fill    = PatternFill("solid", fgColor="8ED973")
    gray_fill   = PatternFill("solid", fgColor="C0C0C0")

    id_key = next(k for k,v in mapping.items() if v=="ID")

    # 4. 更新 or 追加
    for _, new_row in df_new.iterrows():
        new_id = new_row.get(id_key)
        if pd.isna(new_id) or not new_id:
            continue

        if new_id in id2row:
            # 已有行------------------------
            r = id2row[new_id]
            phase_val = ws.cell(r,header2col["Phase"]).value or ""
            if any(k in phase_val for k in ("Concluded","Closed","Resolved")):
                continue

            # 更新日期 (Jira)
            if date_col:
                raw = new_row.get(date_col, "")
                if pd.notna(raw) and raw != "":
                    try:
                        ts = pd.to_datetime(raw)
                        dt = ts.to_pydatetime()
                        c  = header2col[mapping[date_col]]
                        cell = ws.cell(r,c)
                        if cell.value != dt:
                            cell.value         = dt
                            cell.number_format = "m/d/yyyy h:mm:ss AM/PM"
                            cell.fill          = update_fill
                    except:
                        pass

            # 更新其他字段 (跳过 Function)
            for nk, ok in mapping.items():
                if nk==date_col or ok=="Function":
                    continue
                val = new_row.get(nk, "")
                if pd.isna(val) or val=="":
                    continue
                if ok == "Involved I-Step":
                    s = str(val)
                    # 1) 如果以 G070 或 U006 开头，沿用原逻辑
                    if s.startswith("G070") or s.startswith("U006"):
                        val = "NA05" + s[4:]
                    else:
                        # 2) 否则尝试提取全角或半角括号内的编号，如 “（25-07-452 ATS+3...）”
                        m = re.search(r'[（(]([\d-]+)', s)
                        if m:
                            code = m.group(1)               # e.g. "25-07-452"
                            val = f"NA05-{code}"           # 结果 "NA05-25-07-452"

                c = header2col.get(ok)
                if c:
                    cell = ws.cell(r,c)
                    if cell.value != val:
                        cell.value = val
                        cell.fill  = update_fill

            # 更新 Root cause via patterns
            if "Owner" in header2col and "Root cause" in header2col:
                raw_ow = ws.cell(r, header2col["Owner"]).value or ""
                ow = str(raw_ow).lower()
                for cause, names in owner_patterns.items():
                    if any(n.lower() in ow for n in names):
                        cell = ws.cell(r, header2col["Root cause"])
                        if cell.value != cause:
                            cell.value = cause
                            cell.fill  = update_fill
                        break

        else:
            # —— 新增行，填值并染粉色 —— 
            last_row += 1
            for idx, hdr in enumerate(headers, start=1):
                val = ""
                for nk, ok in mapping.items():
                    if ok == hdr:
                        tmp = new_row.get(nk, "")
                        if pd.notna(tmp) and tmp != "":
                            if hdr == mapping.get(date_col):
                                try:
                                    ts = pd.to_datetime(tmp)
                                    val = ts.to_pydatetime()
                                except:
                                    val = tmp
                            else:
                                val = tmp
                            #对于 Involved I-Step的更改统一格式成NA05开头
                            if ok == "Involved I-Step":
                                s = str(val)
                                # 1) 如果以 G070 或 U006 开头，沿用原逻辑
                                if s.startswith("G070") or s.startswith("U006"):
                                    val = "NA05" + s[4:]
                                else:
                                    # 2) 否则尝试提取全角或半角括号内的编号，如 “（25-07-452 ATS+3...）”
                                    m = re.search(r'[（(]([\d-]+)', s)
                                    if m:
                                        code = m.group(1)               # e.g. "25-07-452"
                                        val = f"NA05-{code}"           # 结果 "NA05-25-07-452"

                        break
                cell = ws.cell(last_row, idx)
                cell.value = val
                if val != "":
                    cell.fill = new_fill
                    if hdr == mapping.get(date_col):
                        cell.number_format = "m/d/yyyy h:mm:ss AM/PM"

            # 新增 ID 超链接，继承自Octane的表
            url = id2url_new.get(new_id)
            if url:
                cell = ws.cell(last_row, id_col)
                cell.hyperlink = url
                cell.style     = "Hyperlink"
           
            # 新增 Function via patterns
            if "Found in function" in header2col and "Function" in header2col:
                raw_fv = ws.cell(last_row, header2col["Found in function"]).value or ""
                fv = str(raw_fv).lower()
                for func_name, kws in fund_patterns.items():
                    if any(kw.lower() in fv for kw in kws):
                        cell = ws.cell(last_row, header2col["Function"])
                        cell.value = func_name
                        cell.fill  = new_fill
                        break

            # 新增 Root cause via patterns
            if "Owner" in header2col and "Root cause" in header2col:
                raw_ow = ws.cell(last_row, header2col["Owner"]).value or ""
                ow = str(raw_ow).lower()
                for cause, names in owner_patterns.items():
                    if any(n.lower() in ow for n in names):
                        cell = ws.cell(last_row, header2col["Root cause"])
                        cell.value = cause
                        cell.fill  = new_fill
                        break

    # === 5. 填充公式 & 标记 Octane/Jira & 其它列 ===
    max_row = ws.max_row

    # 5.1 Days 列公式
    d_idx  = header2col.get("Days")
    ct_idx = header2col.get("Creation time")
    if d_idx and ct_idx:
        colL = get_column_letter(ct_idx)
        for r in range(2, max_row + 1):
            ws.cell(r, d_idx).value = f'=DATEDIF(${colL}{r},TODAY(),"D")'

    # 5.2 Octane or Jira 列（仅新增）
    oij_idx = header2col.get("Octane or Jira")
    if oij_idx:
        for r in range(original_last + 1, max_row + 1):
            ws.cell(r, oij_idx).value = source_key

    # 5.3 Open > 20 days 列
    o20 = header2col.get("Open >20 days") or header2col.get("Open > 20 days")
    if o20 and d_idx:
        dl = get_column_letter(d_idx)
        for r in range(2, max_row + 1):
            ws.cell(r, o20).value = f'=IF({dl}{r}>20,1,0)'

    # 5.4 No TIS 列
    nt = header2col.get("No TIS")
    pl = header2col.get("Planned closing version")
    ti = header2col.get("Target I-Step:")
    if nt and pl and ti:
        pL = get_column_letter(pl)
        tL = get_column_letter(ti)
        for r in range(2, max_row + 1):
            ws.cell(r, nt).value = f'=IF(OR({pL}{r}<>"",{tL}{r}<>""),0,1)'

    # 5.5 Top issue 列
    tags_idx = header2col.get("Tags")
    top_idx =  header2col.get("Top issue Candidiate")
    gray_fill = PatternFill("solid", fgColor="C0C0C0")

    if tags_idx and top_idx:
        for r in range(2,max_row+1):
            tags_val = ws.cell(r, tags_idx).value or ""
            top_cell = ws.cell(r, top_idx)
            top_val = top_cell.value or ""
            if "IPN_CN_TopIssue" in tags_val:
                if top_val == "":
                    top_cell.value = "Yes"
                    top_cell.fill = new_fill
            else:
                if top_val == "Yes":
                    top_cell.fill = gray_fill
    #rejected ticket
    br_idx  = header2col.get("Blocking reason")
    ph_idx  = header2col.get("Phase")
    rej_idx = header2col.get("Rejected ticket")

    if br_idx and ph_idx and rej_idx:
        for r in range(2, ws.max_row + 1):
            br_val = ws.cell(r, br_idx).value
            ph_val = str(ws.cell(r, ph_idx).value or "")
            ws.cell(r, rej_idx).value = 1 if (br_val not in (None, "") and "New" in ph_val) else 0
    
    # 保存
    wb.save(orig_fp)
    print(f"[{ts}] 更新完成，保存至原表 {orig_fp}")
    time.sleep(10)
    # -------- PivotTable 自动刷新 ----------------
    _refresh_pivots_in_workbook(orig_fp, sheet)
    

class FolderHandler(FileSystemEventHandler):
    def __init__(self, folders):
        self.folders = folders

    def on_created(self, event):
        self._maybe_update(event.src_path)

    on_modified = on_created

    def on_moved(self, event):
        self._maybe_update(event.dest_path)

    def _maybe_update(self, path):
        if os.path.isdir(path):
            return
        fld = os.path.basename(os.path.dirname(path))
        if fld in (self.folders['jira_dir'], self.folders['octane_dir']):
            update_excel(path)


def main():
    for d in (ORIG_DIR, JIRA_DIR, OCTANE_DIR):
        os.makedirs(d, exist_ok=True)

    observer = Observer()
    handler  = FolderHandler(folders)
    observer.schedule(handler, path=JIRA_DIR, recursive=False)
    observer.schedule(handler, path=OCTANE_DIR, recursive=False)
    observer.start()
    print(f"监控：{JIRA_DIR}, {OCTANE_DIR}")

    '''schedule.every().day.at("18:00").do(lambda:
        [update_excel(os.path.join(JIRA_DIR,fn)) for fn in os.listdir(JIRA_DIR) if fn.lower().endswith((".xlsx",".csv"))] +
        [update_excel(os.path.join(OCTANE_DIR,fn)) for fn in os.listdir(OCTANE_DIR) if fn.lower().endswith((".xlsx",".csv"))]
    )
    print("已设定每日 18:00 全量扫描更新。")'''

    try:
        while True:
            schedule.run_pending()
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

if __name__ == "__main__":
    main()
