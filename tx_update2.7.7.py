import os, json, pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# === 加载配置 ===
with open("config.json", "r", encoding="utf-8") as f:
    cfg = json.load(f)

# 路径 & Sheet
original_file = cfg["paths"]["original_file"]
new_file      = cfg["paths"]["new_file"]
updated_file  = cfg["paths"]["updated_file"]
target_sheet  = cfg["sheet"]["target_sheet"]

# 映射字典
mapping                  = cfg["column_mapping"]
fund_function_mapping    = cfg["fund_function_mapping"]
owner_root_cause_mapping = cfg["owner_root_cause_mapping"]

# 读取原表数据 & 新表数据（用于值提取）
df_new = pd.read_excel(new_file)
# -------------------------------
# 1. 用 openpyxl 打开原表，准备修改
# -------------------------------
wb = load_workbook(original_file)
ws = wb[target_sheet]

# 1.1 构造表头到列号的映射
header2col = { ws.cell(1, c).value: c for c in range(1, ws.max_column+1) }

# 1.2 构造原表 ID->行号 映射，并提取它们的超链接
id_col = header2col.get("ID")
id2row    = {}
id2url_orig = {}
for r in range(2, ws.max_row+1):
    cell = ws.cell(r, id_col)
    val = cell.value
    if val is not None:
        id2row[val] = r
        if cell.hyperlink:
            id2url_orig[val] = cell.hyperlink.target

# -------------------------------
# 2. 从新表中提取 ID->URL
# -------------------------------
new_wb  = load_workbook(new_file)
new_ws  = new_wb.active
# 找新表里 ID 列号
id_col_new = next(
    (c for c in range(1, new_ws.max_column+1)
     if new_ws.cell(1,c).value=="ID"), None)

id2url_new = {}
if id_col_new:
    for r in range(2, new_ws.max_row+1):
        cell = new_ws.cell(r, id_col_new)
        if cell.hyperlink:
            id2url_new[cell.value] = cell.hyperlink.target

# -------------------------------
# 3. 定义高亮样式
# -------------------------------
yellow_fill = PatternFill("solid", fgColor="FFFF00")
green_fill  = PatternFill("solid", fgColor="00FF00")

# -------------------------------
# 4. 遍历新表，更新或追加
# -------------------------------
for _, new_row in df_new.iterrows():
    new_id = new_row.get("ID", "")
    if new_id in id2row:
        # —— 更新原有行 —— 
        r = id2row[new_id]
        for new_key, orig_key in mapping.items():
            if new_key in new_row and pd.notna(new_row[new_key]) and new_row[new_key] != "":
                c = header2col.get(orig_key)
                if c:
                    cell = ws.cell(r, c)
                    cell.value = new_row[new_key]
                    cell.fill  = green_fill
        # 如果新表里对 ID 本身也有新的超链接，可按需更新：
        if new_id in id2url_new:
            cell = ws.cell(r, id_col)
            cell.hyperlink = id2url_new[new_id]
            cell.style     = "Hyperlink"
    else:
        # —— 追加新行 —— 
        # 构建一行与原表列数一致的列表
        row_values = []
        for col_name in header2col:
            # 找到 new_key=>orig_key 里哪个 orig_key==col_name
            for new_key, orig_key in mapping.items():
                if orig_key == col_name:
                    v = new_row.get(new_key, "")
                    if pd.isna(v): v = ""
                    row_values.append(v)
                    break
            else:
                row_values.append("")
        ws.append(row_values)
        new_r = ws.max_row
        # 高亮整行
        for c in range(1, ws.max_column+1):
            ws.cell(new_r, c).fill = yellow_fill
        # 给 ID 单元格加超链接（如果有）
        if new_id in id2url_new:
            cell = ws.cell(new_r, id_col)
            cell.hyperlink = id2url_new[new_id]
            cell.style     = "Hyperlink"
        # 并记住新增行的行号，可以后续公式处理

# -------------------------------
# 5. 会后续你原先的“Days”“Octane or Jira”“Open >20 days”“No TIS”公式
#    以及 fund_function_mapping、owner_root_cause_mapping 的处理
#    这里略，和之前示例里相同 —— 对 ws 进行相应列的写值或写公式即可。
# -------------------------------

wb.save(updated_file)
print("更新完成，ID 的超链接已保留/重建，存在的行单元格已染绿，其它已追加行染黄。")
