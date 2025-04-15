import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -------------------------------
# 1. 文件路径及参数设置
# -------------------------------
original_file = "original.xlsx"      # 原始 Excel 文件，包含多个 sheet
new_file = "new.xlsx"               # 新数据 Excel 文件
target_sheet = "Octane and jira"    # 目标更新的工作表名称
updated_file = "updated_excel.xlsx" # 更新后的文件保存路径

# 定义映射关系：新表列名 -> 原表对应的列名
mapping = {
    "Name": "名字",
    "age": "年龄",
    "sex": "性别"
}

# -------------------------------
# 2. 读取原表中目标工作表的数据（保留原有数据）
# -------------------------------
df_orig = pd.read_excel(original_file, sheet_name=target_sheet)

# -------------------------------
# 3. 读取新表数据，并构造与原表结构一致的新行数据
# -------------------------------
df_new = pd.read_excel(new_file)

# 获取目标工作表的所有列（原表结构）
orig_columns = list(df_orig.columns)

# 构造新行数据列表，每个字典的 key 均为原表的列
new_rows = []
for idx, new_row in df_new.iterrows():
    row_data = {}
    for col in orig_columns:
        # 遍历每一列，判断该列是否是映射关系的目标列
        matched = False
        for new_key, orig_key in mapping.items():
            if col == orig_key:  # 当前原表列在映射表中
                # 如果新表中有该列，则取值，否则填空
                value = new_row.get(new_key, "")
                if pd.isna(value):
                    value = ""
                row_data[col] = value
                matched = True
                break
        if not matched:
            # 原表中该列不在映射关系中，则填入空值
            row_data[col] = ""
    new_rows.append(row_data)

# 将新行数据生成 DataFrame，与原表结构保持一致
df_new_rows = pd.DataFrame(new_rows, columns=orig_columns)

# -------------------------------
# 4. 将新数据追加到原表数据的末尾，不更改原有数据
# -------------------------------
df_updated_target = pd.concat([df_orig, df_new_rows], ignore_index=True)

# -------------------------------
# 5. 保留原表中其他sheet，不更改原有数据，仅更新目标sheet
# -------------------------------
# 利用 openpyxl 载入原始工作簿（包含多个sheet）
wb = load_workbook(original_file)

# 如果目标 sheet 已存在，则先将其删除（后面写入更新的数据）
if target_sheet in wb.sheetnames:
    ws_target = wb[target_sheet]
    wb.remove(ws_target)

# 利用 ExcelWriter 将更新后的目标 sheet 写入到原工作簿中，保留其他 sheet 不变
with pd.ExcelWriter(updated_file, engine="openpyxl") as writer:
    writer.book = wb
    # 为确保已加载其他 sheet，构造一个 sheets 字典
    writer.sheets = {ws.title: ws for ws in wb.worksheets}
    # 将更新后的目标 sheet写入工作簿
    df_updated_target.to_excel(writer, sheet_name=target_sheet, index=False)
    writer.save()

# -------------------------------
# 6. 对追加的新行进行高亮显示
# -------------------------------
# 新数据在目标 sheet 中的位置：原有数据行数之后，新行开始
wb_updated = load_workbook(updated_file)
ws = wb_updated[target_sheet]

highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 注意：Excel 第一行为表头，因此数据行从第2行开始
original_rows = df_orig.shape[0]  # 原有数据行数（不含表头）
start_row = original_rows + 2      # 新数据的起始行号

# 高亮新增行区域
for row in range(start_row, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).fill = highlight_fill

wb_updated.save(updated_file)
print(f"更新完成！新数据已按照映射关系追加到工作表 '{target_sheet}' 的末尾，新追加行已用黄色高亮显示，结果保存在 {updated_file}。")
