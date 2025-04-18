import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# === 辅助函数：清理尾部空行 & 定位最后有数据行 ===
def is_row_blank(ws, row):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True

def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break

def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            return r
    return 1

# === 0. 加载配置 ===
with open("config.json", "r", encoding="utf-8") as f:
    cfg = json.load(f)

original_file            = cfg["paths"]["original_file"]
new_file                 = cfg["paths"]["new_file"]
updated_file             = cfg["paths"]["updated_file"]
target_sheet             = cfg["sheet"]["target_sheet"]
mapping                  = cfg["column_mapping"]
fund_function_mapping    = cfg["fund_function_mapping"]
owner_root_cause_mapping = cfg["owner_root_cause_mapping"]

# === 1. 读取新表数据 ===
df_new = pd.read_excel(new_file)

# === 2. 打开原表 ===
wb = load_workbook(original_file)
ws = wb[target_sheet]

# 2.1 清理尾部空行
trim_trailing_blank_rows(ws)

# 2.2 构建表头→列号映射
header2col = { ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) }

# 2.3 提取原表 ID 行号、超链接
id_col = header2col["ID"]
id2row = {}
id2url_orig = {}
for r in range(2, ws.max_row + 1):
    cell = ws.cell(r, id_col)
    val = cell.value
    if val is not None:
        id2row[val] = r
        if cell.hyperlink:
            id2url_orig[val] = cell.hyperlink.target

# === 3. 提取新表 ID→超链接 ===
new_wb = load_workbook(new_file)
new_ws = new_wb.active
id_col_new = next((c for c in range(1, new_ws.max_column+1)
                   if new_ws.cell(1, c).value == "ID"), None)
id2url_new = {}
if id_col_new:
    for r in range(2, new_ws.max_row + 1):
        nc = new_ws.cell(r, id_col_new)
        if nc.hyperlink:
            id2url_new[nc.value] = nc.hyperlink.target

# 样式 & 表头列表
green_fill  = PatternFill("solid", fgColor="00FF00")
yellow_fill = PatternFill("solid", fgColor="FFFF00")
headers     = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

# 准备追加行起始
last_row = find_last_data_row(ws, id_col)

# === 4. 遍历新表，更新或追加 ===
for _, new_row in df_new.iterrows():
    new_id = new_row.get("ID", "")
    if not new_id:
        continue

    if new_id in id2row:
        # —— 更新已有行 —— 
        r = id2row[new_id]
        if new_id in id2url_new:
            cell = ws.cell(r, id_col)
            cell.hyperlink = id2url_new[new_id]
            cell.style     = "Hyperlink"

        for nk, ok in mapping.items():
            if nk not in new_row:
                continue
            new_val = new_row[nk]
            if pd.isna(new_val) or new_val == "":
                continue

            # **特殊替换：Involved I-Step 前缀**
            if nk == "Involved I-Step":
                s = str(new_val)
                if s.startswith("G070") or s.startswith("U006"):
                    new_val = "NA05" + s[4:]

            c = header2col.get(ok)
            if not c:
                continue
            cell = ws.cell(r, c)
            if cell.value != new_val:
                cell.value = new_val
                cell.fill  = green_fill

        # Function 列
        if "Found in function" in header2col and "Function" in header2col:
            fund_val = ws.cell(r, header2col["Found in function"]).value or ""
            for k, v in fund_function_mapping.items():
                if k in fund_val:
                    c = header2col["Function"]
                    cell = ws.cell(r, c)
                    if cell.value != v:
                        cell.value = v
                        cell.fill  = green_fill
                    break

        # Root cause 列
        if "Owner" in header2col and "Root cause" in header2col:
            ow = ws.cell(r, header2col["Owner"]).value or ""
            for k, v in owner_root_cause_mapping.items():
                if k in ow:
                    c = header2col["Root cause"]
                    cell = ws.cell(r, c)
                    if cell.value != v:
                        cell.value = v
                        cell.fill  = green_fill
                    break

    else:
        # —— 追加新行 —— 
        last_row += 1
        for idx, hdr in enumerate(headers, start=1):
            val = ""
            for nk, ok in mapping.items():
                if ok == hdr:
                    tmp = new_row.get(nk, "")
                    if pd.notna(tmp) and tmp != "":
                        # 前缀替换
                        if nk == "Involved I-Step":
                            s = str(tmp)
                            if s.startswith("G070") or s.startswith("U006"):
                                tmp = "NA05" + s[4:]
                        val = tmp
                    break
            ws.cell(row=last_row, column=idx).value = val
            if val != "":
                ws.cell(row=last_row, column=idx).fill = yellow_fill

        # ID 超链接
        if new_id in id2url_new:
            cell = ws.cell(last_row, id_col)
            cell.hyperlink = id2url_new[new_id]
            cell.style     = "Hyperlink"

        # Function & Root cause
        if "Found in function" in header2col and "Function" in header2col:
            fund_val = new_row.get("Found in function", "") or ""
            for k, v in fund_function_mapping.items():
                if k in fund_val:
                    c = header2col["Function"]
                    ws.cell(last_row, c).value = v
                    ws.cell(last_row, c).fill  = yellow_fill
                    break
        if "Owner" in header2col and "Root cause" in header2col:
            ow = new_row.get("Owner", "") or ""
            for k, v in owner_root_cause_mapping.items():
                if k in ow:
                    c = header2col["Root cause"]
                    ws.cell(last_row, c).value = v
                    ws.cell(last_row, c).fill  = yellow_fill
                    break

# === 5. 填充公式和其它列 ===
max_row = ws.max_row

# 5.1 Days
days_idx     = header2col.get("Days")
creation_idx = header2col.get("Creation time")
if days_idx and creation_idx:
    creation_letter = get_column_letter(creation_idx)
    for r in range(2, max_row + 1):
        ws.cell(r, days_idx).value = f'=DATEDIF(${creation_letter}{r},TODAY(),"D")'

# 5.2 Octane or Jira
oir_idx = header2col.get("Octane or Jira")
if oir_idx:
    fn  = os.path.basename(new_file)
    val = "Octane" if "Octane" in fn else ("Jira" if "Jira" in fn else "")
    for r in range(last_row, max_row + 1):
        ws.cell(r, oir_idx).value = val

# 5.3 Open > 20 days
open20_idx = header2col.get("Open >20 days") or header2col.get("Open > 20 days")
if open20_idx and days_idx:
    dl = get_column_letter(days_idx)
    for r in range(2, max_row + 1):
        ws.cell(r, open20_idx).value = f'=IF({dl}{r}>20,1,0)'

# 5.4 No TIS
nt_idx = header2col.get("No TIS")
pl_idx = header2col.get("Planned closing version")
ti_idx = header2col.get("Target I-Step:")
if nt_idx and pl_idx and ti_idx:
    pl_l = get_column_letter(pl_idx)
    ti_l = get_column_letter(ti_idx)
    for r in range(2, max_row + 1):
        ws.cell(r, nt_idx).value = f'=IF(OR({pl_l}{r}<>"",{ti_l}{r}<>""),0,1)'

# === 6. 保存结果 ===
wb.save(updated_file)
print("更新完成！空行已清理，Involved I-Step 前缀替换，差异染绿，新增染黄，超链接保留，结果保存在", updated_file)
