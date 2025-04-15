import os
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -------------------------------
# 1. 文件路径及参数设置
# -------------------------------

# 原始 Excel 文件（模板），包含多个 sheet
original_file = r"data\Ticket summary.xlsx"
# 新数据 Excel 文件（待追加的新表）
new_file = r"data\Octane_defects_filtered_4_10_2025_10_50_18_AM.xlsx"

# 目标更新的工作表名称（在原始 Excel 文件中）
target_sheet = "Octane and jira"
# 更新后的文件保存路径
updated_file = "tx_updated_excel2.2.xlsx"

# 定义映射关系：用于构造新行数据（新表列名 -> 原表对应的列名）
mapping = {
    "ID": "ID",
    "Ticket no. supplier": "Ticket no. supplier",
    "Name": "Name",
    "Closed in version": "Closed in version",
    "Involved I-Step": "Involved I-Step",
    "First use/SoP of function": "First use/SoP of function",
    "Creation time": "Creation time",
    "Error occurrence": "Error occurrence",
    "Phase Group": "Phase",
    "Found in function": "Found in function",
    "Defect finder": "Defect finder",
    "Owner": "Owner",
    "Involved I-Step": "Target I-Step:",
    "Target Week": "Follow up",
    "Planned closing version": "Planned closing version"
    # 其它字段如 "Top issue Candidate"、"Reporting relevance"、"Comment" 等可根据需要添加
}

# 此处忽略文件：避免更新时处理原始模板或更新后文件
ignored_files = { os.path.basename(original_file), os.path.basename(updated_file) }

# -------------------------------
# 2. 设置新增/删除列的功能参数
# -------------------------------

# 如需删除原表中的某些列，列出名称（必须与原表中的完全匹配）
columns_to_remove = [
    # 例如："SomeColumnName", "AnotherColumn"
]

# 如需增加新列，可通过字典方式指定：键为列名称，值为默认值
columns_to_add = {
    # 例如："新列1": "",
    #          "测试列": 0
}

# -------------------------------
# 3. 定义“Fund in Function”到“Function”映射关系
# -------------------------------
# 如果原表中的“Fund in Function”列存在特定内容，则“Function”列应填写对应值
fund_function_mapping = {
    "adapt speed to route geometry[01.02.02.15.02.14]": "ASRG",
    "change lane [01.02.02.15.02.07]":"CL",
    "allow hands-off driving 130 [01.02.02.15.02.01]":"HOO130",
    "keep distance [01.02.02.15.02.11]":"KD",
    "keep lane [01.02.02.15.02.10] or keep lane extended [01.02.02.15.02.08]":"KL/KLE",
    "display assisted view [01.02.02.15.02.20]":"Adview",
    "stop and go at traffic lights [01.02.02.15.02.06]":"SGTL",
    "Speed Limit Info [SLI] (incl. No Passing Info)   [01.02.02.09.09.01]":"SLI",
    "indicate traffic sign and lights [01.02.02.15.03.01]":"TSLI",
    "ADAS  Interaction with Navigation [01.04.03.01.03.01.01.04] and allow hands-off driving 130 [01.02.02.15.02.01]":"SAM-China",
    "Environment Detection for PA [01.02.02.15.04.03.01.03] or Parking Assistant [01.02.02.15.04.03.01]":"Parking",
    "stop and go at right of way situations [01.02.02.15.02.04]":"SGROW",
    # "some text": "YYY"
}

# -------------------------------
# 4. 读取原表数据（目标工作表）和新表数据，构造追加的新行
# -------------------------------

# 读取原始 Excel 中目标 sheet 的数据（原有数据保持不变）
df_orig = pd.read_excel(original_file, sheet_name=target_sheet)
# 读取新表数据
df_new = pd.read_excel(new_file)

# 获取目标工作表的所有列（原表结构）
orig_columns = list(df_orig.columns)

# 构造新行数据列表，每个字典的 key 均为原表的列
new_rows = []
for idx, new_row in df_new.iterrows():
    row_data = {}
    for col in orig_columns:
        matched = False
        # 遍历 mapping，看原表的列是否在映射关系中出现（即存在于 mapping 的 value 中）
        for new_key, orig_key in mapping.items():
            if col == orig_key:
                # 从新表中根据 new_key 提取数据，如果没有数据或数据为NaN则置空
                value = new_row.get(new_key, "")
                if pd.isna(value):
                    value = ""
                row_data[col] = value
                matched = True
                break
        if not matched:
            # 若原表该列不在映射关系中，填入空字符串
            row_data[col] = ""
    new_rows.append(row_data)

# 将新行数据生成 DataFrame，与原表结构保持一致
df_new_rows = pd.DataFrame(new_rows, columns=orig_columns)

# -------------------------------
# 5. 将新数据追加到原表的末尾
# -------------------------------
df_updated_target = pd.concat([df_orig, df_new_rows], ignore_index=True)

# -------------------------------
# 6. 根据“Fund in Function”字段内容更新“Function”列
# -------------------------------
# 如果原表中包含这两个列，则遍历每一行进行赋值
if "Fund in Function" in df_updated_target.columns and "Function" in df_updated_target.columns:
    def update_function(row):
        fund_value = row.get("Fund in Function", "")
        # 如果 fund_value 存在于映射字典中，则设置 Function 列为对应值
        if fund_value in fund_function_mapping:
            row["Function"] = fund_function_mapping[fund_value]
        return row
    df_updated_target = df_updated_target.apply(update_function, axis=1)

# -------------------------------
# 7. 对 DataFrame 进行列的删减/新增操作
# -------------------------------
# 删除指定的列（如果有）
if columns_to_remove:
    df_updated_target.drop(columns=columns_to_remove, inplace=True, errors="ignore")
# 新增列（如果指定的列不存在则添加，默认值由 columns_to_add 提供）
for col_name, default_val in columns_to_add.items():
    if col_name not in df_updated_target.columns:
        df_updated_target[col_name] = default_val

# -------------------------------
# 8. 将更新后的数据写入新的 Excel 文件，同时保留原表中其它sheet
# -------------------------------
from openpyxl import load_workbook

# 加载原始工作簿（包含多个sheet）
wb = load_workbook(original_file)
# 删除目标 sheet（后续写入更新后的数据）
if target_sheet in wb.sheetnames:
    ws_target = wb[target_sheet]
    wb.remove(ws_target)

# 利用 ExcelWriter 写入更新后的目标 sheet，同时保留其他 sheet 不变
with pd.ExcelWriter(updated_file, engine="openpyxl") as writer:
    writer.book = wb
    writer.sheets = {ws.title: ws for ws in wb.worksheets}
    df_updated_target.to_excel(writer, sheet_name=target_sheet, index=False)
    writer.save()

# -------------------------------
# 9. 对追加的新行进行黄色高亮显示
# -------------------------------
wb_updated = load_workbook(updated_file)
ws = wb_updated[target_sheet]
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Excel 第一行为表头，因此数据行从第2行开始
original_rows = df_orig.shape[0]  # 原有数据行数（不含表头）
start_row = original_rows + 2      # 新数据的起始行号

for row in range(start_row, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).fill = highlight_fill

wb_updated.save(updated_file)
print(f"更新完成！新数据已追加到工作表 '{target_sheet}' 的末尾，已根据“Fund in Function”更新Function列，并对新增行用黄色高亮显示。\n结果保存在 {updated_file}。")
