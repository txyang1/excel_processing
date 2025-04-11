import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ==================================================
# 1. 定义纠错函数（例如对 Defect finder 的名称进行更正）
# ==================================================
def error_correction(value, column_name):
    """
    针对特定字段进行数据纠错，例如纠正 "Defect finder" 的错误名称。
    你可以根据需要添加更多字段的纠错规则。
    """
    if column_name == "Defect finder":
        # 定义更正映射，注意键值区分大小写
        correction_mapping = {
            "finderA": "FinderA",
            "FiderA": "FinderA",
            "finderB": "FinderB",
            # 其他可能的错误映射……
        }
        # 将原始值转换成统一格式（例如转成小写比较），再查找是否需要更正
        corrected = correction_mapping.get(value, value)
        return corrected
    # 对于其他字段，目前直接返回原值，必要时可以扩展
    return value

# ==================================================
# 2. 读取现有总表（updated_excel.xlsx）和数据源文件
# ==================================================
# 更新文件名
updated_file = "excel_code/results/updated_excel.xlsx"
df_old = pd.read_excel(updated_file)

# 读取数据源
df_defect = pd.read_excel("excel_code/data/Defect_data2.xlsx")
df_jira = pd.read_excel("excel_code/data/Jira2.xlsx")

# ==================================================
# 3. 根据唯一标识符（例如 "ID" 和 "issue id"）合并 Defect_data 和 Jira 数据
# ==================================================
# 此处假设：Defect_data 的 "ID" 与 Jira 的 "issue id" 一一对应
df_merged = pd.merge(df_defect, df_jira, left_on="ID", right_on="issue id", how="left")

# ==================================================
# 4. 构造新的总表 DataFrame
# ==================================================
# 根据总表 updated_excel 中的字段定义，构造映射字典：
mapping_dict = {
    "Function": "Assigned ECU Group",                   # 来自 Defect_data
    "ID": "ID",                                          # 同名字段
    "Ticket no. supplier": "Ticket no.supplier",         # 来自 Defect_data
    "Name": "Name",                                      # 来自 Defect_data
    "Closed in Version": "Closed in version",            # 来自 Defect_data
    "Involved I-Step": "Invoved I-step",                 # 来自 Defect_data
    "First use/Sop of function": "First use/SoP of function",  # 来自 Defect_data
    "Top issue Candidate": "Issue key",                  # 来自 Jira
    "Reporting relevance": "Reporting relevance",        # 来自 Defect_data
    "Creation time": "Creation time",                    # 来自 Defect_data
    "Error occurrence": "Error occurrence",              # 来自 Defect_data
    "duplicated": "Inward issue link(Duplicate)",        # 来自 Jira
    "Phase": "Phase",                                    # 来自 Defect_data
    "Found in function": "Found in function",            # 来自 Defect_data
    "Defect finder": "Defect finder",                    # 来自 Defect_data，需要纠错
    "Owner": "Owner",                                    # 来自 Defect_data
    "Comment": "Summary",                                # 来自 Jira
    "Target I-Step:": "Target I-Step",                    # 来自 Defect_data
    "Follow up": "Updated",                              # 来自 Jira
    "Planned closing version": "Planned closing version",# 来自 Defect_data
    "Root cause": "",                                    # 无数据，默认空
    "Days": "Days in phase",                             # 来自 Defect_data
    "open >20 days": "Days in phase",                    # 根据 Days in phase 判断，大于20标记为 "Yes"，否则 "No"
    "No TIS, Octane or Jira": "",                        # 无数据，默认空
    "Days in the phase": "Days in phase"                 # 来自 Defect_data
}

# 构造新的总表数据时，保存一个字典以便后续转 DataFrame
new_data = {}

for target_field, source_field in mapping_dict.items():
    # 针对 "open >20 days" 字段特殊处理
    if target_field == "open >20 days":
        if source_field in df_merged.columns:
            new_data[target_field] = df_merged[source_field].apply(lambda x: "Yes" if x > 20 else "No")
        else:
            new_data[target_field] = ["No"] * len(df_merged)
    # 如果该目标字段没有对应来源，则填充空字符串
    elif source_field == "":
        new_data[target_field] = ""
    else:
        # 对于 "Defect finder" 字段，先进行纠错处理
        if target_field == "Defect finder":
            new_data[target_field] = df_merged[source_field].apply(lambda x: error_correction(x, "Defect finder"))
        else:
            new_data[target_field] = df_merged[source_field]

# 构造新总表 DataFrame
df_new = pd.DataFrame(new_data)

# ==================================================
# 5. 将新总表与原有总表做比较，更新和添加新行
# ==================================================
# 假设 "ID" 是唯一标识符，使用它做对比
# 记录更新的单元格位置：(行号, 列名称)
updated_cells = []

# 将 df_old 的 "ID" 列转换成集合，方便判断新增记录
old_ids = set(df_old["ID"].tolist())

# 用 new_total 中的数据更新已存在的记录
for idx_new, row_new in df_new.iterrows():
    id_val = row_new["ID"]
    # 如果该记录已存在，则检查每个字段是否变化
    if id_val in old_ids:
        # 定位旧记录行的索引
        idx_old = df_old[df_old["ID"] == id_val].index[0]
        for col in df_new.columns:
            new_val = row_new[col]
            old_val = df_old.at[idx_old, col]
            # 如果发现不一致，则更新 df_old，并记录位置
            if pd.isnull(new_val) and pd.isnull(old_val):
                continue
            if new_val != old_val:
                df_old.at[idx_old, col] = new_val
                updated_cells.append((idx_old, col))
    else:
        # 新记录不存在，直接添加到 df_old
        df_old = df_old.append(row_new, ignore_index=True)
        # 记录所有该行的所有列为更新（因为这是新增行）
        new_row_idx = df_old.shape[0] - 1
        for col in df_new.columns:
            updated_cells.append((new_row_idx, col))

# ==================================================
# 6. 保存更新后的总表到文件，并高亮标记更新内容
# ==================================================
df_old.to_excel(updated_file, index=False)

# 使用 openpyxl 对更新的单元格高亮
wb = load_workbook(updated_file)
ws = wb.active

# 设置高亮填充（黄色）
highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 获取 Excel 表头，对应列的位置（openpyxl 中行列均从 1 开始，第一行为表头）
header = [cell.value for cell in ws[1]]

for row_idx, col_name in updated_cells:
    # pandas 行号从0开始，Excel 行号为 pandas 行号 + 2（1 行为表头，另 1 表示行号偏移）
    excel_row = row_idx + 2
    if col_name in header:
        col_idx = header.index(col_name) + 1  # Excel 列号从 1 开始
        ws.cell(row=excel_row, column=col_idx).fill = highlight_fill

wb.save(updated_file)
print(f"已更新 {updated_file}，并高亮显示了更新的单元格。")
