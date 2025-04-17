import os, json, pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# === 一次性加载配置 ===
with open("config.json", "r", encoding="utf-8") as f:
    cfg = json.load(f)

# 文件路径
original_file = cfg["paths"]["original_file"]
new_file      = cfg["paths"]["new_file"]
updated_file  = cfg["paths"]["updated_file"]

# Sheet 名称
target_sheet  = cfg["sheet"]["target_sheet"]

# 各种映射
mapping                 = cfg["column_mapping"]
fund_function_mapping   = cfg["fund_function_mapping"]
owner_root_cause_mapping= cfg["owner_root_cause_mapping"]

# === 然后你的其余脚本保持不变，只需使用上述变量 ===

# 示例：读取原表和新表
df_orig = pd.read_excel(original_file, sheet_name=target_sheet)
df_new  = pd.read_excel(new_file)

# … 其余追加、映射、写文件、高亮、公式都使用 original_file/new_file/updated_file/target_sheet 变量 …
