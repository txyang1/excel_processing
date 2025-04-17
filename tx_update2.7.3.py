import os
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# -------------------------------
# 1. 文件路径及参数设置
# -------------------------------
# 原始 Excel 文件（模板），包含多个 sheet
original_file = r"data\Ticket summary.xlsx"
# 新数据 Excel 文件（待追加的新表）
new_file = r"data\Octane_defects_filtered_4_10_2025_10_50_18_AM.xlsx"

# 目标更新的工作表名称（在原始 Excel 文件中）
target_sheet = "Octane and jira"
# 更新后的文件保存路径
updated_file = "tx_updated_excel2.7.3.xlsx"

# 定义映射关系：用于构造新行数据（新表列名 -> 原表对应的列名）
mapping = {
    "ID": "ID",
    "Ticket no. supplier": "Ticket no. supplier",
    "Name": "Name",
    "Closed in version": "Closed in version",
    "Involved I-Step": "Involved I-Step",
    "First use/SoP of function": "First use/SoP of function",
    "Creation time": "Creation time",
    "Error occurrence": "Error occurrence",
    "Phase Group": "Phase",
    "Found in function": "Found in function",
    "Defect finder": "Defect finder",
    "Owner": "Owner",
    "Involved I-Step": "Target I-Step:",
    "Target Week": "Follow up",
    "Planned closing version": "Planned closing version",
    "Days in phase": "Days in the pahse"
    # 其它字段如 "Top issue Candidate"、"Reporting relevance"、"Comment" 可按需添加
}

# 忽略文件，避免更新时处理原始模板或更新后文件
ignored_files = { os.path.basename(original_file), os.path.basename(updated_file) }

# -------------------------------
# 2. 设置新增/删除列的功能参数
# -------------------------------
columns_to_remove = [
    # 例如："SomeColumnName", "AnotherColumn"
]
columns_to_add = {
    # 例如："新列1": "", "测试列": 0
}

# -------------------------------
# 3. 定义“Fund in Function”到“Function”映射关系
# -------------------------------
fund_function_mapping = {
    #ASRG
    "adapt speed to route geometry [01.02.02.15.02.14]": "ASRG",
    #CL
    "change lane [01.02.02.15.02.07]": "CL",
    "Speed Limit Info 21.0 Mid [SLI21.0_Mi] [01.02.02.02.20]": "CL",
    #HOO130
    "allow hands-off driving 130 [01.02.02.15.02.01]": "HOO130",
    "allow hands-off driving 130 [01.02.02.15.02.01], allow hands-off driving 60 [01.02.02.15.02.02]": "HOO130",
    #KD
    "keep distance [01.02.02.15.02.11]": "KD",
    #KL/KLE
    "keep lane [01.02.02.15.02.10]": "KL/KLE",
    "keep lane extended [01.02.02.15.02.08]": "KL/KLE",
    "BS2": "Kl/KLE",
    "Motion Planning [01.02.02.15.02.08.07], Motion Planning [01.02.02.15.02.10.08]": "Kl/KLE",
    "Processing SRR - Detection  [01.02.01.02.08.03.01.13.05.02]": "Kl/KLE",
    #Asview
    "display assisted view [01.02.02.15.02.20]": "Adview",
    #SGTL
    "stop and go at traffic lights [01.02.02.15.02.06]": "SGTL",
    #SLI
    "Speed Limit Info [SLI] (incl. No Passing Info)  [01.02.02.09.09.01]": "SLI",
    "ADAS  Interaction with Navigation [01.04.03.01.03.01.01.04], FKT_display_settings_Speed_Limit_Info_SLI21.0_basis [01.02.02.01.03.01.01.03.02.03], Provide Navigation 2.0 [01.04.03.01.03.06]": "SLI",
    "ADAS  Interaction with Navigation [01.04.03.01.03.01.01.04], FKT_display_settings_Speed_Limit_Info_SLI21.0_basis [01.02.02.01.03.01.03.02.02.04.02.03], Provide Navigation 2.0 [01.04.03.01.03.06]": "SLI",
    #TSLI
    "indicate traffic sign and lights [01.02.02.15.03.01]": "TSLI",
    #SAM-China
    "ADAS  Interaction with Navigation [01.04.03.01.03.01.01.04],allow hands-off driving 130 [01.02.02.15.02.01]": "SAM-China",
    #Parking
    "Environment Detection for PA [01.02.02.15.04.03.01.03]": "Parking",
    "Parking Assistant [01.02.02.15.04.03.01]": "Parking",
    #SGROW
    "stop and go at right of way situations [01.02.02.15.02.04]": "SGROW",
    #Kufu function test
    "Autonomous Emergency Braking [01.02.02.15.03.03]": "Kufu function test",
    "Autonomous Emergency Braking [01.02.02.15.03.03]": "Kufu function test",
    #V2X
    "Implement basic platform of Vehicle-to-Everything (V2X)": "V2X",
    #Safty
    "Processing SRR - Detection  [01.02.01.02.08.01.02.13.05.02]": "Safty",
    #HOO-SAM
    "ADAS  Interaction with Navigation [01.04.03.01.03.01.01.04]": "HOO-SAM",
    "ADAS  Interaction with Navigation [01.04.03.01.03.01.01.04], Steering and Lane Control Assistant 3 [LSA3] [01.02.02.12.02.02]": "HOO-SAM",

}

#3.2 Owner 到 Root case 的映射
Owner_Root_case_mappig = {
 "Niklas Haeuser" :"Condition evaluate",
 "Ruomeng Guan": "HPL",
 "Cristina delVal": "Icon issue",
 "Daniel Albetal": "IKS",
 "Christoph Romainczyk": "Kufu function issue",
 "Aifa Zhou": "Map issue",
 "Zed Zhang": "Motion plan",
 "zedzhang": "Motion plan",
 "Matthias Stark": "Object fusion",
 "JianLin Zhang": "Obstacle",
 "Fabiao Wang": "Road model",
 "Han Jia": "Road strategy",
 "hanjia": "Road strategy",
 "Juan Carlos Fuentes Michel": "SRR/MRR",
 "juan-carlosfuentes-michel": "SRR/MRR",

}

# -------------------------------
# 4. 读取原表数据（目标工作表）和新表数据，构造追加的新行
# -------------------------------
df_orig = pd.read_excel(original_file, sheet_name=target_sheet)
df_new = pd.read_excel(new_file)
orig_columns = list(df_orig.columns)

new_rows = []
for idx, new_row in df_new.iterrows():
    row_data = {}
    for col in orig_columns:
        matched = False
        for new_key, orig_key in mapping.items():
            if col == orig_key:
                value = new_row.get(new_key, "")
                if pd.isna(value):
                    value = ""
                row_data[col] = value
                matched = True
                break
        if not matched:
            row_data[col] = ""
    new_rows.append(row_data)

df_new_rows = pd.DataFrame(new_rows, columns=orig_columns)

# -------------------------------
# 5. 将新数据追加到原表的末尾
# -------------------------------
df_updated_target = pd.concat([df_orig, df_new_rows], ignore_index=True)

# -------------------------------
# 6. 根据“Found in function”字段内容更新“Function”列
# -------------------------------
if "Found in function" in df_updated_target.columns and "Function" in df_updated_target.columns:
    def update_function(row):
        fund_value = row.get("Found in function", "")
        if isinstance(fund_value, str):
            fund_value = fund_value.strip()
        else:
            fund_value = ""
        for mapping_key, mapped_value in fund_function_mapping.items():
            if mapping_key in fund_value:
                row["Function"] = mapped_value
                break
        return row
    df_updated_target = df_updated_target.apply(update_function, axis=1)


#6.2
if "Owner" in df_updated_target.columns and "Root cause" in df_updated_target.columns:
    def update_function(row):
        fund_value = row.get("Owner", "")
        if isinstance(fund_value, str):
            fund_value = fund_value.strip()
        else:
            fund_value = ""
        for mapping_key, mapped_value in Owner_Root_case_mappig.items():
            if mapping_key in fund_value:
                row["Root cause"] = mapped_value
                break
        return row
    df_updated_target = df_updated_target.apply(update_function, axis=1)


# -------------------------------
# 7. 对 DataFrame 进行列的删减/新增操作
# -------------------------------
if columns_to_remove:
    df_updated_target.drop(columns=columns_to_remove, inplace=True, errors="ignore")
for col_name, default_val in columns_to_add.items():
    if col_name not in df_updated_target.columns:
        df_updated_target[col_name] = default_val

# -------------------------------
# 8. 将更新后的数据写入新的 Excel 文件，同时保留原表中其它 sheet
# -------------------------------
wb = load_workbook(original_file)
if target_sheet in wb.sheetnames:
    ws_target = wb[target_sheet]
    wb.remove(ws_target)
with pd.ExcelWriter(updated_file, engine="openpyxl") as writer:
    writer.book = wb
    writer.sheets = {ws.title: ws for ws in wb.worksheets}
    df_updated_target.to_excel(writer, sheet_name=target_sheet, index=False)
    writer.save()

# -------------------------------
# 9. 对追加的新行进行黄色高亮显示
# -------------------------------
wb_updated = load_workbook(updated_file)
ws = wb_updated[target_sheet]
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
original_rows = df_orig.shape[0]  # 原有数据行数（不含表头）
start_row = original_rows + 2      # 新数据的起始行号
for row in range(start_row, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).fill = highlight_fill

# -------------------------------
# 10. 对所有数据行的 "Days" 列设置公式
# -------------------------------
days_col_index = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "Days":
        days_col_index = col
        break

if days_col_index is not None:
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=days_col_index).value = f'=DATEDIF($J{row},TODAY(),"D")'
else:
    print("警告：未找到 'Days' 列，未设置公式。")

# -------------------------------
# 11. 设置 "Octane or Jira" 列的内容
# -------------------------------
octane_or_jira_col_index = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "Octane or Jira":
        octane_or_jira_col_index = col
        break

if octane_or_jira_col_index is not None:
    base_new_file = os.path.basename(new_file)
    if "Octane" in base_new_file:
        value_to_fill = "Octane"
    elif "Jira" in base_new_file:
        value_to_fill = "Jira"
    else:
        value_to_fill = ""
    for row in range(start_row, ws.max_row + 1):
        ws.cell(row=row, column=octane_or_jira_col_index).value = value_to_fill
else:
    print("警告：未找到 'Octane or Jira' 列，无法设置对应内容。")

# -------------------------------
# 12. 对高亮部分的 "Open > 20 days" 列进行填充
# -------------------------------
open20_col_index = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "Open >20 days":
        open20_col_index = col
        break

if open20_col_index is not None and days_col_index is not None:
    # 取 Days 列的字母
    days_col_letter = get_column_letter(days_col_index)
    for row in range(start_row, ws.max_row + 1):
        ws.cell(row=row, column=open20_col_index).value = f'=IF({days_col_letter}{row}>20,1,0)'
else:
    print("警告：未找到 'Open > 20 days' 列或 'Days' 列，无法设置公式。")

# -------------------------------
# 13. 对 "No TIS" 列进行填充
#    如果 "Planned closing version" 或 "Target I-Step:" 中有一列非空，则输出0，否则1
# -------------------------------
no_tis_col_index = None
planned_col_index = None
target_i_step_col_index = None
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col).value
    if header == "No TIS":
        no_tis_col_index = col
    elif header == "Planned closing version":
        planned_col_index = col
    elif header == "Target I-Step:":
        target_i_step_col_index = col

if no_tis_col_index is not None and planned_col_index is not None and target_i_step_col_index is not None:
    planned_letter = get_column_letter(planned_col_index)
    target_letter = get_column_letter(target_i_step_col_index)
    for row in range(start_row, ws.max_row + 1):
        ws.cell(row=row, column=no_tis_col_index).value = f'=IF(OR({planned_letter}{row}<>"",{target_letter}{row}<>""),0,1)'
else:
    print("警告：未找到 'No TIS' 列或相关参照列（Planned closing version/Target I-Step:），无法设置公式。")

wb_updated.save(updated_file)
print(f"更新完成！新数据已追加到工作表 '{target_sheet}' 的末尾，\n根据 'Found in function' 更新了 Function 列，\n新增行用黄色高亮显示，\n整个 'Days' 列均设置了公式 =DATEDIF($J{{row}},TODAY(),\"D\"),\n'Octane or Jira' 列填入了对应值，\n'Open > 20 days' 列根据 Days 值设定了 1/0，\n'No TIS' 列根据 Planned closing version 和 Target I-Step: 填入了 0 或 1.\n结果保存在 {updated_file}。")


'''
Root cause###
#一对一
Condition evaluate : Owner(Niklas Haeuser)
HPL： Owner(Ruomeng Guan),
Icon issue: Owner(Cristina delVal)
IKS: Owner(Daniel Albetal)
Kufu function issue： Owner(Christoph Romainczyk)
Map issue: Owner(Aifa Zhou)
Motion plan: Owner(Zed Zhang/zedzhang)
Object fusion: Owner(Matthias Stark)
Obstacle: Owner(JianLin Zhang)
Road model: Owner(Fabiao Wang)
Road strategy: Owner(Han Jia/hanjia)
SRR/MRR: Owner(Juan Carlos Fuentes Michel/juan-carlosfuentes-michel)
#多对一
GNSS: Owner(Meijie Fu), Ticket no. supplier(存在ORIONINIT)
Platform issue: Function(Platform), ID(存在ADBK25)
Prediction： Owner(luislupartner), Function(CL),
Traffic sign fusion: Owner(Liang Xue), Function(TSLI)
CV issue : Ticket no. supplier(存在CVDRVN), Reporting relevance(Not Reporting Relevant)
'''
