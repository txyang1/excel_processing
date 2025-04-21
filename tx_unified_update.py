import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# === 辅助函数 ===
def is_row_blank(ws, row):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True

def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break

def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            return r
    return 1

# === 0. 读取配置 ===
with open("config.json", "r", encoding="utf-8") as f:
    cfg = json.load(f)

paths    = cfg["paths"]
orig_fp  = paths["original_file"]
new_fp   = paths["new_file"]
upd_fp   = paths["updated_file"]

sheet    = cfg["sheet"]["target_sheet"]
sources  = cfg["sources"]
fund_map = cfg["fund_function_mapping"]
own_map  = cfg["owner_root_cause_mapping"]

# === 1. 判断来源并加载 mapping ===
base       = os.path.basename(new_fp)
source_key = next((k for k,v in sources.items() if v["pattern"] in base), None)
if not source_key:
    raise RuntimeError("无法识别 new_file 来源")

src_cfg    = sources[source_key]
read_meth  = src_cfg["read_method"]
date_col   = src_cfg.get("date_col")
mapping    = src_cfg["mapping"]

# === 2. 读取新表 ===
df_new = pd.read_excel(new_fp) if read_meth=="excel" else pd.read_csv(new_fp)

# === 3. 打开原表，清除旧高亮 & 空行 ===
wb = load_workbook(orig_fp)
ws = wb[sheet]

# 清除之前留下的绿/黄
green_codes  = {"FF00FF00","00FF00"} 
yellow_codes = {"FFFFFF00","FFFF00"}
for row in ws.iter_rows(min_row=2, min_col=1,
                        max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        f = cell.fill
        if f.fill_type=="solid":
            rgb = f.fgColor.rgb or f.fgColor.value
            if rgb in green_codes or rgb in yellow_codes:
                cell.fill=PatternFill(fill_type=None)

trim_trailing_blank_rows(ws)

# 构建表头映射
header2col = {ws.cell(1,c).value:c for c in range(1, ws.max_column+1)}
headers    = list(header2col.keys())
id_col     = header2col["ID"]

# 记录已有 ID→行号
id2row = {
    ws.cell(r,id_col).value:r
    for r in range(2, ws.max_row+1)
    if ws.cell(r,id_col).value
}

# 如果是 Excel 来源，提取新表 ID→超链接
id2url_new = {}
if read_meth=="excel":
    nw = load_workbook(new_fp)
    nws= nw.active
    col_id = next((c for c in range(1,nws.max_column+1) 
                   if nws.cell(1,c).value=="ID"), None)
    if col_id:
        for r in range(2, nws.max_row+1):
            c = nws.cell(r,col_id)
            if c.hyperlink:
                id2url_new[c.value] = c.hyperlink.target

# 定位追加起始
last_row      = find_last_data_row(ws, id_col)
original_last = last_row

# 高亮样式
green  = PatternFill("solid", fgColor="00FF00")
yellow = PatternFill("solid", fgColor="FFFF00")

# 找出映射中 ID 对应的源列名
id_key = next(k for k,v in mapping.items() if v=="ID")

# === 4. 遍历新表，更新 or 追加 ===
for _, new_row in df_new.iterrows():
    new_id = new_row.get(id_key,"")
    if pd.isna(new_id) or not new_id:
        continue

    if new_id in id2row:
        # —— 更新已有行，只更新非 Function 列 —— 
        r = id2row[new_id]
        # 日期列（Jira）
        if date_col:
            raw = new_row.get(date_col,"")
            if pd.notna(raw) and raw!="":
                try:
                    ts = pd.to_datetime(raw)
                    dt = ts.to_pydatetime()
                    c  = header2col[mapping[date_col]]
                    cell = ws.cell(r,c)
                    if cell.value!=dt:
                        cell.value = dt
                        cell.number_format = "m/d/yyyy h:mm:ss AM/PM"
                        cell.fill = green
                except:
                    pass
        # 其它映射字段
        for nk, ok in mapping.items():
            if nk==date_col or ok=="Function":
                continue
            val = new_row.get(nk,"")
            if pd.isna(val) or val=="": continue
            # 前缀替换
            if ok=="Target I-Step:" and (str(val).startswith("G070") or str(val).startswith("U006")):
                val = "NA05"+str(val)[4:]
            c = header2col.get(ok)
            if c:
                cell = ws.cell(r,c)
                if cell.value!=val:
                    cell.value = val
                    cell.fill  = green
        # Root cause
        if "Owner" in header2col and "Root cause" in header2col:
            ow = ws.cell(r,header2col["Owner"]).value or ""
            for k,v in own_map.items():
                if k in ow:
                    c=header2col["Root cause"]
                    cell=ws.cell(r,c)
                    if cell.value!=v:
                        cell.value=v; cell.fill=green
                    break

    else:
        # —— 追加新行，包括 Function 列 —— 
        last_row += 1
        for idx, hdr in enumerate(headers, start=1):
            val = ""
            for nk, ok in mapping.items():
                if ok==hdr:
                    tmp = new_row.get(nk,"")
                    if pd.notna(tmp) and tmp!="":
                        # 日期
                        if hdr==mapping.get(date_col):
                            try:
                                ts=pd.to_datetime(tmp)
                                val=ts.to_pydatetime()
                            except:
                                val=tmp
                        else:
                            val=tmp
                        # 前缀
                        if ok=="Target I-Step:" and (str(val).startswith("G070") or str(val).startswith("U006")):
                            val="NA05"+str(val)[4:]
                    break
            cell = ws.cell(last_row, idx)
            cell.value = val
            if val!="":
                cell.fill = yellow
                if hdr==mapping.get(date_col):
                    cell.number_format="m/d/yyyy h:mm:ss AM/PM"
        # 新行 ID 超链接
        url = id2url_new.get(new_id)
        if url:
            c= id_col
            cell = ws.cell(last_row,c)
            cell.hyperlink = url
            cell.style     = "Hyperlink"
        # Function 列（只针对新行）
        if "Found in function" in header2col and "Function" in header2col:
            fv = new_row.get("Found in function","") or ""
            for k,v in fund_map.items():
                if k in fv:
                    c=header2col["Function"]
                    ws.cell(last_row,c).value = v
                    ws.cell(last_row,c).fill  = yellow
                    break
        # Root cause
        if "Owner" in header2col and "Root cause" in header2col:
            ow = new_row.get("Owner","") or ""
            for k,v in own_map.items():
                if k in ow:
                    c=header2col["Root cause"]
                    ws.cell(last_row,c).value = v
                    ws.cell(last_row,c).fill  = yellow
                    break

# === 5. 填公式 & 标记新增行 Octane/Jira ===
max_row = ws.max_row

# Days
d_idx  = header2col.get("Days")
ct_idx = header2col.get("Creation time")
if d_idx and ct_idx:
    colL = get_column_letter(ct_idx)
    for r in range(2, max_row+1):
        ws.cell(r,d_idx).value = f'=DATEDIF(${colL}{r},TODAY(),"D")'

# Octane/Jira 列（仅新行）
oij_idx = header2col.get("Octane or Jira")
if oij_idx:
    for r in range(original_last+1, max_row+1):
        ws.cell(r,oij_idx).value = source_key

# Open >20 days
o20 = header2col.get("Open >20 days") or header2col.get("Open > 20 days")
if o20 and d_idx:
    cd = get_column_letter(d_idx)
    for r in range(2, max_row+1):
        ws.cell(r,o20).value = f'=IF({cd}{r}>20,1,0)'

# No TIS
nt = header2col.get("No TIS")
pl = header2col.get("Planned closing version")
ti = header2col.get("Target I-Step:")
if nt and pl and ti:
    pL = get_column_letter(pl)
    tL = get_column_letter(ti)
    for r in range(2, max_row+1):
        ws.cell(r,nt).value = f'=IF(OR({pL}{r}<>"",{tL}{r}<>""),0,1)'

# === 6. 保存 ===
wb.save(upd_fp)
print(f"完成（来源={source_key}）。结果保存在 '{upd_fp}'。")
