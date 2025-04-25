import os
import sys
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# 1. 获取用户输入
directory = input("Please enter the path to the excel folder: ")
timestamp = input("Please enter the timestamp (e.g. 20250421_0905) ")
istep_and_cw = input("Please enter istep and CW  (e.g. 490CW16) ")

# 2. 初始化数据结构
data_dict = {}      # 用于按列累积数据
file_names = []     # 记录每行的来源文件名

# 3. 检查已有汇总文件，避免重复汇总
summary_file_path = os.path.join(directory, f"summary_{timestamp}_{istep_and_cw}.xlsx")
existing_summary = os.path.exists(summary_file_path)

if existing_summary:
    wb = load_workbook(summary_file_path)
    existing_summary_df = pd.read_excel(summary_file_path, sheet_name='TRIGGER')
    existing_file_names = existing_summary_df['File Name'].unique().tolist()
else:
    existing_file_names = []

# 4. 遍历文件夹，读取每个 .xlsm 的 TRIGGER 表
for filename in os.listdir(directory):
    if not filename.endswith(".xlsm"):
        continue
    if filename.startswith("summary"):
        continue
    if filename in existing_file_names:
        continue

    file_path = os.path.join(directory, filename)
    xls = pd.ExcelFile(file_path)
    if 'TRIGGER' not in xls.sheet_names:
        print(f"Ignored (missing TRIGGER sheet): {filename}")
        continue

    # 4.1 读取数据并清洗
    df = pd.read_excel(file_path, sheet_name='TRIGGER', skiprows=6)
    # 删除跳过行后首个多余行
    df = df.drop(df.index[0])
    # 删除第 34 列及以后的冗余列
    cols_to_delete = df.columns[33:200]
    df = df.drop(columns=cols_to_delete, errors='ignore')
    # 丢弃全空行
    df.dropna(how='all', inplace=True)

    # 4.2 记录来源文件名
    appended_filename = f"{filename}_{timestamp}_{istep_and_cw}"
    file_names.extend([appended_filename] * len(df))

    # 4.3 累积各列数据
    for col in df.columns:
        data_dict.setdefault(col, []).extend(df[col].tolist())

# 5. 合并所有数据为 DataFrame
try:
    new_data_df = pd.DataFrame(data_dict)
except Exception as e:
    print(f"\nError when creating DataFrame: {e}")
    print("请检查各 Excel 的第 7 行表头是否一致。")
    sys.exit(1)

# 插入“File Name”列
new_data_df.insert(0, 'File Name', file_names)

# 格式化“Date”列
if 'Date' in new_data_df.columns:
    new_data_df['Date'] = pd.to_datetime(new_data_df['Date'], errors='coerce').dt.strftime('%b %d')

# 添加空的“Comment”列，并把“File Name”移到最后
new_data_df['Comment'] = ''
cols = new_data_df.columns.tolist()
cols.append(cols.pop(cols.index('File Name')))
new_data_df = new_data_df[cols]

# 再次丢弃可能的全空行
new_data_df.dropna(how='all', inplace=True)

# 6. 如果已有旧汇总，则拼接
if existing_summary:
    summary_df = pd.concat([existing_summary_df, new_data_df], ignore_index=True)
else:
    summary_df = new_data_df

# 最后一次丢弃空行，并打印行数
summary_df.dropna(how='all', inplace=True)
print(f"最终行数（不含全空行）：{summary_df.shape[0]}")

# 7. 保存为 Excel
summary_df.to_excel(summary_file_path, index=False, engine='openpyxl', sheet_name='TRIGGER')

# 8. 加载写好的工作簿，开始格式化
wb = load_workbook(summary_file_path)
ws_trigger = wb['TRIGGER']

# 8.1 添加筛选
ws_trigger.auto_filter.ref = ws_trigger.dimensions

# 8.2 创建/更新 OVERVIEW 表
if 'OVERVIEW' in wb.sheetnames:
    ws_overview = wb['OVERVIEW']
    ws_overview.delete_rows(2, ws_overview.max_row - 1)
else:
    ws_overview = wb.create_sheet(title='OVERVIEW')

status_counts = summary_df['Status'].value_counts()
ws_overview.append(['Status', 'Count'])
for status, cnt in status_counts.items():
    ws_overview.append([status, cnt])

# 8.3 插入柱状图
chart = BarChart()
chart.title = "Status Counts"
chart.x_axis.title = "Status"
chart.y_axis.title = "Count"
data = Reference(ws_overview, min_col=2, min_row=1, max_row=len(status_counts)+1)
cats = Reference(ws_overview, min_col=1, min_row=2, max_row=len(status_counts)+1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws_overview.add_chart(chart, "E5")

# 8.4 生成 In Analysis/Done 列表
excluded = ["done", "OCT-Ticket", "known", "Event", "ToDo done"]
in_analysis, done = [], []
for fn in summary_df['File Name'].unique():
    sts = summary_df[summary_df['File Name']==fn]['Status']
    if all((s in excluded) or pd.isna(s) for s in sts):
        done.append(fn)
    else:
        in_analysis.append(fn)

start_row = 20
ws_overview.cell(row=start_row, column=1, value='In Analysis')
for i, fn in enumerate(in_analysis, start=start_row+1):
    ws_overview.cell(row=i, column=1, value=fn)
done_row = start_row + len(in_analysis) + 2
ws_overview.cell(row=done_row, column=1, value='Done')
for i, fn in enumerate(done, start=done_row+1):
    ws_overview.cell(row=i, column=1, value=fn)

# 8.5 隐藏和调整列
hide_cols = ['#', 'Excel-Session', 'GPS Position', 'Road_Ext_QU', 'Software',
             'Unnamed: 11','Unnamed: 12','Unnamed: 14','Unnamed: 16',
             'Canape-Folder','BI','Error occurence','TIS','Category','Cluster',
             'CANape device 1','CANape device 2']
for col in hide_cols:
    if col in summary_df.columns:
        idx = summary_df.columns.get_loc(col) + 1
        ws_trigger.column_dimensions[get_column_letter(idx)].hidden = True

ws_trigger.sheet_view.zoomScale = 80

# 自动调整宽度（限最大 30）
def auto_adjust(ws, df, cols):
    for c in cols:
        if c in df.columns:
            idx = df.columns.get_loc(c) + 1
            length = max(df[c].astype(str).map(len).max(), len(c)) + 2
            ws.column_dimensions[get_column_letter(idx)].width = min(length, 30)

auto_adjust(ws_trigger, summary_df, ['File Name','Date'] + summary_df.columns.tolist())

# 特定列宽调整
for col in ['Event','Analysis','Solution']:
    if col in summary_df.columns:
        ws_trigger.column_dimensions[get_column_letter(summary_df.columns.get_loc(col)+1)].width = 50
for col in ['Canape-Trigger','Vigem-Trigger','Datacenter','Comment']:
    if col in summary_df.columns:
        ws_trigger.column_dimensions[get_column_letter(summary_df.columns.get_loc(col)+1)].width = 20

# 首行灰色背景
grey_fill = PatternFill("solid", fgColor="D3D3D3")
for cell in ws_trigger[1]:
    cell.fill = grey_fill

# 8.6 下拉及条件格式化
status_opts = {
    "open":"FFCCCC","forward":"FFCCCC","in analysis":"FFFFCC",
    "done":"CCFFCC","OCT-Ticket create":"FFFFCC","OCT-Ticket":"CCFFCC",
    "Evaluation open":"FFFFCC","known":"CCFFCC","new trigger":"FFCCCC",
    "see":"E0E0E0","see Cluster":"E0E0E0","Event":"CCCCFF",
    "ToDo":"FFCCCC","ToDo done":"CCFFCC"
}
status_list = ",".join(status_opts.keys())
dv = DataValidation(type="list", formula1=f'"{status_list}"', allow_blank=True)
ws_trigger.add_data_validation(dv)
col_idx = summary_df.columns.get_loc('Status') + 1
for r in range(2, ws_trigger.max_row+1):
    dv.add(ws_trigger.cell(row=r, column=col_idx))

for st, color in status_opts.items():
    fill = PatternFill("solid", fgColor=color)
    rule = CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill)
    addr = f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws_trigger.max_row}"
    ws_trigger.conditional_formatting.add(addr, rule)

# 9. 保存并结束
wb.save(summary_file_path)
print(f"Summary Excel file created successfully at {summary_file_path}!")
