import os
import time
import json
import pandas as pd
import schedule
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# === 加载配置 ===
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
with open(os.path.join(BASE_DIR, "unified_config_auto.json"), 'r', encoding='utf-8') as f:
    cfg = json.load(f)

folders    = cfg['folders']
ORIG_DIR   = os.path.join(BASE_DIR, folders['orig_dir'])
JIRA_DIR   = os.path.join(BASE_DIR, folders['jira_dir'])
OCTANE_DIR = os.path.join(BASE_DIR, folders['octane_dir'])

paths      = cfg['paths']
sheet      = cfg['sheet']['target_sheet']
sources    = cfg['sources']
settings   = cfg.get('settings', {})
clear_old  = settings.get('clear_old_highlight','N').upper() == 'Y'
fund_pats  = cfg.get('fund_function_patterns', {})
owner_pats = cfg.get('owner_root_cause_patterns', {})

def is_row_blank(ws, row):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True

def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break

def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            return r
    return 1

def update_excel(new_fp):
    print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] Start update: {new_fp}")
    folder = os.path.basename(os.path.dirname(new_fp))
    if folder == os.path.basename(JIRA_DIR):
        source_key = "Jira"
    elif folder == os.path.basename(OCTANE_DIR):
        source_key = "Octane"
    else:
        print(f" → Unknown folder ({folder}), skip.")
        return

    # 原表 与 生成文件路径
    orig_fp = os.path.join(ORIG_DIR, paths['original_file'])
    upd_fp  = os.path.join(os.path.dirname(new_fp), paths['updated_file'])

    src_cfg   = sources[source_key]
    read_meth = src_cfg['read_method']
    date_col  = src_cfg.get('date_col')
    mapping   = src_cfg['mapping']

    # 读取 df_new
    df_new = pd.read_excel(new_fp) if read_meth=="excel" else pd.read_csv(new_fp)

    # 打开原表
    wb = load_workbook(orig_fp)
    ws = wb[sheet]

    # 清除旧高亮
    if clear_old:
        green = {"8ED973","008ED973","FF8ED973"}
        blue  = {'FFADD8E6','00ADD8E6','ADD8E6'}
        gray  = {'FFC0C0C0','00C0C0C0','C0C0C0'}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                pf = cell.fill
                if getattr(pf,'patternType',None)=="solid":
                    argb = pf.fgColor.rgb or pf.fgColor.value
                    if argb in green|blue|gray:
                        cell.fill = PatternFill()

    trim_trailing_blank_rows(ws)

    header2col = { ws.cell(1,c).value: c for c in range(1, ws.max_column+1) }
    id_col     = header2col['ID']
    headers    = list(header2col.keys())

    # 已有ID→行
    id2row = {
        ws.cell(r,id_col).value: r
        for r in range(2, ws.max_row+1)
        if ws.cell(r,id_col).value
    }

    # 如果是 Excel，提取超链接
    id2url_new = {}
    if read_meth=="excel":
        tmpwb = load_workbook(new_fp)
        nws   = tmpwb.active
        cid   = next((c for c in range(1,nws.max_column+1)
                      if nws.cell(1,c).value=="ID"), None)
        if cid:
            for r in range(2, nws.max_row+1):
                cell = nws.cell(r,cid)
                if cell.hyperlink:
                    id2url_new[cell.value] = cell.hyperlink.target

    last_row      = find_last_data_row(ws, id_col)
    original_last = last_row

    update_fill = PatternFill("solid", fgColor="ADD8E6")
    new_fill    = PatternFill("solid", fgColor="8ED973")
    gray_fill   = PatternFill("solid", fgColor="C0C0C0")

    id_key = next(k for k,v in mapping.items() if v=="ID")

    # 4. 更新 or 追加
    for _, new_row in df_new.iterrows():
        new_id = new_row.get(id_key)
        if pd.isna(new_id) or not new_id:
            continue

        if new_id in id2row:
            r = id2row[new_id]
            phase_val = ws.cell(r,header2col["Phase"]).value or ""
            if any(k in phase_val for k in ("Concluded","Closed","Resolved")):
                continue

            #—更新日期—
            if date_col:
                raw = new_row.get(date_col)
                if pd.notna(raw):
                    try:
                        dt = pd.to_datetime(raw).to_pydatetime()
                        c  = header2col[mapping[date_col]]
                        cell = ws.cell(r,c)
                        if cell.value != dt:
                            cell.value         = dt
                            cell.number_format = "m/d/yyyy h:mm:ss AM/PM"
                            cell.fill          = update_fill
                    except: pass

            #—更新其他字段—
            for nk, ok in mapping.items():
                if nk==date_col or ok=="Function": continue
                val = new_row.get(nk)
                if pd.isna(val) or val=="": continue
                if ok=="Involved I-Step" and str(val).startswith(("G070","U006")):
                    val = "NA05"+str(val)[4:]
                c = header2col.get(ok)
                if c and ws.cell(r,c).value != val:
                    ws.cell(r,c).value = val
                    ws.cell(r,c).fill  = update_fill

            #—Root cause—
            if "Owner" in header2col and "Root cause" in header2col:
                ow = str(ws.cell(r,header2col["Owner"]).value or "").lower()
                for cause,names in owner_pats.items():
                    if any(n.lower() in ow for n in names):
                        cell = ws.cell(r,header2col["Root cause"])
                        if cell.value != cause:
                            cell.value = cause
                            cell.fill  = update_fill
                        break

        else:
            last_row += 1
            for idx,hdr in enumerate(headers, start=1):
                val = ""
                for nk, ok in mapping.items():
                    if ok==hdr:
                        tmp = new_row.get(nk)
                        if pd.notna(tmp): val = tmp
                        break
                c = ws.cell(last_row, idx)
                c.value = val
                if val:
                    c.fill = new_fill
                    if hdr == mapping.get(date_col):
                        c.number_format = "m/d/yyyy h:mm:ss AM/PM"

            # 超链接
            if new_id in id2url_new:
                hc = ws.cell(last_row, id_col)
                hc.hyperlink, hc.style = id2url_new[new_id], "Hyperlink"

            # Function/Root cause 如上面逻辑

    # 5. 填公式等（同你脚本）
    max_row = ws.max_row
    # Days
    if "Days" in header2col and "Creation time" in header2col:
        colL = get_column_letter(header2col["Creation time"])
        for r in range(2, max_row+1):
            ws.cell(r,header2col["Days"]).value = f'=DATEDIF(${colL}{r},TODAY(),"D")'

    # 保存
    wb.save(upd_fp)
    print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] 完成，保存至 {upd_fp}")

class FolderHandler(FileSystemEventHandler):
    def on_created(self, event):
        if not event.is_directory:
            folder = os.path.basename(os.path.dirname(event.src_path))
            if folder in (folders['jira_dir'], folders['octane_dir']):
                update_excel(event.src_path)
    on_modified = on_created

def main():
    for d in (ORIG_DIR, JIRA_DIR, OCTANE_DIR):
        os.makedirs(d, exist_ok=True)

    observer = Observer()
    observer.schedule(FolderHandler(), path=JIRA_DIR, recursive=False)
    observer.schedule(FolderHandler(), path=OCTANE_DIR, recursive=False)
    observer.start()
    print(f"监控：{JIRA_DIR}, {OCTANE_DIR}")

    schedule.every().day.at("08:00").do(lambda:
        [update_excel(os.path.join(JIRA_DIR,fn)) for fn in os.listdir(JIRA_DIR) if fn.lower().endswith((".xlsx",".csv"))] +
        [update_excel(os.path.join(OCTANE_DIR,fn)) for fn in os.listdir(OCTANE_DIR) if fn.lower().endswith((".xlsx",".csv"))]
    )
    print("已设定每日 08:00 全量扫描更新。")

    try:
        while True:
            schedule.run_pending()
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

if __name__ == "__main__":
    main()



from watchdog.events import FileSystemEventHandler

class FolderHandler(FileSystemEventHandler):
    def __init__(self, folders):
        self.folders = folders

    def on_created(self, event):
        self._maybe_update(event.src_path)

    on_modified = on_created

    def on_moved(self, event):
        self._maybe_update(event.dest_path)

    def _maybe_update(self, path):
        if os.path.isdir(path):
            return
        fld = os.path.basename(os.path.dirname(path))
        if fld in (self.folders['jira_dir'], self.folders['octane_dir']):
            update_excel(path)



from watchdog.observers.polling import PollingObserver as Observer

def main():
    …  
    observer = Observer()
    handler  = FolderHandler(folders)
    observer.schedule(handler, path=JIRA_DIR, recursive=False)
    observer.schedule(handler, path=OCTANE_DIR, recursive=False)
    …

