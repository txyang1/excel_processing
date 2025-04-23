import os
import time
import json
import pandas as pd
import schedule
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# === 读取配置 ===
BASE_DIR  = os.path.abspath(os.path.dirname(__file__))
with open(os.path.join(BASE_DIR, "unified_config_auto.json"), 'r', encoding='utf-8') as f:
    cfg = json.load(f)

# 文件夹配置
folders   = cfg['folders']
ORIG_DIR   = os.path.join(BASE_DIR, folders['orig_dir'])
JIRA_DIR   = os.path.join(BASE_DIR, folders['jira_dir'])
OCTANE_DIR = os.path.join(BASE_DIR, folders['octane_dir'])

# JSON 中的 paths 和其他设置
paths     = cfg['paths']
sheet     = cfg['sheet']['target_sheet']
sources   = cfg['sources']
settings  = cfg.get('settings', {})
clear_old = settings.get('clear_old_highlight','N').upper() == 'Y'
fund_pats = cfg.get('fund_function_patterns', {})
owner_pats= cfg.get('owner_root_cause_patterns', {})

def is_row_blank(ws, row):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row, c).value not in (None, ""):
            return False
    return True

def trim_trailing_blank_rows(ws):
    for r in range(ws.max_row, 1, -1):
        if is_row_blank(ws, r):
            ws.delete_rows(r)
        else:
            break

def find_last_data_row(ws, key_col):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            return r
    return 1

def update_excel(new_fp):
    """针对单个新文件 new_fp，读取 Orig_files/summary.xlsx，更新后保存到同一文件夹下的 updated_file 名称。"""
    print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] 开始更新: {new_fp}")
    folder = os.path.basename(os.path.dirname(new_fp))
    source_key = None
    if folder == os.path.basename(JIRA_DIR):   source_key = "Jira"
    if folder == os.path.basename(OCTANE_DIR): source_key = "Octane"
    if not source_key:
        print(f"  → 未识别来源 ({folder})，跳过。")
        return

    # 加载原表与目标路径
    orig_fp = os.path.join(ORIG_DIR, paths['original_file'])
    upd_name= paths['updated_file']
    dest_dir= os.path.dirname(new_fp)
    upd_fp  = os.path.join(dest_dir, upd_name)

    src_cfg   = sources[source_key]
    read_meth = src_cfg['read_method']
    date_col  = src_cfg.get('date_col')
    mapping   = src_cfg['mapping']

    # 读取新表
    df_new = pd.read_excel(new_fp) if read_meth=="excel" else pd.read_csv(new_fp)

    # 打开原表
    wb = load_workbook(orig_fp)
    ws = wb[sheet]

    # 清除旧高亮
    if clear_old:
        green_codes = {"8ED973","008ED973","FF8ED973"}
        blue_codes = {'FFADD8E6','00ADD8E6','ADD8E6'}
        gray_codes = {'FFC0C0C0','00C0C0C0','C0C0C0'}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                pf = cell.fill
                if getattr(pf,'patternType',None)=="solid":
                    argb = pf.fgColor.rgb or pf.fgColor.value
                    if argb in green_codes|blue_codes|gray_codes:
                        cell.fill = PatternFill()

    trim_trailing_blank_rows(ws)

    # 表头与已有 ID→行
    header2col = {ws.cell(1,c).value:c for c in range(1,ws.max_column+1)}
    headers    = list(header2col.keys())
    id_col     = header2col['ID']

    id2row     = { ws.cell(r,id_col).value:r
                   for r in range(2, ws.max_row+1)
                   if ws.cell(r,id_col).value }

    # 对 Excel 来源，提取新表 ID→超链接
id2url_new = {}
if read_meth == "excel":
    nwb    = load_workbook(new_fp)
    nws    = nwb.active
    col_id = next((c for c in range(1, nws.max_column+1)
                   if nws.cell(1,c).value=="ID"), None)
    if col_id:
        for r in range(2, nws.max_row+1):
            cell = nws.cell(r, col_id)
            if cell.hyperlink:
                id2url_new[cell.value] = cell.hyperlink.target

# 定位追加起始行
last_row      = find_last_data_row(ws, id_col)
original_last = last_row

# 高亮样式
update_fill = PatternFill("solid", fgColor="ADD8E6")   # 淡蓝
new_fill    = PatternFill("solid", fgColor="8ED973")   # 淡绿色
gray_fill   = PatternFill("solid", fgColor="C0C0C0")   # 灰色

# 新表中 ID 列对应的键
id_key = next(k for k,v in mapping.items() if v=="ID")

# === 4. 遍历 df_new，更新 or 追加 ===
for _, new_row in df_new.iterrows():
    new_id = new_row.get(id_key, "")
    if pd.isna(new_id) or not new_id:
        continue

    if new_id in id2row:
        # —— 已有行，更新除 Function 之外字段 并染淡蓝 —— 
        r = id2row[new_id]

        # 跳过已结束行
        phase_val = ws.cell(r, header2col["Phase"]).value or ""
        if any(p in phase_val for p in ("Concluded","Closed","Resolved")):
            continue

        # 更新日期 (Jira)
        if date_col:
            raw = new_row.get(date_col, "")
            if pd.notna(raw) and raw != "":
                try:
                    ts = pd.to_datetime(raw)
                    dt = ts.to_pydatetime()
                    c  = header2col[mapping[date_col]]
                    cell = ws.cell(r,c)
                    if cell.value != dt:
                        cell.value         = dt
                        cell.number_format = "m/d/yyyy h:mm:ss AM/PM"
                        cell.fill          = update_fill
                except:
                    pass

        # 更新其他字段 (跳过 Function)
        for nk, ok in mapping.items():
            if nk==date_col or ok=="Function":
                continue
            val = new_row.get(nk, "")
            if pd.isna(val) or val=="":
                continue
            if ok=="Involved I-Step" and (str(val).startswith("G070") or str(val).startswith("U006")):
                val = "NA05" + str(val)[4:]
            c = header2col.get(ok)
            if c:
                cell = ws.cell(r,c)
                if cell.value != val:
                    cell.value = val
                    cell.fill  = update_fill

        '''# 更新 Function via patterns
        if "Found in function" in header2col and "Function" in header2col:
            raw_fv = ws.cell(r, header2col["Found in function"]).value or ""
            fv = str(raw_fv).lower()
            for func_name, kws in fund_patterns.items():
                if any(kw.lower() in fv for kw in kws):
                    cell = ws.cell(r, header2col["Function"])
                    if cell.value != func_name:
                        cell.value = func_name
                        cell.fill  = update_fill
                    break'''

        # 更新 Root cause via patterns
        if "Owner" in header2col and "Root cause" in header2col:
            raw_ow = ws.cell(r, header2col["Owner"]).value or ""
            ow = str(raw_ow).lower()
            for cause, names in owner_patterns.items():
                if any(n.lower() in ow for n in names):
                    cell = ws.cell(r, header2col["Root cause"])
                    if cell.value != cause:
                        cell.value = cause
                        cell.fill  = update_fill
                    break

    else:
        # —— 新增行，填值并染粉色 —— 
        last_row += 1
        for idx, hdr in enumerate(headers, start=1):
            val = ""
            for nk, ok in mapping.items():
                if ok == hdr:
                    tmp = new_row.get(nk, "")
                    if pd.notna(tmp) and tmp != "":
                        if hdr == mapping.get(date_col):
                            try:
                                ts = pd.to_datetime(tmp)
                                val = ts.to_pydatetime()
                            except:
                                val = tmp
                        else:
                            val = tmp
                        if ok=="Involved I-Step:" and (str(val).startswith("G070") or str(val).startswith("U006")):
                            val = "NA05" + str(val)[4:]
                    break
            cell = ws.cell(last_row, idx)
            cell.value = val
            if val != "":
                cell.fill = new_fill
                if hdr == mapping.get(date_col):
                    cell.number_format = "m/d/yyyy h:mm:ss AM/PM"

        # 新增 ID 超链接
        url = id2url_new.get(new_id)
        if url:
            cell = ws.cell(last_row, id_col)
            cell.hyperlink = url
            cell.style     = "Hyperlink"

        # 新增 Function via patterns
        if "Found in function" in header2col and "Function" in header2col:
            raw_fv = ws.cell(last_row, header2col["Found in function"]).value or ""
            fv = str(raw_fv).lower()
            for func_name, kws in fund_patterns.items():
                if any(kw.lower() in fv for kw in kws):
                    cell = ws.cell(last_row, header2col["Function"])
                    cell.value = func_name
                    cell.fill  = new_fill
                    break

        # 新增 Root cause via patterns
        if "Owner" in header2col and "Root cause" in header2col:
            raw_ow = ws.cell(last_row, header2col["Owner"]).value or ""
            ow = str(raw_ow).lower()
            for cause, names in owner_patterns.items():
                if any(n.lower() in ow for n in names):
                    cell = ws.cell(last_row, header2col["Root cause"])
                    cell.value = cause
                    cell.fill  = new_fill
                    break

# === 5. 填充公式 & 标记 Octane/Jira & 其它列 ===
max_row = ws.max_row

# 5.1 Days 列公式
d_idx  = header2col.get("Days")
ct_idx = header2col.get("Creation time")
if d_idx and ct_idx:
    colL = get_column_letter(ct_idx)
    for r in range(2, max_row + 1):
        ws.cell(r, d_idx).value = f'=DATEDIF(${colL}{r},TODAY(),"D")'

# 5.2 Octane or Jira 列（仅新增）
oij_idx = header2col.get("Octane or Jira")
if oij_idx:
    for r in range(original_last + 1, max_row + 1):
        ws.cell(r, oij_idx).value = source_key

# 5.3 Open > 20 days 列
o20 = header2col.get("Open >20 days") or header2col.get("Open > 20 days")
if o20 and d_idx:
    dl = get_column_letter(d_idx)
    for r in range(2, max_row + 1):
        ws.cell(r, o20).value = f'=IF({dl}{r}>20,1,0)'

# 5.4 No TIS 列
nt = header2col.get("No TIS")
pl = header2col.get("Planned closing version")
ti = header2col.get("Target I-Step:")
if nt and pl and ti:
    pL = get_column_letter(pl)
    tL = get_column_letter(ti)
    for r in range(2, max_row + 1):
        ws.cell(r, nt).value = f'=IF(OR({pL}{r}<>"",{tL}{r}<>""),0,1)'

# 5.5 Top issue 列
tags_idx = header2col.get("Tags")
top_idx =  header2col.get("Top issue Candidiate")
gray_fill = PatternFill("solid", fgColor="C0C0C0")

if tags_idx and top_idx:
    for r in range(2,max_row+1):
        tags_val = ws.cell(r, tags_idx).value or ""
        top_cell = ws.cell(r, top_idx)
        top_val = top_cell.value or ""
        if "IPN_CN_TopIssue" in tags_val:
            if top_val == "":
                top_cell.value = "Yes"
                top_cell.fill = new_fill
        else:
            if top_val == "Yes":
                top_cell.fill = gray_fill
# === 6. 保存 ===
wb.save(upd_fp)
print(f"更新完成 (来源={source_key})，新增行绿色，高亮；更新行淡蓝，高亮。结果已保存至 '{upd_fp}'.")               

class FolderHandler(FileSystemEventHandler):
    def on_created(self, event):
        if not event.is_directory:
            folder = os.path.basename(os.path.dirname(event.src_path))
            if folder in (folders['jira_dir'], folders['octane_dir']):
                update_excel(event.src_path)
    on_modified = on_created

def main():
    # 确保目录存在
    for d in (ORIG_DIR, JIRA_DIR, OCTANE_DIR):
        os.makedirs(d, exist_ok=True)

    # 启动 Watchdog
    observer = Observer()
    observer.schedule(FolderHandler(), path=JIRA_DIR, recursive=False)
    observer.schedule(FolderHandler(), path=OCTANE_DIR, recursive=False)
    observer.start()
    print(f"监控目录：{JIRA_DIR}，{OCTANE_DIR}")

    # 每日 08:00 全量扫描
    schedule.every().day.at("08:00").do(lambda: [
        update_excel(os.path.join(JIRA_DIR, fn)) for fn in os.listdir(JIRA_DIR)
        if fn.lower().endswith((".xlsx",".csv"))
    ] + [
        update_excel(os.path.join(OCTANE_DIR, fn)) for fn in os.listdir(OCTANE_DIR)
        if fn.lower().endswith((".xlsx",".csv"))
    ])
    print("已添加每日 08:00 全量扫描更新任务。")

    try:
        while True:
            schedule.run_pending()
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

if __name__ == "__main__":
    main()
