import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============================
# 1. 参数设置
# ============================
original_file = "original.xlsx"          # 原始 Excel 文件（包含多个 sheet）
jira_file = "jira.csv"                   # Jira 的 CSV 文件
target_sheet = "Octane and jira"         # 目标工作表名称
updated_file = "updated_excel.xlsx"      # 更新后保存的文件

# ============================
# 2. 读取原始 Excel 中目标工作表数据
# ============================
df_original = pd.read_excel(original_file, sheet_name=target_sheet)

# ============================
# 3. 读取 Jira CSV 文件，并映射字段名称
# ============================
df_jira = pd.read_csv(jira_file)

# 定义映射关系：将 Jira 中的字段名称转换为原表对应的字段名称
rename_dict = {
    "Summary": "Name",
    "issue key": "Ticket no. supplier",
    "Assignee": "Owner",
    "Reporter": "Defect finder"
}
df_jira.rename(columns=rename_dict, inplace=True)

# ============================
# 4. 构造与原表结构一致的新数据
# ============================
# 原表的列作为目标结构
target_columns = df_original.columns.tolist()

# 提取 Jira 中与原表中字段相同的部分
common_columns = [col for col in target_columns if col in df_jira.columns]
df_new_subset = df_jira[common_columns].copy()

# 如果原表中有的字段 Jira 中未提供，则补充空值
for col in target_columns:
    if col not in df_new_subset.columns:
        df_new_subset[col] = ""
        
# 调整列顺序，使其与原表完全一致
df_new_subset = df_new_subset[target_columns]

# ============================
# 5. 将 Jira 中的新数据追加到原表数据的末尾
# ============================
df_updated = pd.concat([df_original, df_new_subset], ignore_index=True)

# ============================
# 6. 将更新后的数据写回 Excel 中，同时保留其它工作表不变
# ============================
# 利用 openpyxl 加载整个工作簿
wb = load_workbook(original_file)

# 删除目标工作表，以便替换为更新后的数据
if target_sheet in wb.sheetnames:
    ws_to_remove = wb[target_sheet]
    wb.remove(ws_to_remove)

# 使用 ExcelWriter 写入更新后的目标 sheet，并保留工作簿中其它 sheet
with pd.ExcelWriter(original_file, engine='openpyxl') as writer:
    writer.book = wb
    df_updated.to_excel(writer, sheet_name=target_sheet, index=False)
    writer.save()

# 为防止覆盖，另存一份更新后的文件
wb.save(updated_file)

# ============================
# 7. 高亮显示新追加的行（以绿色填充）
# ============================
# 重新加载更新后的工作簿和目标 sheet
wb_updated = load_workbook(updated_file)
ws = wb_updated[target_sheet]

# 记录原始数据的行数（不含表头）
original_row_count = df_original.shape[0]
# Excel 中第一行为表头，所以新添加的行起始行号为 original_row_count + 2
highlight_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

for row in range(original_row_count + 2, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).fill = highlight_fill

wb_updated.save(updated_file)
print(f"已将 Jira 数据追加到工作表 '{target_sheet}' 的末尾，并以绿色高亮显示新数据行。更新后的文件为 {updated_file}")
